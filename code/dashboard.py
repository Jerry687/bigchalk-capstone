"""
Big Chalk Mix Engine — Bench UI (Claude Design handoff, option 1a).

Visual system: IBM Plex Sans/Mono, chalk-navy #1E4FA3, left workspace rail,
top slice-picker bar, metric strip with REVIEW flags, family color chips,
health dots on the batch screen. Styling lives in assets/bench.css.

Screens (left rail):
  Diagnostics    — metric strip, actual-vs-fitted with holdout band,
                   residuals, coefficient table with family chips + t-bars.
  Variables      — per-product editable config (sign / adstock / bounds /
                   role); force rows tinted; guardrail banner.
  Contributions  — due-tos by model year, signed avg weekly due-tos,
                   YoY table with % pills.
  Batch          — all models with health dots, review flags, click a row
                   to load that slice; Run all combinations.

    pip install dash
    cd code && python dashboard.py     # http://127.0.0.1:8050
"""
import os
import shutil
import time
import uuid
from functools import lru_cache
from typing import Optional

for _v in ("OMP_NUM_THREADS", "OPENBLAS_NUM_THREADS",
           "MKL_NUM_THREADS", "NUMEXPR_NUM_THREADS"):
    os.environ.setdefault(_v, "2")

import numpy as np
import pandas as pd
import plotly.graph_objects as go
from dash import (Dash, dcc, html, dash_table, Input, Output, State, ctx,
                  no_update)

import capstone_pipeline as cp
import multilevel as ml
import rollup as ru
import saturation_curves as sc

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
DEFAULT_DATA = os.path.join(ROOT, "Anonymized Data for Project.xlsx")
CFG_DIR = os.path.join(ROOT, "configs")
SUMMARY = os.path.join(ROOT, "outputs", "all_models_summary.csv")
os.makedirs(CFG_DIR, exist_ok=True)

WINDOW_CHOICES = [52, 104, 156]
CFG_COLS = ["variable", "family", "sign", "adstock_decay", "scale",
            "sat_midpoint", "sat_slope", "coef_lower", "coef_upper", "role"]
# Families a cap can be set on; drives the per-family limit table (F4).
FAMILIES = ["Distribution", "Trade", "Media", "Competitive", "Macro",
            "Price", "Category Price", "Seasonality", "Trend"]
# Canonical variable-template columns (V2 download/upload; also the export's
# "model fit structure" sheet). channel = a specific channel, or ALL = the
# Product default. Priority on run: Product × Channel override > Product default.
TEMPLATE_COLS = ["product", "channel", "variable", "family", "sign", "role",
                 "coef_lower", "coef_upper", "adstock_decay", "scale",
                 "sat_midpoint", "sat_slope"]
PREFERRED_TARGETS = ["Volume Sales", "Dollar Sales", "Unit Sales"]
FLAG_THRESHOLD = 40.0    # holdout MAPE % above which a model is flagged
VIF_WARN = 10.0          # VIF above which collinearity is flagged (non-blocking)
# Every config-file mutation (save / upload / regenerate / reset / default
# creation) goes through this one lock — shared with capstone_pipeline so app
# writes and default creation can't interleave. Process-local (see cp).
_CFG_WRITE_LOCK = cp._CFG_WRITE_LOCK

# design tokens (assets/bench.css holds the full set)
INK, INK2, INK3, MUTED = "#101828", "#475467", "#667085", "#98A2B3"
NAVY, WARN, RED, GREEN = "#1E4FA3", "#B54708", "#B42318", "#067647"
GRID = "#EEF1F5"
FAM_CHIP = {
    "Distribution": ("#E0F2EC", "#067647"), "Trade": ("#EBF1FB", "#1E4FA3"),
    "Seasonality": ("#FDF2E3", "#B54708"), "Competitive": ("#FBE9E7", "#B42318"),
    "Macro": ("#EEEAF7", "#5925DC"), "Media": ("#E0F0F6", "#0E7090"),
    "Price": ("#FBE9E7", "#B42318"), "Category Price": ("#F2F4F7", "#475467"),
    "Trend": ("#F2F4F7", "#475467"), "(intercept)": ("#F2F4F7", "#475467"),
}
NAV_ICONS = {
    "diag": "M1.5 8.5h3l2-5 3 9 2-6.5h3",
    "vars": "M2 4.5h7 M12 4.5h2 M9.5 2.5v4 M2 11.5h2 M7 11.5h7 M4.5 9.5v4",
    "contrib": "M2.5 3h8 M2.5 8h11 M2.5 13h5",
    "total": "M2 13.5h12 M2 13.5V9h3v4.5 M6.5 13.5V5h3v8.5 M11 13.5V2.5h3v11",
    "sat": "M2 13c3.5 0 4-9 11-9 M2 13h12",
    "multi": "M8 2.5v3 M4 8.5v-1a1 1 0 011-1h6a1 1 0 011 1v1 M4 8.5v4 M8 8.5v4 M12 8.5v4",
    "batch": "M2 2h5v5H2z M9 2h5v5H9z M2 9h5v5H2z M9 9h5v5H9z",
    "export": "M8 2v8 M4.5 6.5l3.5 3.5 3.5-3.5 M2.5 13.5h11",
    "defs": "M3 2.5h7a2 2 0 012 2v9 M13 2.5H6a2 2 0 00-2 2v9 M3 13.5h10",
}
# key "batch" is kept for the component ids; only the display label changed
# to "Model Runs" per Alex (slide 4 nav rename).
# "total", "sat" and "multi" are the 2026-08-11 additions: the High Level
# (Total Brand / Total Channel) rollup, saturation-curve viewing, and the
# pooled / hierarchical model levels.
SCREENS = [("diag", "Diagnostics"), ("vars", "Variables"),
           ("contrib", "Contributions"), ("total", "High Level"),
           ("sat", "Saturation"), ("multi", "Multi-Level"),
           ("batch", "Model Runs"), ("export", "Export"),
           ("defs", "Definitions")]


def _slug(s: str) -> str:
    return "".join(ch for ch in str(s).lower() if ch.isalnum())


@lru_cache(maxsize=24)
def load_sheet(path: str, sheet: str) -> pd.DataFrame:
    return pd.read_excel(path, sheet_name=sheet)


@lru_cache(maxsize=4)
def detect_product_sheets(path: str) -> tuple:
    xl = pd.ExcelFile(path)
    named = tuple(s for s in xl.sheet_names if s.lower().startswith("brand"))
    if named:
        return named
    found = []
    for s in xl.sheet_names:
        try:
            head = pd.read_excel(xl, sheet_name=s, nrows=1)
        except Exception:
            continue
        if {"Geography", "Time"} <= set(map(str, head.columns)):
            found.append(s)
    return tuple(found)


def channels_for(path: str, sheet: str) -> list:
    g = load_sheet(path, sheet)["Geography"].unique()
    return [c for c in g if isinstance(c, str)]


def target_choices(path: str, sheet: str) -> list:
    df = load_sheet(path, sheet)
    num = [c for c in df.columns
           if pd.api.types.is_numeric_dtype(df[c])
           and not str(c).startswith("_C_")]
    pref = [c for c in PREFERRED_TARGETS if c in df.columns]
    # key=str: some client sheets carry numeric column headers, which
    # cannot be sorted against strings directly
    rest = sorted((c for c in num if c not in pref), key=str)
    return pref + rest


# Config-path helpers delegate to capstone_pipeline so the dashboard, the
# batch runner, and the CLI runners all resolve the SAME files (single source
# of truth). resolve_cfg_path falls back to the default path (which the
# dashboard's load_or_create_config always creates) so the editor never gets
# a None path.
def cfg_path(path: str, sheet: str) -> str:
    """Product DEFAULT config (channel = ALL)."""
    return cp.default_config_path(path, sheet)


def override_cfg_path(path: str, sheet: str, channel: str) -> str:
    """Product × Channel OVERRIDE config."""
    return cp.override_config_path(path, sheet, channel)


def resolve_cfg_path(path: str, sheet: str, channel: str) -> str:
    """Runtime resolution: Product × Channel override wins if it exists, else
    the Product default. (Jerry's two-tier model.)"""
    return cp.resolve_config_path(path, sheet, channel) or cfg_path(path, sheet)


def save_cfg_path(path: str, sheet: str, scope: str, channel: str) -> str:
    """Where edits are written: 'channel' scope → this channel's override
    file; otherwise → the Product default."""
    if scope == "channel" and channel:
        return override_cfg_path(path, sheet, channel)
    return cfg_path(path, sheet)


def load_cfg_for_edit(path: str, sheet: str, scope: str, channel: str):
    """Config to show in the editor for the chosen scope.
    Returns (df, override_state) where override_state is:
      None       — editing the Product default
      'existing' — editing an existing channel override
      'inherit'  — channel scope with no override yet (seeded from default)."""
    if scope == "channel" and channel:
        ov = override_cfg_path(path, sheet, channel)
        if os.path.exists(ov):
            return pd.read_csv(ov), "existing"
        return load_or_create_config(path, sheet), "inherit"
    return load_or_create_config(path, sheet), None


# ---------------- variable template (V2 download / upload) ----------------

def _cfg_to_template_rows(cfg_df, product, channel):
    return [{"product": product, "channel": channel, "variable": r["variable"],
             "family": r["family"], "sign": r["sign"], "role": r["role"],
             "coef_lower": r["coef_lower"], "coef_upper": r["coef_upper"],
             "adstock_decay": r["adstock_decay"],
             # older configs predate `scale`; treat a missing one as 1 (no
             # scaling) so existing files keep working untouched
             "scale": r["scale"] if "scale" in cfg_df.columns
             and pd.notna(r.get("scale")) else 1.0,
             # saturation is opt-in; blank = the variable stays linear
             "sat_midpoint": r.get("sat_midpoint")
             if "sat_midpoint" in cfg_df.columns else None,
             "sat_slope": r.get("sat_slope")
             if "sat_slope" in cfg_df.columns else None}
            for _, r in cfg_df.iterrows()]


def build_template_df(path, brand) -> pd.DataFrame:
    """Editable template for a product: the Product default (channel = ALL)
    plus every existing Channel override, in the canonical upload format."""
    rows = _cfg_to_template_rows(load_or_create_config(path, brand), brand, "ALL")
    for channel in channels_for(path, brand):
        ov = override_cfg_path(path, brand, channel)
        if os.path.exists(ov):
            rows += _cfg_to_template_rows(pd.read_csv(ov), brand, channel)
    return pd.DataFrame(rows, columns=TEMPLATE_COLS)


def _norm_channel(channel) -> str:
    if channel is None or (isinstance(channel, float) and np.isnan(channel)):
        return "ALL"
    c = str(channel).strip()
    return "ALL" if c == "" or c.upper() == "ALL" else c


def _tmpl_pick(t, col, default):
    return t[col] if col in t and pd.notna(t[col]) else default


def apply_template_df(tdf, path):
    """Apply an uploaded template ATOMICALLY: validate every (product, channel)
    group first and write NOTHING unless all pass, so a later failure can't
    leave a partial edit. Each group → a config file (channel ALL → Product
    default, else the Channel override); variables in the data NOT listed for a
    group become role=exclude (Alex). Returns (status, [(product, channel,
    detail), …]) where status is 'applied' / 'rejected' / 'rollback_incomplete'
    ('rollback_incomplete' means disk changed but couldn't be fully restored —
    the caller must refresh AND surface the recovery path)."""
    tdf = tdf.copy()
    tdf.columns = [str(c).strip().lower() for c in tdf.columns]
    if not {"product", "channel", "variable"} <= set(tdf.columns):
        return "rejected", [("—", "—", "missing required columns (need product, "
                             "channel, variable)")]

    sheets = set(detect_product_sheets(path))
    # Normalize product / variable / channel BEFORE grouping so that "ALL"/
    # "all", channel case, and stray whitespace collapse into one group instead
    # of forming duplicate groups that overwrite the same file. Channel is
    # canonicalized to the product's actual channel name (its override path is
    # case-folded, so two spellings would otherwise collide).
    _chan_map = {}

    def canon_channel(product, ch):
        ch = _norm_channel(ch)
        if ch == "ALL" or product not in sheets:
            return ch
        if product not in _chan_map:
            _chan_map[product] = {str(c).strip().lower(): c
                                  for c in channels_for(path, product)}
        return _chan_map[product].get(ch.lower(), ch)

    tdf["product"] = tdf["product"].map(lambda x: str(x).strip())
    tdf["variable"] = tdf["variable"].map(lambda x: str(x).strip())
    tdf["channel"] = [canon_channel(p, c)
                      for p, c in zip(tdf["product"], tdf["channel"])]

    plans, errors = [], []          # plans: (product, ch, target_path, cfg_df)
    for (product, ch), grp in tdf.groupby(["product", "channel"], dropna=False):
        product, ch = str(product), str(ch)
        tag = (product, ch)
        if product not in sheets:
            errors.append((*tag, f"unknown product '{product}'"))
            continue
        if ch != "ALL" and ch not in set(channels_for(path, product)):
            errors.append((*tag, f"unknown channel '{ch}' for {product}"))
            continue
        base = cp.generate_default_config(load_sheet(path, product))
        universe = {str(v) for v in base["variable"]}
        listed_names = list(grp["variable"])
        dups = sorted({v for v in listed_names if listed_names.count(v) > 1})
        if dups:
            errors.append((*tag, f"duplicate variables: {dups}"))
            continue
        unknown = [v for v in listed_names if v not in universe]
        if unknown:      # typos -> reject (never silently exclude everything)
            errors.append((*tag, f"unknown variable(s): {unknown[:5]}"
                           + (" …" if len(unknown) > 5 else "")))
            continue
        listed = {str(r["variable"]): r for _, r in grp.iterrows()}
        rows = []
        for _, b in base.iterrows():
            v = str(b["variable"])
            if v in listed:
                t = listed[v]
                rows.append({"variable": v,
                             "family": _tmpl_pick(t, "family", b["family"]),
                             "sign": _tmpl_pick(t, "sign", b["sign"]),
                             "adstock_decay": _tmpl_pick(t, "adstock_decay",
                                                         b["adstock_decay"]),
                             "coef_lower": _tmpl_pick(t, "coef_lower",
                                                      b["coef_lower"]),
                             "coef_upper": _tmpl_pick(t, "coef_upper",
                                                      b["coef_upper"]),
                             "scale": _tmpl_pick(t, "scale",
                                                 b.get("scale", 1.0)),
                             "sat_midpoint": _tmpl_pick(
                                 t, "sat_midpoint", b.get("sat_midpoint")),
                             "sat_slope": _tmpl_pick(
                                 t, "sat_slope", b.get("sat_slope")),
                             "role": _tmpl_pick(t, "role", b["role"])})
            else:                                   # unlisted -> excluded
                row = {k: b.get(k, 1.0 if k == "scale" else None)
                       for k in CFG_COLS}
                row["role"] = "exclude"
                rows.append(row)
        ok, cfg, bad = _validate_cfg_rows(rows)       # validate, don't write
        if not ok:
            errors.append((*tag, f"bound lo >= hi for {bad}"))
            continue
        target = (cfg_path(path, product) if ch == "ALL"
                  else override_cfg_path(path, product, ch))
        plans.append((product, ch, target, cfg))

    if errors:                       # atomic: reject the whole upload
        skipped = [(p, c, "not written (upload rejected)") for p, c, _, _ in plans]
        return "rejected", errors + skipped
    if not plans:                    # empty / no-op upload writes nothing
        return "rejected", [("—", "—", "no valid rows to write (empty template)")]

    return _commit_configs_atomic(plans)


def _rm(paths):
    for p in paths:
        try:
            os.remove(p)
        except OSError:
            pass


def _commit_configs_atomic(plans):
    """Write configs as a best-effort multi-file transaction, serialized by
    _CFG_WRITE_LOCK so concurrent uploads/saves can't interleave on a target.
      1. Stage every config to a UNIQUE temp file. Any staging failure (I/O OR
         serialization) → delete all temps (incl. the partial one, tracked
         before the write) and commit nothing.
      2. Commit: back up each existing target, then os.replace(tmp → target).
         If any commit step fails, ROLL BACK — restore backed-up originals,
         remove newly-created targets. A restore that ITSELF fails is reported
         and its backup is PRESERVED (never delete the last copy of an original)
         with an explicit 'rollback incomplete' status.
    os.replace is atomic per file; cross-file atomicity isn't OS-guaranteed, so
    this backup/rollback keeps the set consistent (single process)."""
    with _CFG_WRITE_LOCK:
        staged = []   # dicts: tmp, target, existed, cfg, product, ch, committed
        try:
            for product, ch, target, cfg in plans:
                tmp = f"{target}.{uuid.uuid4().hex}.uploadtmp"
                staged.append({"tmp": tmp, "target": target,
                               "existed": os.path.exists(target), "cfg": cfg,
                               "product": product, "ch": ch, "committed": False})
                cfg.to_csv(tmp, index=False)       # tmp tracked before write
        except Exception as e:                     # I/O or serialization
            _rm([s["tmp"] for s in staged])
            return "rejected", [("—", "—",
                                 f"write failed — nothing committed: {e}")]

        backups = {}          # target -> backup path
        made_backups = []     # every bak path attempted (for cleanup)
        try:
            for s in staged:
                if s["existed"]:
                    bak = f"{s['target']}.{uuid.uuid4().hex}.bak"
                    made_backups.append(bak)       # tracked before copy
                    shutil.copy2(s["target"], bak)
                    backups[s["target"]] = bak
                os.replace(s["tmp"], s["target"])
                s["committed"] = True
        except Exception as e:
            rollback_errors, keep, failed = [], set(), set()
            for s in staged:
                if not s["committed"]:
                    continue
                tgt, bak = s["target"], backups.get(s["target"])
                if s["existed"] and bak:           # restore original
                    try:
                        os.replace(bak, tgt)       # success -> back to original
                    except Exception as re:
                        rollback_errors.append(f"{tgt}: restore failed ({re})")
                        keep.add(bak)              # KEEP — only surviving original
                        failed.add(tgt)            # this one really stayed changed
                elif not s["existed"]:             # remove newly-created target
                    try:
                        os.remove(tgt)
                    except Exception as re:
                        rollback_errors.append(f"{tgt}: cleanup failed ({re})")
                        failed.add(tgt)            # new file couldn't be removed
            _rm([s["tmp"] for s in staged]
                + [b for b in made_backups if b not in keep])
            if rollback_errors:
                changed = [(s["product"], s["ch"], "CHANGED (not restored)")
                           for s in staged if s["target"] in failed]
                return "rollback_incomplete", changed + [
                    ("—", "—", "commit failed and ROLLBACK INCOMPLETE: "
                     f"{e}. {'; '.join(rollback_errors)}. Original(s) preserved "
                     f"at: {sorted(keep)}")]
            return "rejected", [("—", "—", "commit failed — rolled back to "
                                 f"previous configs: {e}")]

        _rm(made_backups)                          # success — drop backups
        return "applied", [(s["product"], s["ch"],
                            f"written ({len(s['cfg'])} vars, "
                            f"{int((s['cfg'].role == 'exclude').sum())} excluded)")
                           for s in staged]


# ---------------- export workbook (X1) ----------------

def _period_cols_for(dates, years, time_map):
    """Model-year label + any uploaded time-map periods, per week. Alex asked
    for the time periods to travel WITH the weekly rows so the file can be
    pivoted in Excel (drop date in rows, driver in columns, period in filters)
    without VLOOKUPing a mapping back in."""
    ds = [d.strftime("%Y-%m-%d") for d in dates]
    per_week = [[] for _ in ds]
    for label, wanted in (time_map or {}).items():
        w = set(wanted)
        for i, d in enumerate(ds):
            if d in w:
                per_week[i].append(str(label))
    return [{"model_year": str(years[i]),
             "period": "; ".join(per_week[i]) if per_week[i] else "",
             "quarter": f"{dates[i].year} Q{((dates[i].month - 1) // 3) + 1}",
             "year": int(dates[i].year)} for i in range(len(ds))]


def _slice_export_rows(path, brand, channel, target, window, time_map=None,
                       payload=None, tuning=None):
    """(fit-stat row, fit-structure rows [template format], weekly due-to rows)
    for one slice. Served from the cached batch when available so Export does
    not silently re-run models that the screens already show."""
    load_or_create_config(path, brand)      # ensure the default exists (new data)
    cfg_file = resolve_cfg_path(path, brand, channel)
    p = payload or cached_slice(path, target, window, brand, channel, tuning)
    if p is None:
        _recomputed = True
        cfg = cp.ModelConfig(target=target, model_weeks=int(window),
                             window_end=dataset_anchor(path),
                             variable_config=cfg_file, **(tuning or {}))
        r = cp.run_slice("", brand, channel, config=cfg,
                         df=load_sheet(path, brand))
        p = _payload(r, brand, channel, target, window)

    _EXPORT_PAYLOADS[(brand, channel)] = p
    s = p["stats"]
    ho = np.nan if s["holdout_mape"] is None else s["holdout_mape"]
    stat = {"product": brand, "channel": channel, "target": target,
            "window_weeks": int(window),
            "window_start": p["window_start"], "window_end": p["window_end"],
            "weeks_modeled": p["n_weeks"], "selling_weeks": p["n_selling"],
            "R2": round(s["r2"], 4),
            "adj_R2": round(s["adj_r2"], 4), "MAPE_in_pct": round(s["mape"], 2),
            "MAPE_holdout_pct": None if np.isnan(ho) else round(ho, 2),
            "grade": _grade_label(ho), "n_selected": s["n_selected"],
            "n_forced": s["n_forced"]}

    # fit structure = the config actually used, in template format (so an
    # exported model can be tweaked in Excel and re-uploaded). Tagged with the
    # specific channel; re-uploading it materializes that channel's override.
    struct = _cfg_to_template_rows(pd.read_csv(cfg_file), brand, channel)
    # Alex: "add another column to the side that says what the coefficient
    # currently is… maybe add a flag for what made it through the model."
    # Both are IGNORED on re-upload (not template columns), so a round-trip
    # still works — they are read-only context for tuning in Excel.
    coefs = p["coef_map"]
    in_model = set(coefs)
    for row in struct:
        v = row.get("variable")
        row["current_coefficient"] = coefs.get(v)
        row["in_last_model"] = "yes" if v in in_model else "no"

    # weekly due-tos — actual per-week contributions, long/tidy, now carrying
    # the period columns so the file pivots straight away.
    dates = pd.to_datetime(p["dates"])
    pcols = _period_cols_for(dates, p["contrib_wk"]["years"], time_map)
    weekly = []
    cw = p["contrib_wk"]
    # include the Intercept line so intercept + Σ(driver due-tos) reproduces
    # the fitted value exactly for each week, as on screen
    drivers = ([d for d in cw["contrib"] if d == "Intercept"]
               + [d for d in cw["drivers"]])
    for i in range(len(dates)):
        base = {"product": brand, "channel": channel,
                "date": dates[i], **pcols[i]}
        for drv in drivers:
            weekly.append({**base, "driver": drv,
                           "due_to": float(cw["contrib"][drv][i])})
    return stat, struct, weekly


GRADE_FILL = {"Good Model": GREEN, "Moderate Model": WARN, "Bad Model": RED,
              "No holdout": "#98A2B3"}

# payloads produced during the current export, so the due-to sheet uses the
# SAME numbers as the other sheets even when nothing was cached
_EXPORT_PAYLOADS: dict = {}


def _dueto_sheet_rows(path, target, window, time_map, slices, tuning=None,
                      payloads=None):
    """Period-over-period DUE-TOs (the change), one row per driver per pair of
    periods — the number Alex/Arko actually present. Built for the model-year
    pairs plus any uploaded time-map periods that a slice covers."""
    rows = []
    for br, ch in slices:
        # prefer the payload the export just built for this slice: relying on
        # the cache alone silently produced an EMPTY due_to_change sheet
        # whenever the export's settings didn't match a cached batch
        p = (payloads or {}).get((br, ch)) \
            or cached_slice(path, target, window, br, ch, tuning)
        if p is None:
            continue
        cw = p["contrib_wk"]
        labels = [o["value"] for o in _avgc_period_options(cw, time_map)
                  if o["value"] != "all"]
        pairs = []
        if "previous" in labels and "current" in labels:
            pairs.append(("previous", "current"))
        mapped = [v for v in labels if str(v).startswith("map:")]
        pairs += [(mapped[i], mapped[i + 1]) for i in range(len(mapped) - 1)]
        for a, b in pairs:
            vals, na, nb, per_week = _dueto_values(cw, a, b, time_map)
            total = sum(vals.values())
            for drv, v in vals.items():
                rows.append({
                    "product": br, "channel": ch,
                    "period_from": _period_label(cw, a, time_map),
                    "period_to": _period_label(cw, b, time_map),
                    "weeks_from": na, "weeks_to": nb,
                    "basis": "per-week avg" if per_week else "totals",
                    "driver": drv, "due_to_change": v,
                    "share_of_total_change_pct":
                        (v / total * 100) if total else None,
                })
    return rows


def build_export_workbook(path, target, window, mode, brand=None, channel=None,
                          time_map=None, tuning=None):
    """Assemble the export: fit_statistics + fit_structure + weekly_due_tos
    (+ errors) for one slice or every slice. Slice enumeration AND per-slice
    runs are guarded so a missing target / bad product lands in the errors
    sheet instead of crashing. Returns (writer_fn, n_ok, n_err)."""
    stats, structs, weekly, errors = [], [], [], []
    slices = []
    if mode == "slice":
        slices = [(brand, channel)]
    else:
        for sheet in detect_product_sheets(path):
            try:
                dfx = load_sheet(path, sheet)
                if target not in dfx.columns:
                    errors.append({"product": sheet, "channel": "*",
                                   "error": f"target '{target}' not in product"})
                    continue
                for ch in channels_for(path, sheet):
                    if dfx.loc[dfx["Geography"] == ch, target].abs().sum() == 0:
                        continue                # not sold in channel — skip
                    slices.append((sheet, ch))
            except Exception as e:
                errors.append({"product": sheet, "channel": "*",
                               "error": f"{type(e).__name__}: {e}"})

    ok_slices = []
    _EXPORT_PAYLOADS.clear()
    for br, ch in slices:
        try:
            s, st, wk = _slice_export_rows(path, br, ch, target, window,
                                           time_map=time_map, tuning=tuning)
            stats.append(s)
            structs.extend(st)
            weekly.extend(wk)
            ok_slices.append((br, ch))
        # a slice with too little data in the window is reported, not modeled —
        # this is what stops the disjoint pre-window chunk Alex found in the
        # weekly due-tos from ever reaching the file again
        except cp.InsufficientWindowData as e:
            errors.append({"product": br, "channel": ch,
                           "error": f"not modeled — {e}"})
        except Exception as e:
            errors.append({"product": br, "channel": ch,
                           "error": f"{type(e).__name__}: {e}"})

    dueto_rows = _dueto_sheet_rows(path, target, window, time_map, ok_slices,
                                   tuning, payloads=dict(_EXPORT_PAYLOADS))
    struct_cols = TEMPLATE_COLS + ["current_coefficient", "in_last_model"]

    def _write(buf):
        from openpyxl.styles import Font, PatternFill
        stat_df = (pd.DataFrame(stats) if stats
                   else pd.DataFrame(columns=["product", "channel", "grade"]))
        with pd.ExcelWriter(buf, engine="openpyxl") as xw:
            stat_df.to_excel(xw, sheet_name="fit_statistics", index=False)
            pd.DataFrame(structs, columns=struct_cols).to_excel(
                xw, sheet_name="fit_structure", index=False)
            (pd.DataFrame(weekly) if weekly else pd.DataFrame(
                columns=["product", "channel", "date", "model_year", "period",
                         "quarter", "year", "driver", "due_to"])).to_excel(
                xw, sheet_name="weekly_due_tos", index=False)
            (pd.DataFrame(dueto_rows) if dueto_rows else pd.DataFrame(
                columns=["product", "channel", "period_from", "period_to",
                         "driver", "due_to_change"])).to_excel(
                xw, sheet_name="due_to_change", index=False)
            if errors:
                pd.DataFrame(errors).to_excel(xw, sheet_name="errors",
                                              index=False)
            # green/amber/red fill on the grade column (Alex: slide 6)
            if stats and "grade" in stat_df.columns:
                ws = xw.sheets["fit_statistics"]
                gcol = list(stat_df.columns).index("grade") + 1
                for i, val in enumerate(stat_df["grade"], start=2):
                    hexc = GRADE_FILL.get(val)
                    if not hexc:
                        continue
                    cell = ws.cell(row=i, column=gcol)
                    cell.fill = PatternFill("solid", fgColor=hexc.lstrip("#"))
                    cell.font = Font(color="FFFFFF", bold=True)
    return _write, len(stats), len(errors)


def batch_summary_path(path: str) -> str:
    ds = _slug(os.path.splitext(os.path.basename(path))[0])[:24]
    return os.path.join(ROOT, "outputs", f"batch_summary_{ds}.csv")


def fresh_default_df(path: str, sheet: str) -> pd.DataFrame:
    """Freshly auto-generate a config from the data (no file write). Delegates
    to the shared generator so ACV force-include lives in one place."""
    return cp.generate_default_config(load_sheet(path, sheet))


def load_or_create_config(path: str, sheet: str,
                          regenerate: bool = False) -> pd.DataFrame:
    # Delegate DEFAULT creation to the shared entry point so app + CLI create
    # the Product default identically (single creation path). Returns the df.
    # df is omitted so the sheet is read ONLY when a config is actually created
    # (no Excel read on the common "already exists" path).
    p = cp.load_or_create_default_config(path, sheet, regenerate=regenerate)
    return pd.read_csv(p)


# ═════════════════════════════════════════════════════════════════════════
# RESULTS CACHE — run once, filter many (Alex, 2026-07-27)
#
# The tool used to re-run the regression every time you changed the Product or
# Channel filter on a results screen. Alex: "instead of having to rerun every
# time, you just create the data that filters into this for everything, and
# then you just filter down… it's going to take a little bit longer for the
# first one, but then when we're evaluating, it's instantaneous."
#
# So: one batch run materializes EVERY slice's results into this cache (the
# same content the Excel export contains), and Diagnostics / Contributions /
# Export just read and filter it. It also removes a real hazard — you can no
# longer end up looking at one slice modeled on 156 weeks while every other
# slice on screen was modeled on 104.
#
# Keyed by (datafile, target, window) so changing a modeling knob can never
# silently show you results built under the old one.
# ═════════════════════════════════════════════════════════════════════════

RESULTS: dict = {}

# Selection strictness presets (Alex: "we want to be more restrictive, we want
# to be less restrictive"). Balanced == the engine's historical defaults, so
# leaving this alone reproduces every existing model exactly.
STRICTNESS = {"strict":   {"p_enter": 0.01, "vif_threshold": 5.0},
              "balanced": {"p_enter": 0.05, "vif_threshold": 10.0},
              "lax":      {"p_enter": 0.15, "vif_threshold": 20.0}}


def parse_caps(rows) -> dict:
    """Family-cap table rows -> {family: max}. Blank / 0 / junk = no cap, so a
    half-filled table can never silently throttle a family."""
    caps = {}
    for r in (rows or []):
        fam = str(r.get("family", "")).strip()
        raw = r.get("max")
        if not fam or raw in (None, "", "None"):
            continue
        try:
            v = int(float(raw))
        except (TypeError, ValueError):
            continue
        if v > 0:
            caps[fam] = v
    return caps


def tuning_of(strictness, cap_rows) -> dict:
    """The full selection-tuning bundle shared by every run path."""
    d = dict(STRICTNESS.get(strictness or "balanced", STRICTNESS["balanced"]))
    d["max_per_family"] = parse_caps(cap_rows)
    return d


def _tuning_key(tuning: dict) -> tuple:
    """Hashable signature of the tuning, so cached results built under one set
    of hyperparameters are never served for another."""
    t = tuning or {}
    return (t.get("p_enter"), t.get("vif_threshold"),
            tuple(sorted((t.get("max_per_family") or {}).items())))


@lru_cache(maxsize=8)
def dataset_anchor(path: str):
    """Latest week in the WHOLE datafile — the common window anchor."""
    return cp.dataset_anchor_week(path)


def window_range(path: str, window) -> tuple:
    return cp.resolve_window(dataset_anchor(path), int(window))


def window_caption(path: str, window) -> str:
    try:
        a, b = window_range(path, window)
        return f"{int(window)} wks · {a:%b %d, %Y} – {b:%b %d, %Y}"
    except Exception:
        return f"{window} wks"


def _results_key(path, target, window, tuning=None) -> tuple:
    return (os.path.abspath(str(path)), str(target), int(window),
            _tuning_key(tuning))


def _media_curves(r) -> list:
    """Response-curve data for each media variable in the model."""
    fit, specs_by_name = r["fit"], r["specs_by_name"]
    refs = r.get("sat_refs", {}) or {}
    out = []
    for c in r["selected"]:
        spec = specs_by_name.get(c)
        if spec is None or spec.adstock_decay is None or c not in r["X_raw"]:
            continue
        # x axis = execution AFTER adstock and scale but BEFORE saturation,
        # which is the quantity the Hill curve is a function of
        x = cp.adstock(r["X_raw"][c].astype(float).values, spec.adstock_decay)
        sc = float(getattr(spec, "scale", 1.0) or 1.0)
        if sc != 1.0:
            x = x / sc
        ref = float(refs.get(c) or (np.nanmax(x) if len(x) else 0.0) or 1.0)
        coef = float(fit.coef[c])
        curve = cp.response_curve(spec, coef, ref,
                                  x_max=float(np.nanmax(x)) if len(x) else ref)
        active = x > 0
        out.append({
            "variable": c,
            "curve_x": curve["x"], "curve_y": curve["y"],
            "saturated": curve["saturated"],
            "half_saturation": curve["half_saturation"],
            "points_x": [float(v) for v in x[active]],
            "points_y": [float(v) for v in
                         fit.contributions[c].values[active]],
            "decay": float(spec.adstock_decay),
            "sat_midpoint": spec.sat_midpoint, "sat_slope": spec.sat_slope,
            "max_effect": coef,
        })
    return out


def _payload(r, brand, channel, target, window) -> dict:
    """Everything the result screens need for ONE slice, as plain types.
    Built once per slice per batch; the screens never refit from this."""
    fit = r["fit"]
    df, sel = r["df"], r["selected"]
    cby = r["contrib_by_year"]
    year_cols = [c for c in cby.columns if not str(c).startswith("YoY")]
    dates = [d.strftime("%Y-%m-%d") for d in df["date"]]

    wkc = fit.contributions.reset_index(drop=True)
    drivers = [str(c) for c in wkc.columns if str(c) != "Intercept"]
    media = [c for c in drivers
             if (sp := r["specs_by_name"].get(c)) is not None
             and sp.adstock_decay is not None]
    n_highvif = int(sum(1 for c in sel
                        if not np.isnan(fit.vif.get(c, np.nan))
                        and fit.vif.get(c, 0) > VIF_WARN))
    coef_rows = []
    for name in ["const"] + list(sel):
        spec = r["specs_by_name"].get(name)
        t = fit.tstats.get(name, np.nan) if name != "const" else np.nan
        coef_rows.append({
            "name": name,
            "family": spec.family if spec else "(intercept)",
            "sign": spec.sign if spec else "",
            "t": None if (name == "const" or pd.isna(t)) else float(t),
            "vif": (None if name == "const"
                    or pd.isna(fit.vif.get(name, np.nan))
                    else float(fit.vif.get(name))),
            "coef": float(fit.coef[name]),
            "beta_std": (None if name == "const"
                         else float(fit.beta_std.get(name, np.nan))),
            "scale": (None if spec is None
                      else float(getattr(spec, "scale", 1.0) or 1.0)),
            "avg_dueto": float(r["avg_contrib"].loc[
                "Intercept" if name == "const" else name,
                "avg_weekly_contribution"]),
            "forced": name in r["forced"],
        })
    dueto_by_period = {f"Y{i+1}": {str(d): float(cby.loc[d, yc])
                                   for d in cby.index if str(d) != "Intercept"}
                       for i, yc in enumerate(year_cols)}
    dueto_by_period["Total"] = {
        str(d): float(cby.loc[d, year_cols].sum())
        for d in cby.index if str(d) != "Intercept"}

    return {
        "brand": brand, "channel": channel, "target": target,
        "window": int(window),
        "window_start": f"{r['window_start']:%Y-%m-%d}",
        "window_end": f"{r['window_end']:%Y-%m-%d}",
        "n_weeks": int(r["n_weeks_window"]),
        "n_selling": int(r["n_weeks_selling"]),
        "dates": dates,
        "actual": [float(v) for v in r["y"].values],
        "fitted": [float(v) for v in fit.fitted],
        "resid": [float(v) for v in fit.resid],
        "split": int(r["split"]),
        "pred_te": [float(v) for v in r["pred_te"]],
        "stats": {
            "r2": float(fit.r2), "adj_r2": float(fit.adj_r2),
            "mape": float(fit.mape),
            "holdout_mape": (None if np.isnan(r["holdout_mape"])
                             else float(r["holdout_mape"])),
            "durbin_watson": float(fit.meta["durbin_watson"]),
            "n_selected": len(sel), "n_forced": len(r["forced"]),
            "n_candidates": len(r["specs"]),
            "n_conflicts": len(r["sign_conflicts"]),
            "n_highvif": n_highvif,
        },
        "coef_rows": coef_rows,
        "cby": {"index": [str(i) for i in cby.index],
                "columns": [str(c) for c in cby.columns],
                "data": [[None if pd.isna(v) else float(v) for v in row]
                         for row in cby.values]},
        "year_labels": [str(c) for c in year_cols],
        "contrib_wk": {
            "dates": dates,
            "years": [str(v) for v in
                      cp.assign_model_years(df["date"].reset_index(drop=True)).values],
            "year_labels": [str(c) for c in year_cols],
            "drivers": drivers,
            # Intercept is kept in `contrib` (the exports decompose the fitted
            # line and must include it) but stays OUT of `drivers`, which is
            # what the charts iterate — it isn't a driver you can act on.
            "contrib": {str(c): [float(x) for x in wkc[c].values]
                        for c in wkc.columns},
            "media": media,
            "exec": {c: [1 if float(x) > 0 else 0 for x in r["X_raw"][c].values]
                     for c in media if c in r["X_raw"]},
        },
        # Media response curves: the fitted volume response against execution,
        # plus the weeks actually observed. On a saturated variable the
        # flattening IS the diminishing return Alex described ("your second
        # dollar spent, less cool"); on a linear one it is a straight line,
        # which is itself informative — no saturation was fitted.
        "media_curves": _media_curves(r),
        "coef_map": {str(k): float(fit.coef[k]) for k in fit.coef.index
                     if k != "const"},
        # comparable coefficient (volume per 1 SD move) — the number that can
        # legitimately be ranked across drivers (Arko)
        "beta_std": {str(k): float(v) for k, v in fit.beta_std.items()},
        "beta_range": {str(k): float(v) for k, v in fit.beta_range.items()},
        "dueto_by_period": dueto_by_period,
        "periods": [f"Y{i+1}" for i in range(len(year_cols))] + ["Total"],
    }


def build_all_results(path: str, target_pref: str, window: int,
                      tuning: Optional[dict] = None) -> dict:
    """Run EVERY Product × Channel once and cache the full results. Slices
    with too little data inside the fixed window are recorded as skipped —
    with the reason — instead of producing a phantom model."""
    anchor = dataset_anchor(path)
    models, skipped, errors = {}, [], []
    for sheet in detect_product_sheets(path):
        df_sheet = load_sheet(path, sheet)
        load_or_create_config(path, sheet)
        tgt = target_pref if target_pref in df_sheet.columns else next(
            (t for t in PREFERRED_TARGETS if t in df_sheet.columns), None)
        if tgt is None:
            continue
        for channel in channels_for(path, sheet):
            if df_sheet.loc[df_sheet["Geography"] == channel, tgt].abs().sum() == 0:
                skipped.append({"brand": sheet, "channel": channel,
                                "weeks_in_window": 0, "selling_weeks": 0,
                                "reason": "not sold in this channel "
                                          "(target is all zero)"})
                continue
            try:
                cfg = cp.ModelConfig(
                    target=tgt, model_weeks=int(window), window_end=anchor,
                    variable_config=resolve_cfg_path(path, sheet, channel),
                    **(tuning or {}))
                r = cp.run_slice("", sheet, channel, config=cfg, df=df_sheet)
                models[(sheet, channel)] = _payload(r, sheet, channel, tgt,
                                                    window)
            except cp.InsufficientWindowData as e:
                skipped.append({"brand": sheet, "channel": channel,
                                "weeks_in_window": e.weeks,
                                "selling_weeks": e.nonzero,
                                "reason": str(e)})
            except Exception as e:
                errors.append({"brand": sheet, "channel": channel,
                               "reason": f"{type(e).__name__}: {e}"})
    store = {"models": models, "skipped": skipped, "errors": errors,
             "target": target_pref, "window": int(window),
             "tuning": tuning or {}, "ts": time.strftime("%H:%M:%S")}
    RESULTS[_results_key(path, target_pref, window, tuning)] = store
    return store


def cached_store(path, target, window, tuning=None):
    if not path:
        return None
    return RESULTS.get(_results_key(path, target, window, tuning))


def cached_slice(path, target, window, brand, channel, tuning=None):
    st = cached_store(path, target, window, tuning)
    return (st or {}).get("models", {}).get((brand, channel))


def summary_from_store(store: dict) -> pd.DataFrame:
    rows = []
    for (b, c), p in store["models"].items():
        s = p["stats"]
        rows.append({
            "brand": b, "channel": c, "target": p["target"],
            "weeks": p["n_weeks"], "selling_weeks": p["n_selling"],
            "n_selected": s["n_selected"], "n_forced": s["n_forced"],
            "R2": round(s["r2"], 4),
            "MAPE_in_pct": round(s["mape"], 2),
            "MAPE_holdout_pct": (np.nan if s["holdout_mape"] is None
                                 else round(s["holdout_mape"], 2)),
            "n_sign_conflicts": s["n_conflicts"],
        })
    out = pd.DataFrame(rows)
    if len(out):
        out = out.sort_values("MAPE_holdout_pct")
    return out


def run_batch(path: str, target_pref: str, window: int,
              tuning: Optional[dict] = None) -> pd.DataFrame:
    """Batch = build the full results cache, then derive the summary table."""
    store = build_all_results(path, target_pref, window, tuning)
    out = summary_from_store(store)
    if len(out):
        out.to_csv(batch_summary_path(path), index=False)
    return out


# ---------------- formatting helpers ----------------

def fmt_compact(v) -> str:
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return "—"
    a = abs(v)
    sign = "−" if v < 0 else "+"
    if a >= 1e6:
        return f"{sign}{a/1e6:.1f}M"
    if a >= 1e3:
        return f"{sign}{a/1e3:.1f}k"
    return f"{sign}{a:.1f}"


def fmt_km(v, decimals: int = 2) -> str:
    """Chart data-label format (Alex, slide 3): 1,200,000→1.2m · 550,000→550k
    · 10,675→10.68k. Up to `decimals` places, trailing zeros trimmed. Shared
    helper — reused by the export tab later."""
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return ""
    a = abs(v)
    sign = "−" if v < 0 else ""
    if a >= 1e6:
        num, unit = a / 1e6, "m"
    elif a >= 1e3:
        num, unit = a / 1e3, "k"
    else:
        num, unit = a, ""
    s = f"{num:.{decimals}f}".rstrip("0").rstrip(".")
    return f"{sign}{s}{unit}"


def _fmt_vif(v):
    """VIF cell text + color. Forced collinear variables can be huge/∞ by
    design (they bypass VIF pruning) — show that plainly, don't mislead."""
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return "—", INK3
    if np.isinf(v) or v >= 1000:
        return ("∞" if np.isinf(v) else "≥1000"), RED
    if v > VIF_WARN:
        return f"{v:.1f}", WARN
    return f"{v:.1f}", INK2


# ---------- Contributions time filter (C1 / E3 / E4) ----------
# The "view period" is decoupled from the modeling WINDOW (E4): the model is fit
# on the window, but the avg-weekly-due-to view can be aggregated over any
# sub-period — entire window, a model year (current/previous), or the latest
# 13/4 weeks — by re-aggregating the stored per-week contributions.
CONTRIB_TAIL_PERIODS = [("q4", "Quarter 4 (Oct–Dec, latest)"),
                        ("last13", "Latest 13 weeks"),
                        ("last4", "Latest 4 weeks")]


def _year_months(cw):
    return [(int(d[:4]), int(d[5:7])) for d in cw.get("dates", []) if len(d) >= 7]


def _period_week_idx(cw, period, time_map=None):
    n = len(cw.get("dates", []))
    if time_map and isinstance(period, str) and period.startswith("map:"):
        wanted = set(time_map.get(period[4:], []))     # uploaded week→period map
        return [i for i, d in enumerate(cw.get("dates", [])) if d in wanted]
    if period == "last13":
        return list(range(max(0, n - 13), n))
    if period == "last4":
        return list(range(max(0, n - 4), n))
    if period == "q4":                          # calendar Q4 of the latest year
        ym = _year_months(cw)
        q4y = [y for (y, m) in ym if m in (10, 11, 12)]
        if not q4y:
            return []
        ty = max(q4y)
        return [i for i, (y, m) in enumerate(ym)
                if y == ty and m in (10, 11, 12)]
    yl = cw.get("year_labels", [])
    # semantic, window-stable: current = latest model year, previous = prior.
    # If the year doesn't exist (e.g. no previous year in a 52-wk window) return
    # EMPTY, never the fallthrough to "all weeks" (which would mislabel).
    if period == "current":
        return [j for j, v in enumerate(cw.get("years", [])) if yl and v == yl[-1]]
    if period == "previous":
        return ([j for j, v in enumerate(cw.get("years", [])) if v == yl[-2]]
                if len(yl) >= 2 else [])
    if period and period.startswith("Y") and period[1:].isdigit():  # legacy
        i = int(period[1:]) - 1
        if 0 <= i < len(yl):
            return [j for j, v in enumerate(cw.get("years", [])) if v == yl[i]]
    return list(range(n))                       # "all" / default


def _period_label(cw, period, time_map=None):
    if isinstance(period, str) and period.startswith("map:"):
        return f"{period[4:]} (mapped)"
    if period == "last13":
        return "Latest 13 weeks"
    if period == "last4":
        return "Latest 4 weeks"
    if period == "q4":
        q4y = [y for (y, m) in _year_months(cw) if m in (10, 11, 12)]
        return f"Q4 {max(q4y)} (Oct–Dec)" if q4y else "Quarter 4"
    yl = cw.get("year_labels", [])
    if period == "current":
        return f"Current year — {yl[-1]}" if yl else "Current year"
    if period == "previous":
        return f"Previous year — {yl[-2]}" if len(yl) >= 2 else "Previous year"
    if period and period.startswith("Y") and period[1:].isdigit():
        i = int(period[1:]) - 1
        return yl[i] if 0 <= i < len(yl) else period
    return "Entire window"


def _avgc_values(cw, period, time_map=None):
    """Avg weekly due-to per driver over the selected period; media averaged
    only over that period's execution weeks (matches the engine's rule)."""
    idx = _period_week_idx(cw, period, time_map)
    media, exec_ = set(cw.get("media", [])), cw.get("exec", {})
    out = {}
    for d in cw.get("drivers", []):
        vals = cw["contrib"][d]
        if d in media and d in exec_:
            ex = exec_[d]
            sel = [vals[i] for i in idx if ex[i] > 0]
        else:
            sel = [vals[i] for i in idx]
        out[d] = (sum(sel) / len(sel)) if sel else 0.0
    return out, len(idx)


def _room_for_labels(fig: go.Figure, values) -> go.Figure:
    """Stop long driver names and outside value labels from colliding (Alex,
    2026-07-27: "Competition ACV Weighted Distribution is kind of cutting off
    the number at the top graph").

    Two things fix it: `automargin` lets the y axis claim as much left gutter
    as the longest category name needs (instead of the name being clipped and
    running into the plot), and padding the x range by ~18% leaves space for
    the `textposition="outside"` value that would otherwise be drawn past the
    axis edge — on the longest bar, which is exactly where they overlapped.
    """
    fig.update_yaxes(automargin=True, ticklabelposition="outside")
    vals = [float(v) for v in np.ravel(list(values))
            if v is not None and not (isinstance(v, float) and np.isnan(v))]
    if vals:
        lo, hi = min(vals + [0.0]), max(vals + [0.0])
        pad = (hi - lo) * 0.18 or (abs(hi) or 1.0) * 0.18
        fig.update_xaxes(range=[lo - pad, hi + pad])
    fig.update_layout(margin=dict(l=8, r=28, t=8, b=8),
                      uniformtext=dict(mode="hide", minsize=8))
    return fig


def _avgc_figure(cw, period, time_map=None, compare=None):
    if not cw or not cw.get("drivers"):
        return _fig_base(go.Figure(), 320)
    vals, _ = _avgc_values(cw, period, time_map)
    if compare and compare != "none":               # period-vs-period (E4)
        vb, _ = _avgc_values(cw, compare, time_map)
        names = sorted(vals, key=lambda d: vals[d])  # order by primary period
        f = go.Figure()
        f.add_bar(y=names, x=[vals[d] for d in names],
                  name=_period_label(cw, period, time_map), orientation="h",
                  marker_color=NAVY, text=[fmt_km(vals[d]) for d in names],
                  textposition="outside", textfont=dict(size=8),
                  cliponaxis=False)
        f.add_bar(y=names, x=[vb[d] for d in names],
                  name=_period_label(cw, compare, time_map), orientation="h",
                  marker_color="#0E7090", text=[fmt_km(vb[d]) for d in names],
                  textposition="outside", textfont=dict(size=8),
                  cliponaxis=False)
        f.update_layout(barmode="group")
        _fig_base(f, max(340, 44 * len(names)))
        _room_for_labels(f, [vals[d] for d in names] + [vb[d] for d in names])
        return f
    ser = sorted(vals.items(), key=lambda kv: kv[1])
    names = [k for k, _ in ser]
    xs = [v for _, v in ser]
    f = go.Figure(go.Bar(
        y=names, x=xs, orientation="h",
        marker_color=[("#D9705E" if v < 0 else NAVY) for v in xs],
        text=[fmt_km(v) for v in xs], textposition="outside",
        textfont=dict(size=9), cliponaxis=False))
    _fig_base(f, max(320, 30 * len(names)))
    _room_for_labels(f, xs)
    return f


# ─────────────────────────────────────────────────────────────────────────
# DUE-TO = CHANGE between two periods (Arko + Alex, 2026-07-27)
#
# Arko: "due-to is like a change, right?" — Alex: "what you have here is
# contribution. What is the total per period? The due-to is just the change…
# that's really what all the people that look at this actually care about:
# how is the change of macroeconomic variables impacting my sales year to
# year." Both views are kept and are now clearly labelled as different things.
#
# Periods of different lengths are compared per-week, otherwise a 52-week
# period trivially "beats" a 4-week one and the due-to is meaningless.
# ─────────────────────────────────────────────────────────────────────────

def _dueto_values(cw, period_a, period_b, time_map=None):
    """(due-to per driver, weeks_a, weeks_b, per_week?) for B minus A."""
    idx_a = _period_week_idx(cw, period_a, time_map)
    idx_b = _period_week_idx(cw, period_b, time_map)
    na, nb = len(idx_a), len(idx_b)
    # unequal-length periods -> compare average weeks, not totals
    per_week = na != nb and na > 0 and nb > 0
    out = {}
    for d in cw.get("drivers", []):
        vals = cw["contrib"][d]
        ta = sum(vals[i] for i in idx_a)
        tb = sum(vals[i] for i in idx_b)
        if per_week:
            ta, tb = ta / max(na, 1), tb / max(nb, 1)
        out[d] = tb - ta
    return out, na, nb, per_week


def _dueto_figure(cw, period_a, period_b, time_map=None):
    # a due-to needs TWO periods; "none" on either side is not an error, it's
    # an incomplete selection — say so rather than drawing a misleading zero
    if not cw or not cw.get("drivers") or not period_a or not period_b \
            or "none" in (period_a, period_b):
        f = _fig_base(go.Figure(), 320)
        f.add_annotation(text="Pick a second period to compare —<br>a due-to "
                              "is the change between two periods.",
                         showarrow=False, font=dict(size=12, color=MUTED))
        return f
    vals, na, nb, per_week = _dueto_values(cw, period_a, period_b, time_map)
    ser = sorted(vals.items(), key=lambda kv: kv[1])
    names = [k for k, _ in ser]
    xs = [v for _, v in ser]
    f = go.Figure(go.Bar(
        y=names, x=xs, orientation="h",
        marker_color=[("#D9705E" if v < 0 else NAVY) for v in xs],
        text=[fmt_km(v) for v in xs], textposition="outside",
        textfont=dict(size=9), cliponaxis=False))
    f.add_vline(x=0, line_color="#D0D5DD")
    _fig_base(f, max(320, 30 * len(names)))
    _room_for_labels(f, xs)
    return f


def _dueto_caption(cw, period_a, period_b, time_map=None):
    if not cw or not period_a or not period_b or "none" in (period_a, period_b):
        return "Due-to = the change between two periods. Select a comparison."
    vals, na, nb, per_week = _dueto_values(cw, period_a, period_b, time_map)
    total = sum(vals.values())
    la = _period_label(cw, period_a, time_map)
    lb = _period_label(cw, period_b, time_map)
    basis = ("per-week averages (periods differ in length: "
             f"{na} vs {nb} wks)" if per_week else f"totals ({na} wks each)")
    return (f"Change from {la} → {lb}, by driver · {basis} · "
            f"total modeled change {fmt_km(total)}")


def _avgc_caption(cw, period, time_map=None, compare=None):
    na = len(_period_week_idx(cw, period, time_map)) if cw else 0
    if compare and compare != "none":
        nb = len(_period_week_idx(cw, compare, time_map))
        return (f"Comparing {_period_label(cw, period, time_map)} ({na} wks) "
                f"vs {_period_label(cw, compare, time_map)} ({nb} wks) · "
                "signed avg weekly (media over execution weeks)")
    return ("Signed; media averaged over execution weeks only · "
            f"{_period_label(cw, period, time_map)} ({na} wks)")


def _compare_options(cw, time_map=None, period=None):
    # exclude the current PRIMARY period so you can't compare A vs A. (Different
    # tokens that happen to cover the same dates — e.g. built-in Q4 vs a mapped
    # Q4 — are still allowed; only the identical token is removed.)
    return [{"label": "No comparison", "value": "none"}] + \
        [o for o in _avgc_period_options(cw, time_map) if o["value"] != period]


def _avgc_period_options(cw, time_map=None):
    # Semantic, window-stable labels (Alex): Current / Previous year always map
    # to the latest / prior model year regardless of the modeling window.
    yl = (cw or {}).get("year_labels", [])
    opts = [{"label": "Entire window", "value": "all"}]
    if yl:
        opts.append({"label": f"Current year — {yl[-1]}", "value": "current"})
    if len(yl) >= 2:
        opts.append({"label": f"Previous year — {yl[-2]}", "value": "previous"})
    opts += [{"label": l, "value": v} for v, l in CONTRIB_TAIL_PERIODS]
    # uploaded mapping-file periods (E3), scoped to weeks present in this run
    if time_map and cw:
        present = set(cw.get("dates", []))
        for lbl, dts in time_map.items():
            if present & set(dts):
                opts.append({"label": f"⤷ {lbl} (mapped)",
                             "value": f"map:{lbl}"})
    return opts


def fam_chip(family: str):
    bg, fg = FAM_CHIP.get(family, ("#F2F4F7", "#475467"))
    return html.Span(family, className="chip",
                     style={"background": bg, "color": fg})


def _fig_base(fig: go.Figure, height: int) -> go.Figure:
    fig.update_layout(
        height=height, margin=dict(l=8, r=8, t=8, b=8),
        paper_bgcolor="#FFFFFF", plot_bgcolor="#FFFFFF",
        font=dict(family="IBM Plex Sans, sans-serif", size=11, color=INK2),
        legend=dict(orientation="h", yanchor="bottom", y=1.01, x=1,
                    xanchor="right", font=dict(size=11)),
    )
    fig.update_xaxes(gridcolor=GRID, zerolinecolor="#D0D5DD",
                     tickfont=dict(family="IBM Plex Mono", size=10, color=MUTED))
    fig.update_yaxes(gridcolor=GRID, zerolinecolor="#D0D5DD",
                     tickfont=dict(family="IBM Plex Mono", size=10, color=MUTED))
    return fig


def card(title, sub, body, pad=True):
    # NB: test `is not None` — an empty Dash component is falsy (len == 0),
    # and dropping it silently invalidates any callback that outputs to it
    head = html.Div([html.Div(title, className="card-title")]
                    + ([html.Div(sub, className="card-sub")]
                       if sub is not None else []),
                    style={"display": "flex", "alignItems": "baseline",
                           "gap": "10px", "marginBottom": "4px"})
    return html.Div([head, body],
                    className="card" + (" card-pad" if pad else ""))


def _defn(term, *body):
    return html.Div(style={"marginBottom": "11px"}, children=[
        html.Span(term, style={"fontWeight": 600, "color": INK}),
        html.Span(" — ", style={"color": MUTED}),
        html.Span(list(body), style={"color": INK2}),
    ])


def _defs_section(title, *items):
    return html.Div(className="card card-pad", style={"marginBottom": "14px"},
                    children=[html.Div(title, className="card-title",
                                       style={"marginBottom": "10px"})]
                    + list(items))


def _icon(path_d, color=INK3):
    return html.Span(
        DangerouslySetSvg(path_d, color),
        style={"display": "flex", "width": "16px", "height": "16px"})


def DangerouslySetSvg(path_d, color):
    # dash html can't inline raw svg easily; use an Img data URL
    parts = "".join(f'<path d="{p.strip()}"/>' for p in path_d.split("M") if p.strip())
    paths = "".join(f'<path d="M{p.strip()}"/>' for p in path_d.split("M") if p.strip())
    svg = (f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 16 16" '
           f'width="16" height="16" fill="none" stroke="{color}" '
           f'stroke-width="1.6" stroke-linecap="round" stroke-linejoin="round">'
           f'{paths}</svg>')
    import base64
    b64 = base64.b64encode(svg.encode()).decode()
    return html.Img(src=f"data:image/svg+xml;base64,{b64}",
                    style={"width": "16px", "height": "16px"})


# ---------------- app & layout ----------------

app = Dash(__name__, title="Big Chalk Mix Engine")

TBL_BASE = dict(
    style_cell={"fontFamily": "IBM Plex Sans, sans-serif", "fontSize": "12.5px",
                "textAlign": "left", "padding": "6px 10px",
                "backgroundColor": "#fff", "color": INK},
    style_header={"fontWeight": "600"},
    style_as_list_view=True,
)


def nav_button(key, label, active=False):
    return html.Button(
        [_icon(NAV_ICONS[key], NAVY if active else INK3), label],
        id=f"nav-{key}", n_clicks=0,
        className="active" if active else "")


def pick(label, dd_id, width="140px", wide=False, help=None):
    lab = html.Div(
        [label] + ([html.Span(" ⓘ", style={"color": MUTED,
                                            "cursor": "help"})] if help else []),
        className="plabel", title=help or "")
    return html.Div([
        lab,
        dcc.Dropdown(id=dd_id, clearable=False,
                     style={"width": width}),
    ], className="bench-pick" + (" wide" if wide else ""), title=help or "")


app.layout = html.Div(className="bench-app", children=[
    dcc.Store(id="datapath"),
    dcc.Store(id="batch_sel"),
    # most-recent Diagnostics run — shared context for the Variables tab
    # (coef/due-to columns), and reused later by data-download & VIF views
    dcc.Store(id="last_run"),
    # active workspace tab — driven by the left nav and by Run (which
    # redirects to Diagnostics). Single renderer keeps show/hide in one place.
    dcc.Store(id="active_tab", data="diag"),
    # bumped AFTER any config write/delete (save / regenerate / reset). The
    # table reload and scope-status readers depend on it, so Dash's dependency
    # graph guarantees they re-read disk only after the write — no race.
    dcc.Store(id="config_rev", data=0),
    # per-week contributions of the latest run — lets the Contributions time
    # filter re-aggregate the avg-weekly view without re-running the model.
    dcc.Store(id="contrib_wk"),
    # uploaded week→period mapping (E3): {period label: [YYYY-MM-DD, …]}.
    dcc.Store(id="time_map"),

    # ═══ left rail ═══
    html.Div(className="bench-rail", children=[
        html.Div(className="bench-logo", children=[
            html.Div("BC", className="bench-logo-mark"),
            html.Div([html.Div("BIG CHALK", className="bench-logo-name"),
                      html.Div("Mix Engine", className="bench-logo-sub")]),
        ]),
        html.Div(className="bench-navwrap", children=[
            html.Div("WORKSPACE", className="bench-navhead"),
            html.Div(className="bench-nav", children=[
                nav_button(k, lbl, active=(k == "diag"))
                for k, lbl in SCREENS]),
        ]),
        html.Div(className="bench-dataset", children=[
            html.Div("DATASET", className="head"),
            html.Div(id="dataset_file", className="file", children="—"),
            html.Div(id="load_status", className="meta",
                     children="Enter a path and load."),
            dcc.Input(id="datafile_input", type="text",
                      value=DEFAULT_DATA if os.path.exists(DEFAULT_DATA) else "",
                      placeholder="C:\\path\\to\\data.xlsx"),
            html.Button("Load data", id="load_data", n_clicks=0,
                        className="btn btn-primary btn-small loadbtn"),
        ]),
    ]),

    # ═══ main column ═══
    html.Div(className="bench-main", children=[

        html.Div(className="bench-topbar", children=[
            # The top bar is a RESULTS FILTER, nothing more (Alex 2026-07-27):
            # "the only thing you should be filtering on are results, which
            # would be your product and channel". TARGET and WINDOW are
            # modeling decisions — changing them means re-running — so they
            # live in ONE place, the Variables screen, and are echoed here
            # read-only so you always know what you're looking at.
            html.Div(className="bench-picks", children=[
                pick("PRODUCT", "brand", width="150px"),
                pick("CHANNEL", "channel", width="150px"),
            ]),
            html.Div(id="model_settings_note", className="bench-settings-note",
                     style={"fontSize": "12px", "color": INK3,
                            "marginLeft": "14px", "lineHeight": "1.35"}),
            html.Div(style={"flex": 1}),
            # Run lives on Variables (tuning) and Batch (per-slice) only — the
            # top bar just reports the latest run. (Alex: slides 1–4)
            html.Div(id="runstat", className="bench-runstat"),
        ]),

        # ── Diagnostics ──
        html.Div(id="scr-diag", className="bench-screen", children=[dcc.Loading(children=[
            html.Div(id="err", style={"color": RED, "fontSize": "13px",
                                      "fontWeight": 600}),
            html.Div(id="metrics", className="bench-metrics"),
            html.Div(style={"display": "flex", "gap": "14px",
                            "marginTop": "14px"}, children=[
                html.Div(style={"flex": 2.1}, children=[card(
                    "Actual vs fitted — weekly", html.Span(id="fit_sub"),
                    dcc.Graph(id="fig_fit", config={"displayModeBar": False}))]),
                html.Div(style={"flex": 1, "display": "flex",
                                "flexDirection": "column", "gap": "14px"},
                         children=[
                    card("Residuals vs fitted", None,
                         dcc.Graph(id="fig_resid",
                                   config={"displayModeBar": False})),
                    card("Residual distribution", None,
                         dcc.Graph(id="fig_residhist",
                                   config={"displayModeBar": False})),
                ]),
            ]),
            html.Div(className="card", style={"marginTop": "14px"}, children=[
                html.Div(className="card-head", children=[
                    html.Div("Coefficients", className="card-title"),
                    html.Div(id="coef_sub", className="card-sub"),
                ]),
                html.Div(id="coef_grid"),
            ]),
            html.Div(className="card card-pad", style={"marginTop": "14px"},
                     children=[
                html.Div("Media response curves", className="card-title"),
                html.Div(id="curves_sub", className="card-sub",
                         style={"marginBottom": "4px"}),
                dcc.Graph(id="fig_curves", config={"displayModeBar": False}),
            ]),
        ])]),

        # ── Variables ──
        html.Div(id="scr-vars", className="bench-screen",
                 style={"display": "none"}, children=[
            # ── modeling settings: the ONE place a target / window is chosen ──
            html.Div(className="card card-pad",
                     style={"marginBottom": "10px", "display": "flex",
                            "alignItems": "center", "gap": "14px"}, children=[
                html.Div([
                    html.Div("MODELING SETTINGS", style={
                        "fontSize": "11px", "fontWeight": 600,
                        "letterSpacing": ".04em", "color": INK3}),
                    html.Div("These define the model itself. Changing them "
                             "requires a re-run — that is why they live here "
                             "and on no results screen.",
                             style={"fontSize": "11.5px", "color": MUTED}),
                ], style={"minWidth": "230px"}),
                pick("TARGET", "target", width="210px", wide=True,
                     help="The dependent variable being modeled. Changing it "
                          "means an entirely new model — nothing updates "
                          "until you Run."),
                pick("WINDOW", "window", width="120px",
                     help=("How many weeks the model covers. The window is a "
                           "FIXED calendar range ending at the dataset's "
                           "latest week and is the SAME for every slice, so "
                           "models are comparable and a delisted slice can't "
                           "be modeled on stale history. The last up to 13 "
                           "weeks inside it are reserved for validation.")),
                html.Div(id="vars_window_note",
                         style={"fontSize": "12px", "color": NAVY,
                                "fontWeight": 500}),
            ]),

            # ── selection strictness + per-family caps (Alex's F4 ask) ──
            html.Div(className="card card-pad",
                     style={"marginBottom": "10px"}, children=[
                html.Div(style={"display": "flex", "alignItems": "center",
                                "gap": "14px", "flexWrap": "wrap"}, children=[
                    html.Div([
                        html.Div("VARIABLE SELECTION", style={
                            "fontSize": "11px", "fontWeight": 600,
                            "letterSpacing": ".04em", "color": INK3}),
                        html.Div("How hard the model is to get into.",
                                 style={"fontSize": "11.5px",
                                        "color": MUTED}),
                    ], style={"minWidth": "230px"}),
                    html.Div([
                        html.Div("STRICTNESS", className="plabel",
                                 title="Punitive = only strongly significant "
                                       "variables enter (p<0.01) and "
                                       "collinearity is cut hard (VIF>5). "
                                       "Lax = lets more in (p<0.15, VIF>20). "
                                       "Balanced is the current default and "
                                       "reproduces every existing model."),
                        dcc.Dropdown(id="strictness", clearable=False,
                                     options=[
                                         {"label": "Punitive (p<0.01, VIF 5)",
                                          "value": "strict"},
                                         {"label": "Balanced (p<0.05, VIF 10)",
                                          "value": "balanced"},
                                         {"label": "Lax (p<0.15, VIF 20)",
                                          "value": "lax"}],
                                     value="balanced",
                                     style={"width": "215px",
                                            "fontSize": "12px"}),
                    ]),
                    html.Div(id="strictness_note",
                             style={"fontSize": "12px", "color": MUTED}),
                ]),
                html.Div(style={"display": "flex", "gap": "16px",
                                "marginTop": "10px",
                                "alignItems": "flex-start"}, children=[
                    html.Div(style={"flex": "0 0 330px"}, children=[
                        html.Div("Max variables per family", style={
                            "fontSize": "12px", "fontWeight": 600,
                            "marginBottom": "4px"}),
                        dash_table.DataTable(
                            id="family_caps", editable=True,
                            columns=[{"id": "family", "name": "family",
                                      "editable": False},
                                     {"id": "max", "name": "max",
                                      "type": "numeric"}],
                            data=[{"family": f, "max": None}
                                  for f in FAMILIES],
                            style_cell_conditional=[
                                {"if": {"column_id": "max"},
                                 "textAlign": "right", "width": "80px",
                                 "fontFamily": "IBM Plex Mono, monospace"}],
                            **TBL_BASE),
                    ]),
                    html.Div(style={"flex": 1, "fontSize": "12px",
                                    "color": INK3, "lineHeight": "1.5"},
                             children=[
                        html.Div([
                            html.B("Blank = no limit."), " Set a number to cap "
                            "how many of that family may enter — e.g. Trade 3, "
                            "Competitive 2, Macro 2, Media left blank is "
                            "Alex's “only two to three trade, but all media”."],
                            style={"marginBottom": "6px"}),
                        html.Div("The cap is applied while selecting, not by "
                                 "trimming afterwards, so a family spends its "
                                 "budget on its most significant members and "
                                 "the other families still fill up normally."),
                        html.Div(["Forced variables are always kept and count "
                                  "toward the budget — a cap of 2 with 2 "
                                  "forced Trade variables admits no more "
                                  "Trade."], style={"marginTop": "6px"}),
                    ]),
                ]),
            ]),
            html.Div(style={"display": "flex", "gap": "10px",
                            "alignItems": "center"}, children=[
                html.Div("CONFIG", className="plabel",
                         title="Edit the Product default (all channels) or a "
                               "Channel-specific override. At runtime an "
                               "override wins over the default for that "
                               "channel."),
                dcc.Dropdown(id="config_scope", clearable=False,
                             options=[{"label": "Product default",
                                       "value": "default"}],
                             value="default",
                             style={"width": "210px", "fontSize": "12px"}),
                html.Div("DUE-TO PERIOD", className="plabel",
                         title="Which model period the due-to context column "
                               "shows (Y1 / Y2 / … / Total). Coefficient is "
                               "the same across periods."),
                dcc.Dropdown(id="dueto_period", clearable=False,
                             options=[{"label": "Total", "value": "Total"}],
                             value="Total",
                             style={"width": "110px", "fontSize": "12px"}),
                dcc.Checklist(id="hide_excluded",
                              options=[{"label": " Hide excluded",
                                        "value": "hide"}],
                              value=[], style={"fontSize": "12.5px",
                                               "color": INK2}),
                html.Div(style={"flex": 1}),
                dcc.Upload(id="tmpl_upload", multiple=False, children=html.Button(
                    "Upload template", className="btn btn-ghost",
                    title="Upload a filled template (.xlsx/.csv). Rows with "
                          "channel=ALL write the Product default; a specific "
                          "channel writes that override; variables not listed "
                          "for a group become excluded.")),
                html.Button("Download template", id="tmpl_dl_btn", n_clicks=0,
                            className="btn btn-ghost",
                            title="Download this product's config (default + "
                                  "overrides) as an editable template."),
                dcc.Download(id="tmpl_download"),
                html.Button("Reset to default", id="cfg_reset", n_clicks=0,
                            className="btn btn-ghost",
                            title="In a channel scope, delete this channel's "
                                  "override so it inherits the Product default "
                                  "again."),
                html.Button("Regenerate defaults", id="cfg_regen", n_clicks=0,
                            className="btn btn-ghost"),
                html.Button("Save config", id="cfg_save", n_clicks=0,
                            className="btn btn-ghost"),
                html.Button("◈  Optimize media curves", id="opt_curves",
                            n_clicks=0, className="btn btn-ghost",
                            title="Search adstock decay + Hill saturation "
                                  "(midpoint, slope) for each media variable "
                                  "in this slice's model. Scored by "
                                  "rolling-origin cross-validation that never "
                                  "touches the reported holdout, so the "
                                  "holdout stays an honest check on whether "
                                  "the optimization actually helped. Writes "
                                  "the winning curves into the table below; "
                                  "review, then Run."),
                html.Button("▶  Run model", id="run_vars", n_clicks=0,
                            className="btn btn-primary",
                            style={"height": "36px"}),
            ]),
            html.Div(id="opt_status", style={"fontSize": "12px",
                                             "color": INK2,
                                             "marginTop": "2px"}),
            html.Div(id="cfg_scope_status",
                     style={"fontSize": "12px", "color": NAVY,
                            "fontWeight": 500, "marginTop": "2px"}),
            html.Div(id="cfg_status", style={"fontSize": "12px",
                                             "color": INK3}),
            html.Div(className="bench-banner", children=[html.Span([
                html.B("Sign is a guardrail, not a wish. ",
                       style={"fontWeight": 600}),
                "Positive-signed variables can never enter with a negative "
                "coefficient. Role ", html.B("force", style={"fontWeight": 600}),
                " keeps a variable in every run; ",
                html.B("exclude", style={"fontWeight": 600}),
                " bans it. Custom bounds override the sign. "
                "Adstock applies to media only."])]),
            html.Div(className="card", children=[dash_table.DataTable(
                id="cfg_table", editable=True, page_size=200,
                filter_action="native",
                columns=[
                    {"id": "variable", "name": "variable", "editable": False},
                    {"id": "family", "name": "family"},
                    {"id": "sign", "name": "expected sign",
                     "presentation": "dropdown"},
                    {"id": "adstock_decay", "name": "adstock", "type": "numeric"},
                    {"id": "scale", "name": "scale ÷", "type": "numeric"},
                    {"id": "sat_midpoint", "name": "sat mid",
                     "type": "numeric"},
                    {"id": "sat_slope", "name": "sat slope",
                     "type": "numeric"},
                    {"id": "coef_lower", "name": "bound lo", "type": "numeric"},
                    {"id": "coef_upper", "name": "bound hi", "type": "numeric"},
                    {"id": "cur_coef", "name": "coef now", "editable": False},
                    {"id": "cur_impact", "name": "impact / SD",
                     "editable": False},
                    {"id": "cur_dueto", "name": "contribution",
                     "editable": False},
                    {"id": "role", "name": "role", "presentation": "dropdown"},
                ],
                style_cell_conditional=(
                    [{"if": {"column_id": c}, "textAlign": "right",
                      "fontFamily": "IBM Plex Mono, monospace"}
                     for c in ("adstock_decay", "scale", "sat_midpoint",
                               "sat_slope", "coef_lower", "coef_upper",
                               "cur_coef", "cur_impact", "cur_dueto")]
                    + [{"if": {"column_id": c}, "color": INK3,
                        "backgroundColor": "#FBFCFD"}
                       for c in ("cur_coef", "cur_impact", "cur_dueto")]),
                dropdown={
                    "sign": {"options": [{"label": s, "value": s} for s in
                             ["positive", "negative", "unconstrained"]]},
                    "role": {"options": [{"label": s, "value": s} for s in
                             ["auto", "force", "exclude"]]},
                },
                style_data_conditional=[
                    {"if": {"filter_query": '{role} = "force"'},
                     "backgroundColor": "#F5F8FD"},
                    {"if": {"filter_query": '{role} = "exclude"'},
                     "backgroundColor": "#FFF6F5"},
                ],
                **TBL_BASE,
            )]),
        ]),

        # ── Contributions ──
        html.Div(id="scr-contrib", className="bench-screen",
                 style={"display": "none"}, children=[dcc.Loading(children=[
            html.Div(style={"display": "flex", "alignItems": "center",
                            "gap": "10px", "marginBottom": "10px"}, children=[
                html.Div("TIME MAP", className="plabel",
                         title="Optional: upload a .csv/.xlsx with 'date' (or "
                               "'week') + 'period' columns to define custom "
                               "periods (Current Year, Quarter 4, Latest 13 "
                               "weeks, …). They appear in the period filter as "
                               "“⤷ … (mapped)”."),
                dcc.Upload(id="timemap_upload", multiple=False,
                           children=html.Button("Upload time map",
                                                className="btn btn-ghost")),
                html.Div(id="timemap_status",
                         style={"fontSize": "12px", "color": INK3}),
            ]),
            html.Div(style={"display": "flex", "gap": "14px"}, children=[
                html.Div(style={"flex": 1.2}, children=[card(
                    "Due-to totals by model year", None,
                    dcc.Graph(id="fig_cby",
                              config={"displayModeBar": False}))]),
                html.Div(style={"flex": 1}, children=[html.Div(
                    className="card card-pad", children=[
                    html.Div(style={"display": "flex", "alignItems": "center",
                                    "justifyContent": "space-between",
                                    "gap": "10px", "marginBottom": "2px"},
                             children=[
                        html.Div(id="avgc_title",
                                 children="Avg weekly contribution per driver",
                                 className="card-title"),
                        html.Div(style={"display": "flex", "gap": "6px",
                                        "alignItems": "center"}, children=[
                            dcc.Dropdown(id="avgc_period", clearable=False,
                                         options=[{"label": "Entire window",
                                                   "value": "all"}],
                                         value="all",
                                         style={"width": "160px",
                                                "fontSize": "12px"}),
                            html.Span("vs", style={"fontSize": "12px",
                                                   "color": MUTED}),
                            dcc.Dropdown(id="avgc_compare", clearable=False,
                                         options=[{"label": "No comparison",
                                                   "value": "none"}],
                                         value="none",
                                         style={"width": "160px",
                                                "fontSize": "12px"}),
                        ]),
                    ]),
                    # Contribution (level in a period) vs due-to (change
                    # between two periods) — different questions, one toggle.
                    dcc.RadioItems(
                        id="contrib_mode",
                        options=[
                            {"label": " Contribution (level in a period)",
                             "value": "contrib"},
                            {"label": " Due-to (change between the two)",
                             "value": "dueto"}],
                        value="contrib", inline=True,
                        style={"fontSize": "12px", "margin": "2px 0 2px"},
                        labelStyle={"marginRight": "14px"}),
                    html.Div("Signed; media averaged over execution weeks only",
                             id="avgc_sub", className="card-sub",
                             style={"marginBottom": "4px"}),
                    dcc.Graph(id="fig_avgc",
                              config={"displayModeBar": False}),
                ])]),
            ]),
            html.Div(className="card", style={"marginTop": "14px"}, children=[
                html.Div(className="card-head", children=[
                    html.Div("Due-tos by model year", className="card-title"),
                    html.Div("Decomposes the fitted line exactly: intercept "
                             "+ Σ coefficient × value, per week",
                             className="card-sub"),
                ]),
                html.Div(id="yoy_grid"),
            ]),
        ])]),

        # ── Batch ──
        # ── High Level (Total Brand / Total Channel rollup) ──
        # Alex 2026-08-11: "Add a 'Total' view for Brand or Channel. This means
        # that each brand*channel is rolled up to either Total Brand or Total
        # Channel, to see the overall fit and contributions."
        html.Div(id="scr-total", className="bench-screen",
                 style={"display": "none"}, children=[
            html.Div(style={"display": "flex", "alignItems": "center",
                            "gap": "12px"}, children=[
                html.Div("High level view", style={"fontSize": "15px",
                                                   "fontWeight": 600}),
                dcc.RadioItems(
                    id="total_mode",
                    options=[{"label": " Total Brand", "value": "brand"},
                             {"label": " Total Channel", "value": "channel"},
                             {"label": " Grand total", "value": "all"}],
                    value="brand", inline=True,
                    labelStyle={"marginRight": "14px", "fontSize": "13px"}),
                html.Div(style={"flex": 1}),
                html.Div(id="total_status",
                         style={"fontSize": "12px", "color": INK3}),
            ]),
            html.Div("Each Product × Channel model is kept as it is; what is "
                     "summed is the OUTPUT — actual, fitted and every driver's "
                     "weekly contribution. The decomposition stays exact, so "
                     "the rolled-up contributions still add to the rolled-up "
                     "fitted line. Totals fit better than their parts because "
                     "independent errors cancel when series are added, which "
                     "is why the member grid is shown underneath.",
                     style={"fontSize": "12px", "color": INK3,
                            "margin": "4px 0 12px", "maxWidth": "980px"}),
            dcc.Loading(children=[
                html.Div(id="total_summary"),
                html.Div(style={"display": "flex", "gap": "14px",
                                "marginTop": "14px", "alignItems": "center"},
                         children=[
                    html.Div("GROUP", className="plabel"),
                    dcc.Dropdown(id="total_group", clearable=False,
                                 style={"width": "260px"}),
                ]),
                html.Div(id="total_detail", style={"marginTop": "12px"}),
            ]),
        ]),

        # ── Saturation curves ──
        # Alex 2026-08-11 + the call: the weekly raw/decayed/saturated
        # diagnostic with correlations, the execution-vs-ROI curve from his
        # SatCurve class, and — his words — "some way of either a screen or
        # pop up or something that lets the user play around with it".
        html.Div(id="scr-sat", className="bench-screen",
                 style={"display": "none"}, children=[
            html.Div(style={"display": "flex", "alignItems": "center",
                            "gap": "12px"}, children=[
                html.Div("Saturation & response curves",
                         style={"fontSize": "15px", "fontWeight": 600}),
                html.Div("MEDIA VARIABLE", className="plabel",
                         style={"marginLeft": "8px"}),
                dcc.Dropdown(id="sat_var", clearable=False,
                             style={"width": "300px"}),
                html.Div(style={"flex": 1}),
                html.Div(id="sat_status",
                         style={"fontSize": "12px", "color": INK3}),
            ]),
            html.Div(id="sat_note",
                     style={"fontSize": "12px", "color": INK3,
                            "margin": "4px 0 12px", "maxWidth": "980px"}),

            # ── the override panel ──
            html.Div(className="card card-pad", style={"marginBottom": "14px"},
                     children=[
                html.Div("Curve parameters — override and see the effect",
                         className="card-title"),
                html.Div("These start at the values the model actually used. "
                         "Change one and every chart below redraws instantly; "
                         "the correlations and the optimal point move with it. "
                         "Nothing here is saved and no model is re-run — this "
                         "is a what-if bench. “Apply to config” writes the "
                         "current three numbers into this product's variable "
                         "config so the next Run uses them.",
                         style={"fontSize": "12px", "color": INK3,
                                "margin": "2px 0 10px"}),
                html.Div(style={"display": "flex", "gap": "26px",
                                "flexWrap": "wrap"}, children=[
                    html.Div(style={"minWidth": "230px"}, children=[
                        html.Div("ADSTOCK DECAY", className="plabel"),
                        dcc.Slider(id="sat_decay", min=0, max=0.9, step=0.05,
                                   value=0.5, tooltip={"placement": "bottom",
                                                       "always_visible": True},
                                   marks={0: "0", 0.45: "0.45", 0.9: "0.9"}),
                    ]),
                    html.Div(style={"minWidth": "230px"}, children=[
                        html.Div("HILL MIDPOINT (γ)", className="plabel"),
                        dcc.Slider(id="sat_mid", min=0.05, max=0.95, step=0.05,
                                   value=0.5, tooltip={"placement": "bottom",
                                                       "always_visible": True},
                                   marks={0.05: "0.05", 0.5: "0.5", 0.95: "0.95"}),
                    ]),
                    html.Div(style={"minWidth": "230px"}, children=[
                        html.Div("HILL SLOPE (s)", className="plabel"),
                        dcc.Slider(id="sat_slope", min=0.3, max=4.0, step=0.1,
                                   value=1.0, tooltip={"placement": "bottom",
                                                       "always_visible": True},
                                   marks={0.3: "0.3", 1: "1", 2: "2", 4: "4"}),
                    ]),
                    html.Div(style={"display": "flex", "alignItems": "flex-end",
                                    "gap": "8px"}, children=[
                        html.Button("Reset to model", id="sat_reset",
                                    n_clicks=0, className="btn btn-ghost"),
                        html.Button("Apply to config", id="sat_apply",
                                    n_clicks=0, className="btn btn-primary"),
                    ]),
                ]),
                html.Div(id="sat_apply_status",
                         style={"fontSize": "12px", "color": INK3,
                                "marginTop": "8px"}),
            ]),

            dcc.Loading(children=[
                html.Div(id="sat_weekly"),
                html.Div(id="sat_corr", style={"marginTop": "14px"}),

                # ── ROI curve (Alex's SatCurve) ──
                html.Div(className="card card-pad",
                         style={"marginTop": "14px"}, children=[
                    html.Div("Execution vs ROI — annual view",
                             className="card-title"),
                    html.Div("Alex's SatCurve, over the latest 52 weeks. "
                             "x is % of current spend (100% = today). The "
                             "shaded area is sales driven; the two lines are "
                             "average and marginal return. Where they cross is "
                             "the optimal spend level.",
                             style={"fontSize": "12px", "color": INK3,
                                    "margin": "2px 0 10px"}),
                    html.Div(style={"display": "flex", "gap": "18px",
                                    "flexWrap": "wrap",
                                    "marginBottom": "10px"}, children=[
                        html.Div([html.Div("ANNUAL SPEND ($)",
                                           className="plabel"),
                                  dcc.Input(id="sat_spend", type="number",
                                            debounce=True,
                                            style={"width": "150px"})]),
                        html.Div([html.Div("PRICE PER UNIT ($)",
                                           className="plabel"),
                                  dcc.Input(id="sat_price", type="number",
                                            debounce=True, step=0.01,
                                            style={"width": "120px"})]),
                        html.Div([html.Div("MARGIN (%)", className="plabel"),
                                  dcc.Input(id="sat_margin", type="number",
                                            value=30, debounce=True, min=1,
                                            max=100, step=1,
                                            style={"width": "100px"})]),
                        html.Div(id="sat_inputs_note",
                                 style={"fontSize": "11px", "color": MUTED,
                                        "alignSelf": "flex-end",
                                        "maxWidth": "380px"}),
                    ]),
                    html.Div(id="sat_roi"),
                ]),
            ]),
        ]),

        # ── Multi-level modeling ──
        # Alex 2026-08-11: unpooled (done) / pooled / hierarchical.
        html.Div(id="scr-multi", className="bench-screen",
                 style={"display": "none"}, children=[
            html.Div(style={"display": "flex", "alignItems": "center",
                            "gap": "12px"}, children=[
                html.Div("Multi-level modeling", style={"fontSize": "15px",
                                                        "fontWeight": 600}),
                html.Div(id="multi_brand_note",
                         style={"fontSize": "12px", "color": INK3}),
                html.Div(style={"flex": 1}),
                html.Button("▶  Run all three levels", id="multi_run",
                            n_clicks=0, className="btn btn-primary",
                            style={"height": "36px"}),
            ]),
            html.Div("Unpooled = one model per Product × Channel (what the "
                     "other screens show). Pooled = ONE model for the product, "
                     "every channel's weeks stacked, one coefficient per "
                     "predictor. Hierarchical = each channel's coefficient is "
                     "the pooled one times an index taken from that channel's "
                     "own unconstrained fit. Runs on the PRODUCT selected in "
                     "the top bar; the channel picker does not apply here.",
                     style={"fontSize": "12px", "color": INK3,
                            "margin": "4px 0 12px", "maxWidth": "980px"}),
            html.Div(className="card card-pad", style={"marginBottom": "14px"},
                     children=[
                html.Div("Hierarchy controls", className="card-title"),
                html.Div(style={"display": "flex", "gap": "28px",
                                "flexWrap": "wrap", "marginTop": "8px"},
                         children=[
                    html.Div(style={"minWidth": "300px"}, children=[
                        html.Div("CHANNEL VARIATION ALLOWED (shrink λ)",
                                 className="plabel"),
                        dcc.Slider(id="multi_shrink", min=0, max=1, step=0.05,
                                   value=1.0,
                                   marks={0: "0 = pooled", 0.5: "0.5",
                                          1: "1 = full index"},
                                   tooltip={"placement": "bottom",
                                            "always_visible": True}),
                    ]),
                    html.Div(style={"minWidth": "260px"}, children=[
                        html.Div("INDEX CAP (± standard deviations)",
                                 className="plabel"),
                        dcc.Slider(id="multi_sdlimit", min=0.25, max=3,
                                   step=0.25, value=1.0,
                                   marks={0.25: "0.25", 1: "1", 3: "3"},
                                   tooltip={"placement": "bottom",
                                            "always_visible": True}),
                    ]),
                    html.Div(style={"alignSelf": "flex-end"}, children=[
                        dcc.Checklist(
                            id="multi_opts",
                            options=[
                                {"label": " Fit λ on the holdout",
                                 "value": "fit"},
                                {"label": " Keep sign priors", "value": "sign"},
                                {"label": " Channel intercepts",
                                 "value": "fe"}],
                            value=["fit", "sign", "fe"],
                            labelStyle={"display": "block",
                                        "fontSize": "12.5px"}),
                    ]),
                ]),
                html.Div("λ is the dial between the two extremes: 0 gives every "
                         "channel the product-level coefficient, 1 gives each "
                         "channel its full index. “Fit λ on the holdout” picks "
                         "it from data the coefficients never saw instead of "
                         "assuming a value. “Channel intercepts” mean-centres "
                         "within channel so a channel that is 30× bigger than "
                         "another does not force the shared coefficients to "
                         "absorb a level difference.",
                         style={"fontSize": "12px", "color": INK3,
                                "marginTop": "10px", "maxWidth": "980px"}),
            ]),
            dcc.Loading(children=[html.Div(id="multi_out")]),
        ]),

        html.Div(id="scr-batch", className="bench-screen",
                 style={"display": "none"}, children=[
            html.Div(style={"display": "flex", "alignItems": "center",
                            "gap": "10px"}, children=[
                html.Div("All models", style={"fontSize": "15px",
                                              "fontWeight": 600}),
                html.Div(id="batch_pills", style={"display": "flex",
                                                  "gap": "8px"}),
                html.Div(style={"flex": 1}),
                html.Button("Reload summary", id="batch_refresh", n_clicks=0,
                            className="btn btn-ghost"),
                html.Button("▶  Run model", id="run_batch_one", n_clicks=0,
                            className="btn btn-ghost"),
                html.Button("Run all combinations", id="batch_run", n_clicks=0,
                            className="btn btn-primary",
                            style={"height": "36px"}),
            ]),
            html.Div("Every product × channel of the loaded datafile, using "
                     "each product's saved config, over one fixed calendar "
                     "window. Running here caches the FULL results — the "
                     "other screens then just filter them, no re-running. "
                     "Click a row to load that slice.",
                     style={"fontSize": "12px", "color": INK3}),
            dcc.Loading(children=[html.Div(className="card", children=[
                dash_table.DataTable(
                    id="batch_table", sort_action="native",
                    filter_action="native", row_selectable="single",
                    page_size=100,
                    style_cell_conditional=[
                        {"if": {"column_id": "grade"},
                         "textAlign": "center", "minWidth": "128px",
                         "width": "128px"}],
                    **TBL_BASE,
                )])]),
            # Slices deliberately NOT modeled. Alex: a slice with no data in
            # the window "would not even show" as a model — but it must not
            # vanish silently either, or nobody notices the data problem.
            html.Div(id="skipped_panel", style={"marginTop": "14px"}),
        ]),

        # ── Export ──
        html.Div(id="scr-export", className="bench-screen",
                 style={"display": "none"}, children=[
            html.Div("Export", style={"fontSize": "15px", "fontWeight": 600}),
            html.Div("One Excel with three tabs — fit statistics (+ grade), "
                     "fit structure (the variable setup, same format as the "
                     "upload template), and weekly due-tos (actual by date). "
                     "Uses the current TARGET and WINDOW from the top bar and "
                     "the override-wins config.",
                     style={"fontSize": "12px", "color": INK3,
                            "margin": "4px 0 12px"}),
            html.Div(className="card card-pad", children=[
                dcc.RadioItems(
                    id="export_mode",
                    options=[{"label": " This slice (current Product × Channel)",
                              "value": "slice"},
                             {"label": " All Product × Channel combinations",
                              "value": "all"}],
                    value="slice",
                    labelStyle={"display": "block", "margin": "4px 0",
                                "fontSize": "13px"}),
                html.Div(style={"display": "flex", "alignItems": "center",
                                "gap": "12px", "marginTop": "10px"}, children=[
                    html.Button("⤓  Download Excel", id="export_btn",
                                n_clicks=0, className="btn btn-primary",
                                style={"height": "36px"}),
                    html.Div(id="export_status",
                             style={"fontSize": "12px", "color": INK3}),
                ]),
                dcc.Download(id="export_download"),
                html.Div("Note: “All” re-runs every slice and can take a few "
                         "seconds.", style={"fontSize": "11px",
                                            "color": MUTED, "marginTop": "8px"}),
            ]),
        ]),

        # ── Definitions & Methods ──
        html.Div(id="scr-defs", className="bench-screen",
                 style={"display": "none", "fontSize": "13px",
                        "lineHeight": "1.55", "maxWidth": "920px"}, children=[
            html.Div("Definitions & Methods", style={"fontSize": "15px",
                                                     "fontWeight": 600}),
            html.Div("What the numbers mean and how a model is built. "
                     "Definitions track the engine's actual behavior.",
                     style={"fontSize": "12px", "color": INK3,
                            "margin": "4px 0 14px"}),

            _defs_section(
                "Fit metrics",
                _defn("R² / adjusted R²", "share of weekly variance the model "
                      "explains; adjusted R² penalizes adding predictors. Shown "
                      "together so a high R² from over-fitting is visible."),
                _defn("In-sample MAPE", "mean absolute % error on the training "
                      "window, over non-zero-actual weeks only."),
                _defn("Holdout MAPE", "out-of-sample error on a reserved tail. "
                      "The last up to 13 weeks are ALWAYS reserved (independent "
                      "of the window, as long as ≥5 training weeks remain); a "
                      "validation model — same structure, selected without "
                      "seeing the tail — is scored there, and that number is "
                      "kept as the reported holdout."),
                _defn("Durbin–Watson", "residual autocorrelation check (~2 ≈ no "
                      "autocorrelation)."),
                _defn("T-stat", "from an unconstrained OLS as an inference "
                      "reference; |t| ≥ 2 ≈ significant at 5%."),
                _defn("VIF", "variance-inflation factor (collinearity), computed "
                      f"on the final in-model variables. > {int(VIF_WARN)} flags "
                      "redundancy — a non-blocking warning; forced variables can "
                      "be high by design.")),

            _defs_section(
                "Model grade (Model Runs tab)",
                _defn("Good", "holdout MAPE ≤ 10%"),
                _defn("Moderate", f"> 10% and ≤ {int(FLAG_THRESHOLD)}%"),
                _defn("Bad", f"> {int(FLAG_THRESHOLD)}%"),
                _defn("No holdout", "the slice has no holdout period to score."),
                html.Div("These cutoffs are a team convention (reusing the "
                         "review-flag threshold) — confirm with the client.",
                         style={"fontSize": "12px", "color": MUTED,
                                "marginTop": "2px"})),

            _defs_section(
                "Contribution vs due-to — not the same number",
                _defn("Contribution", "a ", html.B("level"), ": how much "
                      "volume a driver accounts for IN a period (Σ βᵢ·xᵢ over "
                      "that period's weeks). The additive decomposition of the "
                      "fitted line — intercept as its own line plus "
                      "coefficient × value per driver — so intercept + "
                      "Σ(βᵢ·xᵢ) reproduces the fitted value exactly each week."),
                _defn("Due-to", "a ", html.B("change"), ": how much of the "
                      "movement between TWO periods a driver explains "
                      "(its contribution in period B minus its contribution "
                      "in period A). This is the “how is the change in my "
                      "macro variables affecting my sales year over year” "
                      "number. The due-tos sum exactly to the change in "
                      "modeled volume between the two periods."),
                _defn("Periods of unequal length", "compared as ", html.B(
                    "per-week averages"), " automatically — otherwise a "
                    "52-week period always looks bigger than a 4-week one for "
                    "reasons that have nothing to do with the drivers."),
                _defn("Where to find each", "the Contributions screen has a "
                      "toggle: Contribution (level in a period) or Due-to "
                      "(change between the two selected periods). The export "
                      "carries both — weekly_due_tos (per-week levels, with "
                      "period columns) and due_to_change (period-over-period "
                      "changes)."),
                _defn("Avg weekly due-to", "the mean of a driver's weekly "
                      "contribution. For ", html.B("media"), " (any variable "
                      "with an adstock decay) the average is taken ", html.B(
                          "only over weeks with real execution"), " (non-zero "
                      "raw spend/impressions) — never diluted by the zero weeks "
                      "of a flight that started mid-window. Non-media drivers "
                      "average over all weeks. (This is the metric Alex asked to "
                      "have defined here.)"),
                _defn("Due-to by model year / YoY", "signed sums per model year "
                      "with the year-over-year change, to see what drove the "
                      "change.")),

            _defs_section(
                "Variables — roles & configuration",
                _defn("Expected sign (guardrail)", "a positive-signed variable "
                      "can never enter with a negative coefficient. Custom "
                      "bounds override the sign."),
                _defn("Role — auto / force / exclude", html.B("auto"), ": the "
                      "model decides (selection). ", html.B("force"), ": "
                      "client-mandated — kept in every run, bypasses selection "
                      "and VIF pruning. ", html.B("exclude"), ": never enters."),
                _defn("Adstock", "applied to ", html.B("any variable with an "
                      "adstock decay set"), " (typically media, but the engine "
                      "doesn't restrict by family): normalized carry-over "
                      "aₜ = (1−decay)·xₜ + decay·aₜ₋₁; the adstocked total ≈ the "
                      "raw total (never inflated). Search ≈ 0.2, TV/CTV ≈ 0.7, "
                      "default 0.5."),
                _defn("Bounds (lo / hi)", "optional hard limits on a "
                      "coefficient, written in ", html.B("original units"),
                      "; lo must be strictly < hi. If the variable has a "
                      "scale, the bound is transformed with it, so setting a "
                      "scale never silently tightens a bound."),
                _defn("Scale (÷)", "divide a variable by this before "
                      "modeling, so its coefficient reads “per 1,000 "
                      "impressions” instead of “per impression”. This is a ",
                      html.B("units change only"), ": the coefficient shrinks "
                      "by exactly the same factor the data does, so every "
                      "contribution, due-to and fitted value is identical to "
                      "the unscaled run (verified to ~1e-10)."),
                _defn("Impact / SD (comparable coefficient)", "coefficient × "
                      "the variable's standard deviation — the volume moved "
                      "by a one-SD change. Raw coefficients ", html.B("cannot"),
                      " be compared across variables because the inputs sit "
                      "on different supports: a savings rate moves between 2 "
                      "and 8, impressions between 0 and millions, so their "
                      "coefficients differ by orders of magnitude for reasons "
                      "that have nothing to do with importance. This column "
                      "puts every driver in the same unit (volume), so it is "
                      "the one you can legitimately rank."),
                _defn("Two-tier config", "a Product default (channel = ALL) "
                      "applies to every channel; a Product × Channel override, "
                      "when present, wins for that channel — for the ", html.B(
                          "Batch, Export and CLI"), ". A ", html.B("Variables-"
                      "tab Run is WYSIWYG"), ": it runs exactly the config "
                      "you're editing (default or override), which can differ "
                      "from what the Batch would resolve.")),

            _defs_section(
                "Media curves — adstock & saturation",
                _defn("Adstock (carry-over)", "aₜ = (1−d)·xₜ + d·aₜ₋₁. The "
                      "(1−d) share lands this week, the rest carries forward. "
                      "Normalized so the adstocked total ≈ the raw total. "
                      "Search ≈ 0.2, TV/CTV ≈ 0.7, default 0.5."),
                _defn("Hill saturation", "H(x) = xˢ / (xˢ + kˢ), with "
                      "k = midpoint × peak execution. The ", html.B("Hill "
                      "equation"), " — A. V. Hill, 1910, oxygen binding to "
                      "haemoglobin, and since then the standard dose–response "
                      "curve in pharmacology. It is also the saturation "
                      "function in Meta's Robyn and Google's Meridian. "
                      "sat mid = where the response reaches half its maximum, "
                      "as a fraction of peak execution; sat slope > 1 gives an "
                      "S-curve, < 1 diminishes from the origin. Leave both "
                      "blank for a linear variable (the default)."),
                _defn("Coefficient on a saturated variable", "because H ends "
                      "up in [0,1), the coefficient is the ", html.B("maximum "
                      "weekly volume that variable can deliver"), ". "
                      "Contributions stay additive, so the due-to "
                      "decomposition is unchanged."),
                _defn("Optimize media curves", "searches decay × midpoint × "
                      "slope per media variable. Scored by rolling-origin "
                      "cross-validation on folds that exclude the reported "
                      "holdout entirely, and a candidate must not be worse in "
                      "ANY fold. So the holdout stays an honest check on "
                      "whether the optimization helped. It writes candidates "
                      "into the config table — it does not save or run."),
                _defn("Read this before trusting it", "on this dataset the "
                      "optimizer improves its cross-validated score every "
                      "time it fires, but the untouched holdout improves only "
                      "about half the time (mean +0.85 pp, i.e. slightly "
                      "worse). Media is a small share of these models and 104 "
                      "weeks is thin for two extra non-linear parameters per "
                      "variable. Treat the curves as candidates: run, compare "
                      "the holdout, and revert if it worsened. Details in "
                      "docs/Media_Curves_Math.md."),
                _defn("Response curve chart", "on Diagnostics. A flattening "
                      "curve is a diminishing return and the dotted line is "
                      "the half-saturation point; a straight line means no "
                      "saturation was fitted. Where the dots sit is the "
                      "current operating point — out on the flat means extra "
                      "spend is buying little.")),

            _defs_section(
                "High Level — Total Brand / Total Channel",
                _defn("What is summed", "the OUTPUT of the existing models, "
                      "not a new model. Each Product × Channel keeps its own "
                      "coefficients; actual, fitted and every driver's weekly "
                      "contribution are added up across the group. Because "
                      "each slice's decomposition is exact, the total's is "
                      "too."),
                _defn("Why a total fits better than its parts", "independent "
                      "errors partly cancel when series are added, so a "
                      "rolled-up R² and MAPE will flatter the portfolio. That "
                      "is a real property of the aggregate, not a trick — but "
                      "it is why the member grid sits underneath every total. "
                      "A good total can still hide a bad slice."),
                _defn("WMAPE", "error volume ÷ actual volume. At the total "
                      "level this is the number that answers “how far off is "
                      "my volume”; plain MAPE averages week-level percentages "
                      "and lets a small week count as much as a big one."),
                _defn("Total vs pooled", "a Total view is ten models added "
                      "together. A ", html.B("pooled"), " model (Multi-Level "
                      "screen) is genuinely ONE model over the stacked rows. "
                      "If the question is “what does my portfolio do”, use "
                      "this screen; if it is “what is one coefficient for this "
                      "brand”, use that one."),
                _defn("Partial coverage warning", "a slice that starts "
                      "part-way through the window contributes nothing before "
                      "it starts, so the left edge of a total can dip for "
                      "reasons that have nothing to do with the business. The "
                      "screen says so when it happens.")),

            _defs_section(
                "Saturation — weekly transforms and response curves",
                _defn("The weekly panel", "raw execution (bars), decayed "
                      "execution after adstock (solid), and decayed + "
                      "saturated after Hill (dotted). This is literally the "
                      "column that entered the regression, not a "
                      "re-derivation. The target is overlaid, rescaled, so the "
                      "relationship is visible."),
                _defn("Why the saturated series is rescaled", "Hill returns "
                      "values in [0, 1) by construction, so on a shared axis "
                      "it would be a flat line at the bottom of the chart. It "
                      "is rescaled to the decayed peak for shape comparison "
                      "and shown again on its own true 0–1 axis underneath. "
                      "Correlations are computed on the unscaled series — a "
                      "positive linear rescale cannot change Pearson r."),
                _defn("The three correlations", "|r| with the target for raw, "
                      "decayed and saturated. This is the evidence behind a "
                      "decay or midpoint choice. A higher |r| does not by "
                      "itself guarantee a better model, because the variable "
                      "competes with everything else in the regression — hit "
                      "Run to see the effect on holdout MAPE."),
                _defn("The sliders", "drag decay, midpoint or slope and every "
                      "chart redraws from the cached slice. No model is "
                      "re-run and nothing is saved: the coefficient is held at "
                      "its fitted value while the shape of the input changes. "
                      "“Apply to config” writes the three numbers into the "
                      "product's variable config — candidates, not results, "
                      "until you Run."),
                _defn("Execution vs ROI curve", "Alex's SatCurve over the "
                      "latest 52 weeks. x is % of current spend (100% = "
                      "today), the shaded area is sales driven, and the two "
                      "lines are average and marginal return. Where they cross "
                      "is the optimal spend level — past it, the next dollar "
                      "returns less than the average dollar already spent."),
                _defn("Spend, price and margin", "spend = Σ of the media "
                      "column over the latest 52 weeks (these columns ARE "
                      "spend); price = mean Price per Volume over the same "
                      "weeks; margin is not in the data and defaults to 30%. "
                      "All three are editable. Margin moves the ROI axis but ",
                      html.B("cannot"), " move the optimal point — it scales "
                      "average and marginal return by the same factor."),
                _defn("“No crossover” verdict", "with a slope below 1 the "
                      "curve is concave from the origin and marginal return "
                      "sits under average return at every spend level, so the "
                      "lines never meet. Reporting “optimal = 300%” there "
                      "would be a graphing artifact dressed up as a "
                      "recommendation, so it is flagged as inconclusive "
                      "instead.")),

            _defs_section(
                "Multi-Level — unpooled, pooled, hierarchical",
                _defn("Unpooled", "one independent model per Product × "
                      "Channel — what every other screen shows. Most "
                      "flexibility, fewest degrees of freedom (104 weeks "
                      "each), coefficients free to disagree between channels "
                      "for no reason but noise."),
                _defn("Pooled", "ONE model per product, every channel's weeks "
                      "stacked — roughly 700–900 rows against the same "
                      "predictor count. One coefficient per predictor, shared "
                      "across channels. Valid here because these are models "
                      "OF a time series, not time-series models: adstock and "
                      "saturation are applied within each channel before the "
                      "rows are stacked, so each row is independent."),
                _defn("National predictors", "predictors carrying identical "
                      "values in every channel (macro, most media). Pooled, "
                      "one coefficient would produce that contribution once "
                      "per channel — nine times the effect it can have had. "
                      "Each is multiplied by its channel's share of product "
                      "volume so the pieces sum back to the national figure."),
                _defn("A false positive worth knowing about", "the quick "
                      "screen “equal sums by channel ⇒ national” flags "
                      "Seasonality_Index, which is a per-channel index "
                      "normalised to average 1.0 — every channel sums to "
                      "exactly 104 over a 104-week window while the weekly "
                      "values differ a lot. The engine compares weekly values "
                      "instead and leaves it channel-specific; the screen "
                      "lists any such near-miss."),
                _defn("Channel intercepts", "each channel gets its own base "
                      "level via mean-centring within channel (Arko's "
                      "“mean centered models”). Without it, a single intercept "
                      "has to sit between channels that differ 30× in size, "
                      "and the shared coefficients waste themselves absorbing "
                      "a level gap. Mathematically identical to one dummy per "
                      "channel, without putting dummies through variable "
                      "selection."),
                _defn("Hierarchical", "each channel's coefficient = the pooled "
                      "coefficient × an index from that channel's own "
                      "unconstrained fit (index = its coefficient ÷ the mean "
                      "across channels, so the indices average to 1). This is "
                      "the arithmetic in Alex's ",
                      html.B("Hierarchical Modeling Explanation.xlsx"),
                      ", reproduced cell for cell in test_multilevel.py."),
                _defn("Not Bayesian, deliberately", "textbook hierarchical "
                      "models get partial pooling from a prior and a sampler. "
                      "Alex was explicit that this stays an optimised "
                      "least-squares engine, so the shrinkage is explicit "
                      "instead: an index, a cap, and a λ dial."),
                _defn("λ (shrink)", "0 gives every channel the product-level "
                      "coefficient (= pooled); 1 gives each channel its full "
                      "index (= Alex's method). In between is literal partial "
                      "pooling. “Fit λ on the holdout” chooses it from data "
                      "the coefficients never saw rather than assuming a "
                      "value, and the curve of λ against holdout error is "
                      "shown — a flat curve means the channels do not "
                      "measurably differ."),
                _defn("Index cap", "clips each index to 1 ± k standard "
                      "deviations, from Alex's sheet. It is what keeps this "
                      "partial pooling rather than a relabelled unpooled "
                      "model: a channel whose own fit produced a wild "
                      "coefficient is pulled back toward the product effect."),
                _defn("Sign guard", "the index is clipped at 0 so a channel's "
                      "coefficient can never flip sign against the pooled "
                      "one. On real data the raw index runs from −7 to +12; "
                      "a negative index would hand a channel media that "
                      "destroys sales, silently undoing the sign priors set in "
                      "the variable config. Not in Alex's sheet — added "
                      "because the sheet's example coefficients are far better "
                      "behaved than real ones."),
                _defn("Unstable index warning", "the index divides by the mean "
                      "coefficient across channels. When that mean is small "
                      "relative to its spread, the ratio is unstable and the "
                      "index means little. The screen reports how many "
                      "predictors are in that state — those are exactly the "
                      "ones the cap and λ exist to tame."),
                _defn("Why WMAPE here", "all three levels are scored on the "
                      "same held-out last 13 weeks per channel. Plain MAPE "
                      "across channels of very different size lets the "
                      "smallest channel decide the winner; WMAPE weights each "
                      "week by the volume at stake.")),

            _defs_section(
                "Variable selection — strictness & family caps",
                _defn("Strictness", "Punitive (p<0.01, VIF>5 pruned) admits "
                      "only strongly significant variables; Lax (p<0.15, "
                      "VIF>20) lets far more in; Balanced (p<0.05, VIF>10) is "
                      "the default and reproduces every model built before "
                      "this control existed."),
                _defn("Max per family", "caps how many variables of a family "
                      "may enter — “only two to three trade, but all media”. "
                      "Blank means no limit. The cap is enforced ", html.B(
                          "during"), " selection, not by trimming afterwards, "
                      "so a capped family spends its budget on its most "
                      "significant members while the other families still "
                      "fill normally."),
                _defn("Caps and forced variables", "a forced variable is "
                      "always kept and counts toward its family's budget, so "
                      "a cap of 2 with 2 forced Trade variables admits no "
                      "further Trade. Forcing wins over the cap, exactly as "
                      "it wins over VIF pruning."),
                _defn("Results are never mixed", "the cached results are keyed "
                      "by target, window AND these hyperparameters. Change "
                      "any of them and the screens ask for a fresh batch "
                      "rather than showing you a blend of two settings.")),

            _defs_section(
                "Window, holdout & method",
                _defn("Window (fixed calendar range)", "the window is an ",
                      html.B("absolute date range"), " — it ends at the "
                      "latest week present in the DATASET and runs back 52 / "
                      "104 / 156 weeks — and is ", html.B("identical for "
                      "every Product × Channel"), ". The data is dependent on "
                      "the window, not the window on the data. Previously the "
                      "window was the last N rows each slice happened to have, "
                      "so a product delisted in mid-2023 was still modeled — "
                      "on 18-month-old history — and its due-tos leaked into "
                      "the exports as a disjoint chunk."),
                _defn("Not modeled", "a slice with fewer than 52 weeks of data "
                      "or 26 selling weeks inside the window is deliberately "
                      "not modeled. It appears in the “Not modeled” list on "
                      "Model Runs with the counts and the reason, and in the "
                      "export's errors sheet — never as a model."),
                _defn("Holdout", "the last up to 13 weeks ", html.B("inside "
                      "the window"), " are reserved for validation, so there "
                      "is normally an out-of-sample metric (exceptions: a very "
                      "short history, which needs ≥5 training weeks, or "
                      "holdout explicitly set to 0)."),
                _defn("Run once, filter many", "a Model Runs batch computes "
                      "and caches EVERY slice's results; the Diagnostics, "
                      "Contributions and Export screens then filter that "
                      "cache. Changing the Product or Channel filter does not "
                      "re-run anything, and you cannot end up comparing a "
                      "slice modeled on one window against slices modeled on "
                      "another. Re-tuning a slice on Variables and hitting Run "
                      "refreshes just that slice in the cache."),
                _defn("How a model is built", "the ", html.B("validation"),
                      " model selects the structure — VIF prune to cut "
                      "collinearity, then forward stepwise — on the pre-tail "
                      "window (no leakage) and is scored on the tail; the ",
                      html.B("reported"), " model refits that ", html.B("same "
                      "structure"), " on the full window via ", html.B("sign/"
                      "box-constrained least-squares"), " (a plain ridge can't "
                      "honor the guardrails). Verified numerically equivalent to "
                      "the sponsor's curve_fit approach.")),
        ]),
    ]),
])


# ---------------- callbacks ----------------

@app.callback(Output("active_tab", "data"),
              [Input(f"nav-{k}", "n_clicks") for k, _ in SCREENS],
              prevent_initial_call=True)
def _nav(*_):
    try:
        trig = ctx.triggered_id
    except Exception:          # outside a live callback (tests)
        trig = None
    return (trig or "nav-diag").replace("nav-", "")


@app.callback(Output("active_tab", "data", allow_duplicate=True),
              Input("run_vars", "n_clicks"), Input("run_batch_one", "n_clicks"),
              prevent_initial_call=True)
def _redirect_after_run(_v, _b):
    # Hitting Run (Variables tuning or a Batch single-slice run) jumps to
    # Diagnostics so you land on the model you just ran. (Alex: slide 2)
    return "diag"


@app.callback(
    [Output(f"scr-{k}", "style") for k, _ in SCREENS]
    + [Output(f"nav-{k}", "className") for k, _ in SCREENS],
    Input("active_tab", "data"))
def _render_tabs(active):
    active = active or "diag"
    styles = [{} if k == active else {"display": "none"} for k, _ in SCREENS]
    classes = ["active" if k == active else "" for k, _ in SCREENS]
    return styles + classes


@app.callback(Output("datapath", "data"), Output("load_status", "children"),
              Output("dataset_file", "children"),
              Input("load_data", "n_clicks"),
              State("datafile_input", "value"))
def _load_data(_n, path):
    path = (path or "").strip().strip('"')
    if not path:
        return None, "Enter a path to an .xlsx datafile.", "—"
    if not os.path.exists(path):
        return None, f"File not found: {path}", "—"
    try:
        sheets = detect_product_sheets(path)
    except Exception as e:
        return None, f"Could not read workbook: {e}", "—"
    if not sheets:
        return None, ("No product sheets found (need Geography + Time "
                      "columns, or sheets named 'Brand *')."), "—"
    n_ch = len(channels_for(path, sheets[0]))
    n_wk = len(load_sheet(path, sheets[0])) // max(n_ch, 1)
    meta = f"{len(sheets)} products · ~{n_ch} channels · ~{n_wk} weeks each"
    return ({"path": path, "sheets": list(sheets)}, meta,
            os.path.basename(path))


@app.callback(Output("brand", "options"), Output("brand", "value"),
              Input("datapath", "data"))
def _brands(dp):
    if not dp:
        return [], None
    return dp["sheets"], dp["sheets"][0]


@app.callback(Output("channel", "options"), Output("channel", "value"),
              Input("brand", "value"), Input("batch_sel", "data"),
              State("datapath", "data"))
def _channels(brand, sel, dp):
    if not dp or not brand:
        return [], None
    ch = channels_for(dp["path"], brand)
    if sel and sel.get("brand") == brand and sel.get("channel") in ch:
        return ch, sel["channel"]
    return ch, (ch[0] if ch else None)


@app.callback(Output("target", "options"), Output("target", "value"),
              Input("brand", "value"), State("datapath", "data"),
              State("target", "value"))
def _targets(brand, dp, current):
    if not dp or not brand:
        return [], None
    opts = target_choices(dp["path"], brand)
    if current in opts:
        return opts, current
    return opts, (opts[0] if opts else None)


@app.callback(Output("window", "options"), Output("window", "value"),
              Input("datapath", "data"))
def _window(dp):
    return [{"label": f"{w} wks", "value": w} for w in WINDOW_CHOICES], 104


def _fmt_coef(c) -> str:
    if c is None or (isinstance(c, float) and np.isnan(c)):
        return "—"
    a = abs(c)
    if a != 0 and a < 0.01:
        return f"{c:.2e}"
    if a < 1000:
        return f"{c:,.2f}"
    return f"{c:,.0f}"


_CTX_BLANK = {"cur_coef": "—", "cur_impact": "—", "cur_dueto": "—"}


def _last_run_ctx_map(last_run, brand, channel, target, period) -> dict:
    """{variable -> {cur_coef, cur_dueto}} for the latest run — current
    coefficient and current due-to for the selected period (Y1 / Y2 / … /
    Total). Shown ONLY when the run matches the current Product × Channel ×
    Target, so switching channel/target blanks the context until you re-run
    (never surface a different slice's numbers). (Alex: slide 2, V1.)"""
    if not last_run or not all(
            str(last_run.get(k)) == str(v) for k, v in
            (("brand", brand), ("channel", channel), ("target", target))):
        return {}
    coefs = last_run.get("coef", {}) or {}
    impact = last_run.get("beta_std", {}) or {}
    by_period = last_run.get("dueto_by_period", {}) or {}
    duetos = by_period.get(period) or by_period.get("Total", {}) or {}
    out = {}
    for var in set(coefs) | set(duetos) | set(impact):
        out[var] = {"cur_coef": _fmt_coef(coefs.get(var)),
                    # comparable across variables; the raw coef is not
                    "cur_impact": fmt_compact(impact.get(var)),
                    "cur_dueto": fmt_compact(duetos.get(var))}
    return out


def _apply_ctx(rows, ctx_map):
    for row in rows:
        row.update(ctx_map.get(str(row.get("variable")), _CTX_BLANK))
    return rows


@app.callback(Output("model_settings_note", "children"),
              Output("vars_window_note", "children"),
              Input("target", "value"), Input("window", "value"),
              Input("datapath", "data"))
def _settings_note(target, window, dp):
    """Echo the modeling settings read-only wherever results are shown, with
    the actual calendar range — so "these are Volume Sales due-tos on the
    Jan 2024–Dec 2025 window" is always on screen without offering a filter
    that would silently invalidate what's displayed."""
    if not (target and window):
        return "", ""
    rng = window_caption(dp["path"], window) if dp else f"{window} wks"
    top = [html.Span(str(target), style={"fontWeight": 600, "color": INK2}),
           html.Span(f"  ·  {rng}", style={"color": MUTED}),
           html.Div("modeling settings — change on Variables",
                    style={"fontSize": "10.5px", "color": MUTED})]
    return top, f"Window: {rng} (same range for every Product × Channel)"


@app.callback(Output("strictness_note", "children"),
              Input("strictness", "value"), Input("family_caps", "data"))
def _strictness_note(strictness, cap_rows):
    t = tuning_of(strictness, cap_rows)
    caps = t["max_per_family"]
    txt = f"p<{t['p_enter']} · VIF>{t['vif_threshold']:.0f} pruned"
    if caps:
        txt += " · caps " + ", ".join(f"{k}≤{v}" for k, v in sorted(caps.items()))
    else:
        txt += " · no family caps"
    return txt + " — re-run to apply"


@app.callback(Output("dueto_period", "options"),
              Output("dueto_period", "value"),
              Input("last_run", "data"), State("dueto_period", "value"))
def _dueto_period_opts(last_run, current):
    periods = (last_run or {}).get("periods") or ["Total"]
    opts = [{"label": p, "value": p} for p in periods]
    val = (current if current in periods
           else "Total" if "Total" in periods else periods[-1])
    return opts, val


@app.callback(Output("cfg_table", "filter_query"),
              Input("hide_excluded", "value"))
def _hide_excluded(val):
    # native filter only hides rows from view — cfg_table.data (what Save
    # reads) stays complete, so excluded rows are never dropped on save.
    return '{role} != "exclude"' if val and "hide" in val else ""


@app.callback(Output("config_scope", "options"),
              Output("config_scope", "value"),
              Input("channel", "value"), State("config_scope", "value"))
def _scope_opts(channel, current):
    opts = [{"label": "Product default (all channels)", "value": "default"}]
    if channel:
        opts.append({"label": f"This channel — {channel}", "value": "channel"})
    # channel scope is only valid when a channel is selected — otherwise force
    # back to default so Save/Run can't silently write the default (#4).
    val = "channel" if (current == "channel" and channel) else "default"
    return opts, val


@app.callback(Output("cfg_status", "children", allow_duplicate=True),
              Output("config_rev", "data", allow_duplicate=True),
              Input("cfg_reset", "n_clicks"),
              State("config_scope", "value"), State("channel", "value"),
              State("brand", "value"), State("datapath", "data"),
              prevent_initial_call=True)
def _reset_override(_n, scope, channel, brand, dp):
    # Delete this channel's override so it inherits the Product default again.
    # config_rev is bumped ONLY on an actual delete, so the table/status readers
    # re-read disk after the file is gone (no race). (#3)
    if not dp or not brand:
        return "Load a datafile first.", no_update
    if scope != "channel" or not channel:
        return ("Reset to default applies only when editing a channel override.",
                no_update)
    ov = override_cfg_path(dp["path"], brand, channel)
    if not os.path.exists(ov):
        return (f"{brand} × {channel} has no override — already inherits default.",
                no_update)
    try:
        with _CFG_WRITE_LOCK:
            os.remove(ov)
    except OSError as e:
        return f"Could not remove override: {e}", no_update
    return (f"{brand} × {channel} override removed — now inherits the Product "
            "default (and future default changes apply).", time.time())


@app.callback(Output("config_rev", "data", allow_duplicate=True),
              Output("cfg_status", "children", allow_duplicate=True),
              Input("cfg_regen", "n_clicks"),
              State("config_scope", "value"), State("channel", "value"),
              State("brand", "value"), State("datapath", "data"),
              prevent_initial_call=True)
def _regen_cfg(_n, scope, channel, brand, dp):
    # Regenerate defaults into the CURRENT scope's file (never clobber the
    # Product default while editing a channel override), then bump config_rev
    # so the table reloads the freshly written config deterministically. (#1)
    if not dp or not brand:
        return no_update, "Load a datafile first."
    fresh = fresh_default_df(dp["path"], brand)
    target = save_cfg_path(dp["path"], brand, scope, channel)
    with _CFG_WRITE_LOCK:
        fresh.to_csv(target, index=False)
    where = (f"{brand} × {channel} override" if scope == "channel" and channel
             else f"{brand} default (all channels)")
    return time.time(), f"Regenerated {where} from data — {len(fresh)} variables."


@app.callback(Output("cfg_scope_status", "children"),
              Input("config_scope", "value"), Input("channel", "value"),
              Input("brand", "value"), Input("config_rev", "data"),
              Input("last_run", "data"), State("datapath", "data"))
def _scope_status(scope, channel, brand, _rev, _lr, dp):
    if not dp or not brand:
        return ""
    if scope == "channel" and channel:
        if os.path.exists(override_cfg_path(dp["path"], brand, channel)):
            return (f"Editing {brand} × {channel} override — wins over the "
                    "default for this channel at runtime.")
        return (f"Editing {brand} × {channel} — no override yet (seeded from "
                "the Product default). Save or Run creates the override.")
    return (f"Editing {brand} Product default — used by every channel that "
            "has no override.")


@app.callback(Output("cfg_table", "data"),
              Input("brand", "value"), Input("last_run", "data"),
              Input("dueto_period", "value"), Input("channel", "value"),
              Input("target", "value"), Input("config_scope", "value"),
              Input("config_rev", "data"),
              State("datapath", "data"), State("cfg_table", "data"))
def _load_cfg(brand, last_run, period, channel, target, scope, _rev,
              dp, current):
    if not dp or not brand:
        return []
    try:
        trig_ids = {t["prop_id"].split(".")[0] for t in (ctx.triggered or [])}
    except Exception:
        trig_ids = set()
    ctx_map = _last_run_ctx_map(last_run, brand, channel, target, period)
    # Reload the scoped config from disk when the product or config scope
    # changed, when a write happened (config_rev — save / regenerate / reset),
    # on first paint, or (channel-scope) when the channel changed. For
    # period/target/run/last_run (and channel changes while editing the
    # default) keep the on-screen rows so unsaved edits survive; just re-apply
    # the (slice-gated) context columns.
    reload = ("brand" in trig_ids or "config_scope" in trig_ids
              or "config_rev" in trig_ids or not current
              or ("channel" in trig_ids and scope == "channel"))
    if not reload:
        return _apply_ctx(current, ctx_map)
    cfg, _state = load_cfg_for_edit(dp["path"], brand, scope, channel)
    # configs written before `scale` existed have no such column — default it
    # to 1 (no scaling) so the table shows a real value instead of a blank the
    # analyst has to guess about
    if "scale" not in cfg.columns:
        cfg["scale"] = 1.0
    cfg["scale"] = pd.to_numeric(cfg["scale"], errors="coerce").fillna(1.0)
    for c in ("sat_midpoint", "sat_slope"):
        if c not in cfg.columns:
            cfg[c] = np.nan
    return _apply_ctx(cfg[CFG_COLS].to_dict("records"), ctx_map)


def _validate_cfg_rows(rows):
    """Coerce numerics + check bounds WITHOUT writing. Returns (ok, cfg_df,
    bad_list). lsq_linear needs lo STRICTLY < hi, so `lo >= hi` is rejected."""
    cfg = pd.DataFrame(rows, columns=CFG_COLS)
    for c in ("adstock_decay", "coef_lower", "coef_upper", "scale",
              "sat_midpoint", "sat_slope"):
        cfg[c] = pd.to_numeric(cfg[c], errors="coerce")
    # a blank / 0 / negative scale would divide by zero or flip the data's
    # sign — normalise it away before it can reach the engine
    cfg["scale"] = cfg["scale"].where(cfg["scale"] > 0).fillna(1.0)
    # saturation needs a midpoint in (0,1] and a positive slope, and needs
    # BOTH: half a curve is not a curve, so an incomplete pair is cleared to
    # "no saturation" rather than being guessed at
    cfg["sat_midpoint"] = cfg["sat_midpoint"].where(
        (cfg["sat_midpoint"] > 0) & (cfg["sat_midpoint"] <= 1))
    cfg["sat_slope"] = cfg["sat_slope"].where(cfg["sat_slope"] > 0)
    incomplete = cfg["sat_midpoint"].isna() | cfg["sat_slope"].isna()
    cfg.loc[incomplete, ["sat_midpoint", "sat_slope"]] = np.nan
    bad = cfg[(cfg.coef_lower.notna()) & (cfg.coef_upper.notna())
              & (cfg.coef_lower >= cfg.coef_upper)]
    return (len(bad) == 0), cfg, list(bad.variable)


def _persist_cfg_rows(rows, target_path):
    """Validate + write a config to `target_path` (Product default or a
    Product×Channel override). Returns (ok, cfg_df, bad_list). Shared by Save
    and by a Variables-tab Run (auto-persist so 'modify variables and hit run'
    works — Alex, slide 2). Context/helper columns are dropped by CFG_COLS."""
    ok, cfg, bad = _validate_cfg_rows(rows)
    if not ok:
        return False, cfg, bad
    with _CFG_WRITE_LOCK:            # serialize with template-upload commits
        cfg.to_csv(target_path, index=False)
    return True, cfg, []


@app.callback(Output("cfg_status", "children"), Output("config_rev", "data"),
              Input("cfg_save", "n_clicks"),
              State("cfg_table", "data"), State("brand", "value"),
              State("config_scope", "value"), State("channel", "value"),
              State("datapath", "data"), prevent_initial_call=True)
def _save_cfg(_n, rows, brand, scope, channel, dp):
    if not dp or not brand:
        return "Load a datafile first.", no_update
    target = save_cfg_path(dp["path"], brand, scope, channel)
    ok, cfg, bad = _persist_cfg_rows(rows, target)
    if not ok:
        return f"NOT saved: bound lo must be < bound hi for {bad}", no_update
    where = (f"{brand} × {channel} override" if scope == "channel" and channel
             else f"{brand} default (all channels)")
    return (f"Saved to {where} — {len(cfg)} variables, "
            f"{int((cfg.role == 'force').sum())} forced, "
            f"{int((cfg.role == 'exclude').sum())} excluded. "
            "Next run uses it.", time.time())


@app.callback(Output("cfg_table", "data", allow_duplicate=True),
              Output("opt_status", "children"),
              Input("opt_curves", "n_clicks"),
              State("cfg_table", "data"), State("brand", "value"),
              State("channel", "value"), State("target", "value"),
              State("window", "value"), State("datapath", "data"),
              State("strictness", "value"), State("family_caps", "data"),
              prevent_initial_call=True)
def _optimize_curves(_n, cfg_rows, brand, channel, target, window, dp,
                     strictness, cap_rows):
    """Search adstock decay + Hill saturation for this slice's media
    variables and write the winners into the on-screen config table.

    Deliberately does NOT save or run: the analyst sees what the search chose,
    can override any of it, and then hits Run. An optimizer that silently
    rewrote the config and re-ran would make it impossible to tell which
    numbers came from the data and which from the search.
    """
    if not (dp and brand and channel and target and window):
        return no_update, "Load data and pick a product, channel and target."
    path = dp["path"]
    try:
        tuning = tuning_of(strictness, cap_rows)
        cfg_df = (pd.DataFrame(cfg_rows, columns=CFG_COLS) if cfg_rows
                  else pd.read_csv(resolve_cfg_path(path, brand, channel)))
        base_cfg = cp.ModelConfig(target=target, model_weeks=int(window),
                                  window_end=dataset_anchor(path),
                                  variable_config=cfg_df, **tuning)
        r = cp.run_slice("", brand, channel, config=base_cfg,
                         df=load_sheet(path, brand))
        res = cp.optimize_media_curves(r["df"], r["specs"], target,
                                       selected=r["selected"])
    except cp.InsufficientWindowData as e:
        return no_update, f"Not optimized — {e}"
    except Exception as e:
        import traceback
        traceback.print_exc()
        return no_update, f"Optimization failed: {type(e).__name__}: {e}"

    if res.get("note"):
        return no_update, f"Not optimized — {res['note']}."
    changed = [p for p in res["per_variable"] if p["changed"]]
    rows = [dict(x) for x in (cfg_rows or cfg_df.to_dict("records"))]
    for row in rows:
        st = res["params"].get(str(row.get("variable")))
        if st:
            row["adstock_decay"] = st["adstock_decay"]
            row["sat_midpoint"] = st["sat_midpoint"]
            row["sat_slope"] = st["sat_slope"]
    folds = res.get("folds", [])
    detail = " · ".join(
        f"{p['variable'].split('_')[-2] if '_' in p['variable'] else p['variable']}: "
        f"decay {p['adstock_decay']}"
        + (f", Hill {p['sat_midpoint']}/{p['sat_slope']}"
           if p["sat_midpoint"] is not None else ", linear")
        for p in changed) or "no curve beat what you already had"
    msg = [
        html.B(f"{len(changed)} of {len(res['per_variable'])} media curves "
               f"changed. "),
        f"Cross-validated MAPE {res['baseline_inner_mape']:.2f}% → "
        f"{res['optimized_inner_mape']:.2f}% "
        f"({len(folds)} rolling folds, {res['n_fits']} fits). ",
        html.Span(detail, style={"color": INK3}),
        html.Div("These are candidates, not results — hit Run to see what "
                 "they do to the reported holdout, which the search never "
                 "saw. If the holdout gets worse, the curves were fitting "
                 "noise: revert them.",
                 style={"color": MUTED, "marginTop": "2px"}),
    ]
    return rows, msg


@app.callback(Output("tmpl_download", "data"),
              Input("tmpl_dl_btn", "n_clicks"),
              State("brand", "value"), State("datapath", "data"),
              prevent_initial_call=True)
def _download_template(_n, brand, dp):
    if not dp or not brand:
        return no_update
    tdf = build_template_df(dp["path"], brand)

    def _to_xlsx(buf):
        with pd.ExcelWriter(buf, engine="openpyxl") as xw:
            tdf.to_excel(xw, index=False, sheet_name="variable_template")
    return dcc.send_bytes(_to_xlsx, f"variable_template_{_slug(brand)}.xlsx")


@app.callback(Output("cfg_status", "children", allow_duplicate=True),
              Output("config_rev", "data", allow_duplicate=True),
              Input("tmpl_upload", "contents"),
              State("tmpl_upload", "filename"), State("datapath", "data"),
              prevent_initial_call=True)
def _upload_template(contents, filename, dp):
    if not contents or not dp:
        return no_update, no_update
    import base64
    import io
    try:
        _, b64 = contents.split(",", 1)
        raw = base64.b64decode(b64)
        if str(filename).lower().endswith((".xlsx", ".xls")):
            tdf = pd.read_excel(io.BytesIO(raw))
        else:
            tdf = pd.read_csv(io.BytesIO(raw))
    except Exception as e:
        return f"Could not read template '{filename}': {e}", no_update
    status, results = apply_template_df(tdf, dp["path"])
    summary = "; ".join(f"{p}×{c}: {s}" for p, c, s in results)
    if status == "applied":
        return f"Template applied — '{filename}': {summary}", time.time()
    if status == "rollback_incomplete":
        # disk DID change and couldn't be fully restored — refresh the UI and
        # surface the recovery path (originals kept as .bak).
        return (f"⚠ PARTIAL WRITE — '{filename}': {summary}", time.time())
    return (f"Template REJECTED (nothing written — fix and re-upload) — "
            f"'{filename}': {summary}", no_update)


@app.callback(Output("export_download", "data"),
              Output("export_status", "children"),
              Input("export_btn", "n_clicks"),
              State("export_mode", "value"), State("brand", "value"),
              State("channel", "value"), State("target", "value"),
              State("window", "value"), State("datapath", "data"),
              State("time_map", "data"), State("strictness", "value"),
              State("family_caps", "data"),
              prevent_initial_call=True)
def _export(_n, mode, brand, channel, target, window, dp, time_map,
            strictness, cap_rows):
    if not dp or not target or not window:
        return no_update, "Load data and pick a target + window first."
    if mode == "slice" and (not brand or not channel):
        return no_update, "Pick a Product and Channel for a single-slice export."
    try:
        writer, n_ok, n_err = build_export_workbook(
            dp["path"], target, window, mode, brand, channel,
            time_map=time_map, tuning=tuning_of(strictness, cap_rows))
    except Exception as e:
        return no_update, f"Export failed: {type(e).__name__}: {e}"
    if n_ok == 0 and n_err == 0:
        return no_update, "Nothing to export (no runnable slices)."
    ds = _slug(os.path.splitext(os.path.basename(dp["path"]))[0])
    fname = (f"export_{_slug(brand)}_{_slug(channel)}.xlsx" if mode == "slice"
             else f"export_all_{ds}.xlsx")
    # download even if every slice failed — the workbook carries the errors
    # sheet so the user can see WHY.
    if n_ok == 0:
        status = f"All {n_err} slice(s) failed — see the 'errors' sheet."
    else:
        status = f"Exported {n_ok} slice(s)" + (f" · {n_err} failed" if n_err
                                                else "")
    return dcc.send_bytes(writer, fname), status


def _summary_for(dp):
    if dp and os.path.exists(batch_summary_path(dp["path"])):
        return batch_summary_path(dp["path"])
    if dp and os.path.normpath(dp["path"]) == os.path.normpath(DEFAULT_DATA) \
            and os.path.exists(SUMMARY):
        return SUMMARY
    return None


def _grade_label(m) -> str:
    # Good ≤10% · Moderate ≤40% · Bad >40% holdout MAPE (Alex: slide 4)
    if pd.isna(m):
        return "No holdout"
    if m <= 10:
        return "Good Model"
    if m <= FLAG_THRESHOLD:
        return "Moderate Model"
    return "Bad Model"


def _batch_payload(s):
    s = s.copy()
    s.insert(0, "grade", s["MAPE_holdout_pct"].apply(_grade_label))
    cols = ([{"name": "MODEL", "id": "grade"}]
            + [{"name": c.replace("_", " "), "id": c}
               for c in s.columns if c != "grade"])
    flagged = int((s["MAPE_holdout_pct"] > FLAG_THRESHOLD).sum())
    pills = [
        html.Span(f"{len(s)} models", className="pill"),
        html.Span(["median R² ", html.Span(f"{s.R2.median():.2f}",
                                           className="mono")],
                  className="pill"),
        html.Span(["median holdout MAPE ",
                   html.Span(f"{s.MAPE_holdout_pct.median():.1f}%",
                             className="mono")], className="pill"),
    ]
    if flagged:
        pills.append(html.Span(f"{flagged} need review",
                               className="pill pill-red"))
    # Label + fully-colored cell (green / amber / red); white text for
    # contrast. NaN-holdout rows fall through to a neutral grey.
    q_mod = (f"{{MAPE_holdout_pct}} > 10 && "
             f"{{MAPE_holdout_pct}} <= {FLAG_THRESHOLD}")
    styles = [
        {"if": {"column_id": "grade"},
         "backgroundColor": "#F2F4F7", "color": INK3, "fontWeight": 600},
        {"if": {"column_id": "grade",
                "filter_query": "{MAPE_holdout_pct} <= 10"},
         "backgroundColor": GREEN, "color": "#fff", "fontWeight": 600},
        {"if": {"column_id": "grade", "filter_query": q_mod},
         "backgroundColor": WARN, "color": "#fff", "fontWeight": 600},
        {"if": {"column_id": "grade",
                "filter_query": f"{{MAPE_holdout_pct}} > {FLAG_THRESHOLD}"},
         "backgroundColor": RED, "color": "#fff", "fontWeight": 600},
    ]
    return s.round(3).to_dict("records"), cols, pills, styles


@app.callback(Output("batch_table", "data"), Output("batch_table", "columns"),
              Output("batch_pills", "children"),
              Output("batch_table", "style_data_conditional"),
              Input("batch_refresh", "n_clicks"), Input("datapath", "data"))
def _load_batch(_n, dp):
    p = _summary_for(dp)
    if p is None:
        return [], [], [html.Span("No batch yet — hit Run all combinations.",
                                  className="pill")], []
    return _batch_payload(pd.read_csv(p))


def _skipped_panel(store) -> list:
    """Visible list of what was NOT modeled and why."""
    if not store:
        return []
    rows = list(store.get("skipped", [])) + [
        dict(r, weeks_in_window="—", selling_weeks="—")
        for r in store.get("errors", [])]
    if not rows:
        return [html.Div("Every product × channel had enough data inside the "
                         "window to be modeled.",
                         style={"fontSize": "12px", "color": MUTED})]
    return [html.Div(className="card", children=[
        html.Div(className="card-head", children=[
            html.Div(f"Not modeled ({len(rows)})", className="card-title"),
            html.Div("Slices with too little data inside the fixed modeling "
                     "window. These produce no model and no due-tos — "
                     "deliberately. Investigate the ones you expected to see.",
                     className="card-sub"),
        ]),
        dash_table.DataTable(
            data=[{"product": r["brand"], "channel": r["channel"],
                   "weeks in window": r.get("weeks_in_window", "—"),
                   "selling weeks": r.get("selling_weeks", "—"),
                   "reason": r["reason"]} for r in rows],
            columns=[{"name": c, "id": c} for c in
                     ["product", "channel", "weeks in window",
                      "selling weeks", "reason"]],
            page_size=25, sort_action="native", **TBL_BASE),
    ])]


@app.callback(Output("skipped_panel", "children"),
              Input("batch_table", "data"), State("datapath", "data"),
              State("target", "value"), State("window", "value"),
              State("strictness", "value"), State("family_caps", "data"))
def _show_skipped(_data, dp, target, window, strictness, cap_rows):
    if not dp:
        return []
    return _skipped_panel(cached_store(dp["path"], target, window,
                                       tuning_of(strictness, cap_rows)))


@app.callback(Output("batch_table", "data", allow_duplicate=True),
              Output("batch_table", "columns", allow_duplicate=True),
              Output("batch_pills", "children", allow_duplicate=True),
              Output("batch_table", "style_data_conditional",
                     allow_duplicate=True),
              Input("batch_run", "n_clicks"),
              State("datapath", "data"), State("target", "value"),
              State("window", "value"), State("strictness", "value"),
              State("family_caps", "data"), prevent_initial_call=True)
def _run_batch(_n, dp, target, window, strictness, cap_rows):
    if not dp:
        return [], [], [html.Span("Load a datafile first.",
                                  className="pill")], []
    t0 = time.time()
    path = dp["path"]
    target = target or "Volume Sales"
    window = window or 104
    tuning = tuning_of(strictness, cap_rows)
    s = run_batch(path, target, window, tuning)
    if not len(s):
        return [], [], [html.Span("No slice had enough data in the window.",
                                  className="pill pill-red")], []
    data, cols, pills, styles = _batch_payload(s)
    store = cached_store(path, target, window, tuning) or {}
    a, b = window_range(path, window)
    pills.append(html.Span(f"window {a:%b %Y}–{b:%b %Y}", className="pill"))
    if store.get("skipped"):
        pills.append(html.Span(f"{len(store['skipped'])} not modeled",
                               className="pill"))
    if tuning.get("max_per_family"):
        pills.append(html.Span(
            "caps: " + ", ".join(f"{k}≤{v}" for k, v in
                                 sorted(tuning["max_per_family"].items())),
            className="pill"))
    if strictness and strictness != "balanced":
        pills.append(html.Span(f"{strictness} selection", className="pill"))
    pills.append(html.Span(f"ran in {time.time()-t0:.0f}s", className="pill"))
    return data, cols, pills, styles


@app.callback(Output("brand", "value", allow_duplicate=True),
              Output("batch_sel", "data"),
              Input("batch_table", "selected_rows"),
              State("batch_table", "data"), prevent_initial_call=True)
def _pick_slice(sel, data):
    if not sel or not data:
        from dash import no_update
        return no_update, no_update
    row = data[sel[0]]
    return row["brand"], {"brand": row["brand"], "channel": row["channel"]}


def metric_card(label, value, unit=None, warn=False, chip=None):
    lab = [label] + ([html.Span(chip, className="review-chip")] if chip else [])
    val = [value] + ([html.Span(unit, className="munit")] if unit else [])
    return html.Div([html.Div(lab, className="mlabel"),
                     html.Div(val, className="mvalue")],
                    className="mcard" + (" warn" if warn else ""))


COEF_GRID = "2.4fr 0.9fr 0.7fr 1.1fr 0.7fr 1.1fr 1.1fr 1.1fr"
YOY_GRID = "2.4fr 1fr 1fr 1fr 0.8fr"


def coef_grid_rows(coef_rows):
    """Rows come from the cached payload (`payload["coef_rows"]`), so the
    coefficient table renders from stored results — no refit to look at a
    different slice."""
    head = html.Div([html.Div(x) for x in
                     ["VARIABLE", "FAMILY", "SIGN", "T-STAT",
                      html.Div("VIF", style={"textAlign": "right"}),
                      html.Div("COEFFICIENT", style={"textAlign": "right"}),
                      # comparable across rows — the raw coefficient is not,
                      # because the inputs sit on different supports (Arko)
                      html.Div("IMPACT / SD", style={"textAlign": "right"},
                               title="Coefficient × the variable's standard "
                                     "deviation = volume moved by a 1-SD "
                                     "change. Same unit for every driver, so "
                                     "THIS column is the one you can rank."),
                      html.Div("AVG WKLY CONTRIB",
                               style={"textAlign": "right"})]],
                    className="bgrid-head",
                    style={"gridTemplateColumns": COEF_GRID})
    sign_sym = {"positive": ("+", GREEN), "negative": ("−", RED),
                "unconstrained": ("~", INK2)}
    rows = [head]
    for cr in coef_rows:
        name = cr["name"]
        family = cr["family"]
        sym, scolor = sign_sym.get(cr["sign"], ("—", INK2))
        t = np.nan if cr["t"] is None else float(cr["t"])
        t_abs = 0 if np.isnan(t) else abs(t)
        vif_v = np.nan if cr["vif"] is None else float(cr["vif"])
        _vif_cell, _vif_color = _fmt_vif(vif_v)
        due = float(cr["avg_dueto"])
        coef_v = float(cr["coef"])
        namecell = [html.Span("const (intercept)" if name == "const" else name,
                              style={"overflow": "hidden",
                                     "textOverflow": "ellipsis",
                                     "whiteSpace": "nowrap"})]
        if cr["forced"]:
            namecell.append(html.Span("FORCED", className="forced-chip"))
        rows.append(html.Div([
            html.Div(namecell, style={"display": "flex",
                                      "alignItems": "center",
                                      "overflow": "hidden",
                                      "fontWeight": 500}),
            html.Div(fam_chip(family)),
            html.Div(sym, className="mono", style={"color": scolor}),
            html.Div([
                html.Span(html.Span(style={
                    "width": f"{min(100, t_abs/10*100):.0f}%",
                    "background": NAVY if t_abs >= 2 else "#D0D5DD"}),
                    className="tbar"),
                html.Span("—" if np.isnan(t) else f"{t:.2f}",
                          className="mono",
                          style={"fontSize": "11.5px", "color": INK2,
                                 "marginLeft": "7px"}),
            ], style={"display": "flex", "alignItems": "center"}),
            html.Div(_vif_cell, className="mono",
                     style={"textAlign": "right", "fontSize": "12px",
                            "color": _vif_color,
                            "fontWeight": 600 if _vif_color in (WARN, RED)
                            else 400}),
            html.Div([html.Span(f"{coef_v:,.2f}"),
                      html.Span(f" ÷{fmt_km(cr['scale'])}",
                                style={"color": MUTED, "fontSize": "10px"})
                      if (cr.get("scale") or 1) != 1 else ""],
                     className="mono",
                     style={"textAlign": "right", "fontSize": "12px"}),
            html.Div("—" if cr.get("beta_std") is None
                     else fmt_compact(cr["beta_std"]), className="mono",
                     style={"textAlign": "right", "fontSize": "12px",
                            "fontWeight": 600}),
            html.Div(fmt_compact(due), className="mono",
                     style={"textAlign": "right", "fontSize": "12px",
                            "color": RED if due < 0 else GREEN}),
        ], className="bgrid", style={"gridTemplateColumns": COEF_GRID}))
    return rows


def yoy_grid_rows(cby):
    year_cols = [c for c in cby.columns if not str(c).startswith("YoY")]
    head_cells = [html.Div("DRIVER")] + \
        [html.Div(str(c).upper(), style={"textAlign": "right"})
         for c in year_cols] + \
        [html.Div("YOY Δ", style={"textAlign": "right"}),
         html.Div("YOY %", style={"textAlign": "right"})]
    n_years = len(year_cols)
    grid = f"2.4fr {'1fr ' * n_years}1fr 0.8fr"
    rows = [html.Div(head_cells, className="bgrid-head",
                     style={"gridTemplateColumns": grid})]
    for driver, rr in cby.iterrows():
        chg = rr.get("YoY_change", np.nan)
        pct = rr.get("YoY_pct", np.nan)
        cells = [html.Div(str(driver), style={"fontWeight": 500})]
        for c in year_cols:
            cells.append(html.Div(f"{rr[c]:,.0f}", className="mono",
                                  style={"textAlign": "right",
                                         "fontSize": "12px"}))
        chg_color = MUTED if (np.isnan(chg) or chg == 0) \
            else (RED if chg < 0 else GREEN)
        cells.append(html.Div(
            "—" if np.isnan(chg) else f"{chg:+,.0f}".replace("-", "−"),
            className="mono", style={"textAlign": "right", "fontSize": "12px",
                                     "color": chg_color}))
        if np.isnan(pct):
            pill = html.Span("—", className="pct-pill",
                             style={"background": "#F2F4F7", "color": INK3})
        else:
            bg, fg = (("#E0F2EC", GREEN) if pct > 0 else
                      ("#FBE9E7", RED) if pct < 0 else ("#F2F4F7", INK3))
            pill = html.Span(f"{pct:+.1f}%", className="pct-pill",
                             style={"background": bg, "color": fg})
        cells.append(html.Div(pill, style={"textAlign": "right"}))
        rows.append(html.Div(cells, className="bgrid",
                             style={"gridTemplateColumns": grid}))
    return rows


def _curves_figure(curves):
    """Fitted response curve per media variable, with the weeks actually
    executed plotted on it. A flattening curve is a diminishing return; a
    straight line means no saturation was fitted for that variable."""
    if not curves:
        f = _fig_base(go.Figure(), 260)
        f.add_annotation(text="No media variables in this model.",
                         showarrow=False, font=dict(size=12, color=MUTED))
        return f
    palette = [NAVY, "#0E7090", "#B54708", "#5925DC", "#067647", "#B42318"]
    f = go.Figure()
    for i, c in enumerate(curves):
        col = palette[i % len(palette)]
        short = c["variable"].replace("_Spend", "").replace("_", " ")
        f.add_scatter(x=c["curve_x"], y=c["curve_y"], mode="lines",
                      name=short, line=dict(color=col, width=2),
                      hovertemplate=(f"<b>{short}</b><br>execution %{{x:,.0f}}"
                                     "<br>modeled volume %{y:,.0f}"
                                     "<extra></extra>"))
        f.add_scatter(x=c["points_x"], y=c["points_y"], mode="markers",
                      name=f"{short} — weeks run", showlegend=False,
                      marker=dict(color=col, size=5, opacity=0.35),
                      hovertemplate=("execution %{x:,.0f}<br>"
                                     "volume %{y:,.0f}<extra></extra>"))
        if c["half_saturation"]:
            f.add_vline(x=c["half_saturation"], line_dash="dot",
                        line_color=col, opacity=0.5)
    _fig_base(f, 320)
    f.update_layout(margin=dict(l=8, r=8, t=8, b=8))
    f.update_xaxes(title=dict(text="weekly execution (adstocked)",
                              font=dict(size=10, color=MUTED)))
    f.update_yaxes(title=dict(text="modeled volume", font=dict(size=10,
                                                              color=MUTED)))
    return f


def _curves_caption(curves):
    if not curves:
        return "No media in this model."
    sat = [c for c in curves if c["saturated"]]
    bits = [f"{len(curves)} media variable(s)"]
    if sat:
        bits.append(f"{len(sat)} with a fitted saturation curve — the dotted "
                    "line is the half-saturation point, where the next unit "
                    "of execution starts returning less than the last")
    else:
        bits.append("none saturated — each is linear in execution, so no "
                    "diminishing return has been fitted (run Optimize media "
                    "curves on Variables to search for one)")
    return " · ".join(bits)


def _period_view(cw, period, compare, time_map, mode):
    """(figure, caption, card title) for the period panel, in either mode."""
    if mode == "dueto":
        # The COMPARE period is the baseline and the PRIMARY period is the
        # focus, so the due-to reads "what changed getting TO the period I
        # selected" — the direction a reader expects from "Q3 vs Q2".
        return (_dueto_figure(cw, compare, period, time_map),
                _dueto_caption(cw, compare, period, time_map),
                "Due-to — change by driver")
    return (_avgc_figure(cw, period, time_map, compare),
            _avgc_caption(cw, period, time_map, compare),
            "Avg weekly contribution per driver")


def _cby_df(p) -> pd.DataFrame:
    """Rebuild the by-year due-to table from a cached payload."""
    c = p["cby"]
    return pd.DataFrame(c["data"], index=c["index"], columns=c["columns"])


def _render_payload(p, avgc_period, avgc_compare, time_map, source="",
                    mode="contrib"):
    """Render every result view for ONE slice straight from its cached
    payload. This is the ONLY renderer — a Run and a filter change produce
    identical screens, because both go through here."""
    stats = p["stats"]
    ho = np.nan if stats["holdout_mape"] is None else float(stats["holdout_mape"])
    ho_warn = (not np.isnan(ho)) and ho > FLAG_THRESHOLD
    dates = pd.to_datetime(p["dates"])
    split = p["split"]

    metrics = [
        metric_card("R²", f"{stats['r2']:.3f}"),
        metric_card("ADJ R²", f"{stats['adj_r2']:.3f}"),
        metric_card("IN-SAMPLE MAPE", f"{stats['mape']:.1f}", unit="%"),
        metric_card("HOLDOUT MAPE",
                    "n/a" if np.isnan(ho) else f"{ho:.1f}",
                    unit=None if np.isnan(ho) else "%",
                    warn=ho_warn, chip="REVIEW" if ho_warn else None),
        metric_card("DURBIN–WATSON", f"{stats['durbin_watson']:.2f}"),
        metric_card("PREDICTORS", f"{stats['n_selected']}",
                    unit=f" · {stats['n_forced']} forced"
                    if stats["n_forced"] else None),
    ]

    f1 = go.Figure()
    if p["pred_te"]:
        f1.add_vrect(x0=dates[split], x1=dates[-1], fillcolor="#FEF3E2",
                     opacity=0.6, line_width=0)
    f1.add_scatter(x=dates, y=p["actual"], name="Actual",
                   line=dict(color=INK, width=1.8))
    f1.add_scatter(x=dates, y=p["fitted"], name="Fitted (all data)",
                   line=dict(color=NAVY, width=1.7, dash="dash"))
    if p["pred_te"]:
        f1.add_scatter(x=dates[split:], y=p["pred_te"],
                       name="Holdout forecast (out-of-sample)",
                       line=dict(color=WARN, width=1.7, dash="dash"))
        f1.add_vline(x=dates[split], line_dash="dot", line_color="#D0D5DD")
    _fig_base(f1, 300)

    f2 = go.Figure(go.Scatter(x=p["fitted"], y=p["resid"], mode="markers",
                              marker=dict(color=NAVY, opacity=0.42, size=6)))
    f2.add_hline(y=0, line_dash="dash", line_color="#D0D5DD")
    _fig_base(f2, 140)
    f3 = go.Figure(go.Histogram(x=p["resid"], marker_color="#B7CBEA"))
    _fig_base(f3, 140)

    coef_children = coef_grid_rows(p["coef_rows"])
    coef_sub = (f"{stats['n_selected']} selected of {stats['n_candidates']} "
                "candidates · "
                + ("all pass sign checks" if stats["n_conflicts"] == 0
                   else f"{stats['n_conflicts']} sign conflicts")
                + (f" · {stats['n_highvif']} with VIF>{int(VIF_WARN)} "
                   "(collinearity, non-blocking)" if stats["n_highvif"]
                   else " · VIF all ≤10"))

    cby = _cby_df(p)
    year_cols = p["year_labels"]
    plot_tbl = cby[year_cols].drop("Intercept", errors="ignore")
    f4 = go.Figure()
    palette = ["#C5CEDC", NAVY, "#0E7090"]
    for i, yc in enumerate(year_cols):
        f4.add_bar(y=plot_tbl.index, x=plot_tbl[yc], name=str(yc),
                   orientation="h", marker_color=palette[i % len(palette)],
                   text=[fmt_km(v) for v in plot_tbl[yc]],
                   textposition="outside", textfont=dict(size=9),
                   cliponaxis=False)
    f4.update_layout(barmode="group")
    _fig_base(f4, max(320, 30 * len(plot_tbl)))
    _room_for_labels(f4, plot_tbl.values.ravel())

    contrib_wk = p["contrib_wk"]
    valid = {o["value"] for o in _avgc_period_options(contrib_wk, time_map)}
    period = avgc_period if avgc_period in valid else "all"
    compare = (avgc_compare if avgc_compare in (valid | {"none"})
               and avgc_compare != period else "none")
    f5, avgc_sub, avgc_title = _period_view(contrib_wk, period, compare,
                                            time_map, mode)

    yoy_children = yoy_grid_rows(cby.round(0))
    fit_sub = (f"{p['brand']} × {p['channel']} · {p['target']} · window "
               f"{p['window_start']} → {p['window_end']} "
               f"({p['n_weeks']} wks)" + (f" · {source}" if source else ""))

    last_run = {
        "brand": p["brand"], "channel": p["channel"], "target": p["target"],
        "window": p["window"], "periods": p["periods"],
        "year_labels": year_cols,
        "coef": p["coef_map"], "dueto_by_period": p["dueto_by_period"],
        "beta_std": p.get("beta_std", {}),
        "stats": {"r2": stats["r2"], "adj_r2": stats["adj_r2"],
                  "mape": stats["mape"],
                  "holdout_mape": stats["holdout_mape"],
                  "durbin_watson": stats["durbin_watson"],
                  "n_selected": stats["n_selected"],
                  "n_forced": stats["n_forced"]},
        "ts": time.strftime("%H:%M:%S"),
    }
    curves = p.get("media_curves", [])
    return (metrics, f1, fit_sub, f2, f3, coef_children, coef_sub, f4, f5,
            yoy_children, "", last_run, contrib_wk, avgc_sub, avgc_title,
            _curves_figure(curves), _curves_caption(curves))


VIEW_OUTPUTS = [
    Output("metrics", "children"), Output("fig_fit", "figure"),
    Output("fit_sub", "children"),
    Output("fig_resid", "figure"), Output("fig_residhist", "figure"),
    Output("coef_grid", "children"), Output("coef_sub", "children"),
    Output("fig_cby", "figure"), Output("fig_avgc", "figure"),
    Output("yoy_grid", "children"),
    Output("err", "children"), Output("runstat", "children"),
    Output("last_run", "data"),
    Output("config_rev", "data", allow_duplicate=True),
    Output("contrib_wk", "data"), Output("avgc_sub", "children"),
    Output("avgc_title", "children"),
    Output("fig_curves", "figure"), Output("curves_sub", "children"),
]


@app.callback(
    [Output(o.component_id, o.component_property, allow_duplicate=True)
     for o in VIEW_OUTPUTS],
    Input("brand", "value"), Input("channel", "value"),
    Input("datapath", "data"),
    State("target", "value"), State("window", "value"),
    State("avgc_period", "value"), State("avgc_compare", "value"),
    State("time_map", "data"), State("contrib_mode", "value"),
    State("strictness", "value"), State("family_caps", "data"),
    prevent_initial_call=True)
def show_cached_slice(brand, channel, dp, target, window, avgc_period,
                      avgc_compare, time_map, mode, strictness, cap_rows):
    """Changing the Product / Channel FILTER re-reads the cached batch — it
    does not refit (Alex 2026-07-27). If the slice was never run, or was
    skipped for insufficient data in the window, say so instead of showing
    the previous slice's numbers, which would be actively misleading."""
    empty = _fig_base(go.Figure(), 260)
    blank = ([], empty, "", empty, empty, [], "", empty, empty, [])
    if not (dp and brand and channel):
        return (*blank, "", "", None, no_update, None, "", no_update,
                empty, "")
    path = dp["path"]
    tuning = tuning_of(strictness, cap_rows)
    p = cached_slice(path, target, window, brand, channel, tuning)
    if p is None:
        st = cached_store(path, target, window, tuning)
        if st is None:
            msg = ("No results cached for the current modeling settings — "
                   "go to Model Runs and hit “Run all combinations”; every "
                   "screen then filters those results instantly. (Changing "
                   "the target, window, strictness or a family cap needs a "
                   "fresh batch — results are never mixed across settings.)")
        else:
            sk = next((s for s in st["skipped"] if s["brand"] == brand
                       and s["channel"] == channel), None)
            er = next((s for s in st["errors"] if s["brand"] == brand
                       and s["channel"] == channel), None)
            msg = (f"{brand} × {channel} was not modeled — {sk['reason']}"
                   if sk else
                   f"{brand} × {channel} failed — {er['reason']}" if er else
                   f"{brand} × {channel} is not in the cached run.")
        return (*blank, msg, "", None, no_update, None, "", no_update,
                empty, "")
    out = _render_payload(p, avgc_period, avgc_compare, time_map,
                          source="cached", mode=mode)
    runstat = [html.Span("filtered", className="pill"),
               html.Span(" from cached run", style={"marginLeft": "6px"})]
    return (*out[:11], runstat, out[11], no_update, out[12], out[13],
            out[14], out[15], out[16])


@app.callback(
    VIEW_OUTPUTS,
    Input("run_vars", "n_clicks"), Input("run_batch_one", "n_clicks"),
    State("brand", "value"), State("channel", "value"),
    State("target", "value"), State("window", "value"),
    State("datapath", "data"), State("cfg_table", "data"),
    State("config_scope", "value"), State("avgc_period", "value"),
    State("avgc_compare", "value"), State("time_map", "data"),
    State("contrib_mode", "value"), State("strictness", "value"),
    State("family_caps", "data"), prevent_initial_call=True)
def run_model(_v, _b, brand, channel, target, window, dp, cfg_rows, scope,
              avgc_period, avgc_compare, time_map, mode, strictness,
              cap_rows):
    empty = _fig_base(go.Figure(), 260)
    blank = ([], empty, "", empty, empty, [], "", empty, empty, [])
    rev = no_update            # bumped only if a Variables Run writes a config
    missing = [lbl for lbl, v in
               [("datafile (hit Load data in the rail)", dp),
                ("product", brand), ("channel", channel),
                ("target", target), ("window", window)] if not v]
    if missing:
        msg = "Cannot run — missing: " + ", ".join(missing)
        print(f"[run_model] {msg}", flush=True)
        return (*blank, msg, "", no_update, rev, None, no_update,
                no_update, empty, "")  # clear stale
    path = dp["path"]
    t0 = time.time()
    print(f"[run_model] {brand} x {channel} · target={target} "
          f"window={window}", flush=True)
    try:
        trig = ctx.triggered_id
    except Exception:
        trig = None
    # Variables-tab Run: persist the on-screen edits first so "modify
    # variables and hit run" takes effect (Alex, slide 2), writing to the
    # scoped file (Product default or this channel's override). Invalid bounds
    # (lo >= hi) ABORT with a visible error rather than silently running the
    # old config. A Variables run is WYSIWYG — it runs the file just edited.
    # Batch/other runs use override-wins resolution.
    try:
        # everything that can raise (config read/create, sheet load, fit) is in
        # ONE try, so a bad path / missing sheet routes to the failure return
        # that clears the contributions store — not an unhandled callback error.
        load_or_create_config(path, brand)      # ensure the default exists
        if trig == "run_vars":
            run_cfg_path = save_cfg_path(path, brand, scope, channel)
            if cfg_rows:
                ok, _, bad = _persist_cfg_rows(cfg_rows, run_cfg_path)
                if not ok:
                    return (*blank,
                            f"Not run — bound lo must be < bound hi for: {bad}."
                            " Fix these and re-run; your edits were not saved.",
                            "edits invalid — not saved", no_update, rev,
                            None, no_update, no_update, empty, "")
                # config written — signal readers even if the model then fails,
                # so the scope status reflects the new override.
                rev = time.time()
        else:
            run_cfg_path = resolve_cfg_path(path, brand, channel)
        tuning = tuning_of(strictness, cap_rows)
        cfg = cp.ModelConfig(target=target, model_weeks=int(window),
                             window_end=dataset_anchor(path),
                             variable_config=run_cfg_path, **tuning)
        r = cp.run_slice("", brand, channel, config=cfg,
                         df=load_sheet(path, brand))
    # A slice with too little data inside the fixed window is a deliberate
    # non-result, not a crash — explain it in the words the analyst needs.
    except cp.InsufficientWindowData as e:
        return (*blank, f"{brand} × {channel} was not modeled — {e}",
                "not modeled", None, rev, None, no_update, no_update,
                empty, "")
    except Exception as e:
        import traceback
        traceback.print_exc()
        return (*blank, f"Model failed: {type(e).__name__}: {e}",
                "run failed — see console", no_update, rev,
                None, no_update, no_update, empty, "")
    dur = time.time() - t0

    # Cache this fresh result under the same key the batch uses, so the
    # Product/Channel filter on the result screens immediately picks up the
    # re-tuned model instead of serving the pre-tuning cached one.
    p = _payload(r, brand, channel, target, window)
    key = _results_key(path, target, window, tuning)
    store = RESULTS.setdefault(key, {"models": {}, "skipped": [], "errors": [],
                                     "target": target, "window": int(window),
                                     "tuning": tuning,
                                     "ts": time.strftime("%H:%M:%S")})
    store["models"][(brand, channel)] = p
    store["skipped"] = [s for s in store["skipped"]
                        if not (s["brand"] == brand and s["channel"] == channel)]

    out = _render_payload(p, avgc_period, avgc_compare, time_map, mode=mode)
    runstat = ["Config saved · last run ",
               html.Span(time.strftime("%H:%M"), className="mono"),
               f" ({dur:.1f}s)"]
    return (*out[:11], runstat, out[11], rev, out[12], out[13], out[14],
            out[15], out[16])


@app.callback(Output("avgc_period", "options"), Output("avgc_period", "value"),
              Input("contrib_wk", "data"), Input("time_map", "data"),
              State("avgc_period", "value"))
def _avgc_period_opts(cw, time_map, current):
    opts = _avgc_period_options(cw, time_map)
    values = {o["value"] for o in opts}
    return opts, (current if current in values else "all")


@app.callback(Output("avgc_compare", "options"), Output("avgc_compare", "value"),
              Input("contrib_wk", "data"), Input("time_map", "data"),
              Input("avgc_period", "value"), State("avgc_compare", "value"))
def _compare_opts(cw, time_map, period, current):
    # depends on the primary period so changing it drops the now-duplicate
    # compare option (and resets the value to "none" if it collided).
    opts = _compare_options(cw, time_map, period)
    values = {o["value"] for o in opts}
    return opts, (current if current in values else "none")


@app.callback(Output("time_map", "data"), Output("timemap_status", "children"),
              Input("timemap_upload", "contents"),
              State("timemap_upload", "filename"), prevent_initial_call=True)
def _timemap_upload(contents, filename):
    # Parse an uploaded week→period map: 'date' (or 'week') + 'period' columns.
    # A date may appear in multiple periods (e.g. Current Year AND Quarter 4).
    if not contents:
        return no_update, no_update
    import base64
    import io
    try:
        _, b64 = contents.split(",", 1)
        raw = base64.b64decode(b64)
        df = (pd.read_excel(io.BytesIO(raw))
              if str(filename).lower().endswith((".xlsx", ".xls"))
              else pd.read_csv(io.BytesIO(raw)))
    except Exception as e:
        return no_update, f"Could not read map '{filename}': {e}"
    cols = {str(c).strip().lower(): c for c in df.columns}
    dcol = cols.get("date") or cols.get("week")
    pcol = cols.get("period") or cols.get("label")
    if not dcol or not pcol:
        return no_update, ("Map needs a 'date' (or 'week') column and a "
                           "'period' column.")
    # Coerce dates; rows with an invalid/empty date are SKIPPED and reported
    # (an unfiltered NaT would poison sorted() with mixed str/NaN — TypeError).
    # format="mixed" parses each cell independently so a CSV mixing YYYY-MM-DD
    # and MM/DD/YYYY doesn't drop the latter (pandas <2.0 falls back to infer).
    try:
        parsed = pd.to_datetime(df[dcol], errors="coerce", format="mixed")
    except (ValueError, TypeError):
        parsed = pd.to_datetime(df[dcol], errors="coerce")
    bad = parsed.isna()
    n_bad = int(bad.sum())
    bad_rows = [int(i) + 2 for i in np.where(bad.values)[0]]   # 1-based + header
    m = {}
    for d, p in zip(parsed[~bad].dt.strftime("%Y-%m-%d"),
                    df[pcol][~bad].astype(str)):
        p = p.strip()
        if p and p.lower() != "nan":
            m.setdefault(p, set()).add(d)
    m = {k: sorted(v) for k, v in m.items()}
    if not m:
        return no_update, (f"No valid (date, period) rows — {n_bad} row(s) had "
                           "an invalid/empty date." if n_bad
                           else "No (date, period) rows found in the map.")
    skipped = (f" · skipped {n_bad} row(s) with invalid/empty date "
               f"(rows {bad_rows[:5]}{'…' if len(bad_rows) > 5 else ''})"
               if n_bad else "")
    return m, (f"Loaded {len(m)} mapped period(s): "
               + ", ".join(list(m)[:6]) + (" …" if len(m) > 6 else "") + skipped)


@app.callback(Output("fig_avgc", "figure", allow_duplicate=True),
              Output("avgc_sub", "children", allow_duplicate=True),
              Output("avgc_title", "children", allow_duplicate=True),
              Input("avgc_period", "value"), Input("avgc_compare", "value"),
              Input("time_map", "data"), Input("contrib_mode", "value"),
              State("contrib_wk", "data"), prevent_initial_call=True)
def _contrib_period(period, compare, time_map, mode, cw):
    # Re-aggregate the avg-weekly view over the selected period — optionally
    # side-by-side with a COMPARE period (E4: quarter-to-quarter etc.) — without
    # re-running. `time_map` is an INPUT so re-uploading a map with the SAME
    # period names but different dates still redraws. contrib_wk holds the LAST
    # SUCCESSFUL run and is CLEARED on any failed run — the same event that
    # blanks the other Contributions charts — so this can't diverge.
    if not cw:
        return no_update, no_update, no_update
    valid = {o["value"] for o in _avgc_period_options(cw, time_map)}
    period = period if period in valid else "all"       # re-clamp (map changes)
    if compare not in (valid | {"none"}) or compare == period:  # never A-vs-A
        compare = "none"
    return _period_view(cw, period, compare, time_map, mode)


# ═══════════════════════════════════════════════════════════════════════════
# HIGH LEVEL — Total Brand / Total Channel  (Alex 2026-08-11)
# ═══════════════════════════════════════════════════════════════════════════

def _pct(v, dec=1):
    return "—" if v is None or (isinstance(v, float) and np.isnan(v)) \
        else f"{v:.{dec}f}%"


def _metric(label, value, sub=None, tone=INK):
    return html.Div(className="card card-pad", style={"flex": 1,
                                                      "minWidth": "140px"},
                    children=[
        html.Div(label, style={"fontSize": "11px", "color": MUTED,
                               "letterSpacing": ".04em"}),
        html.Div(value, style={"fontSize": "22px", "fontWeight": 600,
                               "color": tone, "fontFamily": "IBM Plex Mono",
                               "margin": "2px 0"}),
        html.Div(sub or "", style={"fontSize": "11px", "color": INK3}),
    ])


@app.callback(Output("total_summary", "children"),
              Output("total_group", "options"),
              Output("total_group", "value"),
              Output("total_status", "children"),
              Input("total_mode", "value"), Input("active_tab", "data"),
              State("datapath", "data"), State("target", "value"),
              State("window", "value"), State("strictness", "value"),
              State("family_caps", "data"), State("total_group", "value"))
def _total_summary(mode, tab, dp, target, window, strictness, caps, current):
    if tab != "total":
        return no_update, no_update, no_update, no_update
    if not dp:
        return html.Div("Load a datafile first.",
                        style={"fontSize": "13px", "color": INK3}), [], None, ""
    store = cached_store(dp["path"], target, window,
                         tuning_of(strictness, caps))
    if not store or not store.get("models"):
        return (html.Div("No cached results yet — go to Model Runs and hit "
                         "“Run all combinations”, then come back.",
                         style={"fontSize": "13px", "color": WARN}),
                [], None, "")
    payloads = list(store["models"].values())
    rolled = ru.rollup(payloads, mode)
    tbl = ru.rollup_table(rolled)
    disp = tbl.copy()
    disp["volume"] = disp["volume"].map(lambda v: fmt_km(v, 1))
    for c in ("r2",):
        disp[c] = disp[c].map(lambda v: f"{v:.3f}")
    for c in ("mape", "wmape", "bias_pct", "worst_slice_mape"):
        disp[c] = disp[c].map(lambda v: _pct(v))
    disp.columns = ["Group", "Slices", "Weeks", "Volume", "R²", "MAPE",
                    "WMAPE", "Bias", "Worst slice MAPE"]
    body = dash_table.DataTable(
        data=disp.to_dict("records"),
        columns=[{"name": c, "id": c} for c in disp.columns],
        sort_action="native", page_size=25,
        style_data_conditional=[
            {"if": {"filter_query": "{Slices} > 0"}, "cursor": "default"}],
        **TBL_BASE)
    opts = list(tbl["group"])
    val = current if current in opts else (opts[0] if opts else None)
    note = ("Rolled up by product" if mode == "brand"
            else "Rolled up by channel" if mode == "channel"
            else "Everything in one line")
    return (card(f"{len(opts)} total view" + ("s" if len(opts) != 1 else ""),
                 note, body, pad=False),
            opts, val, f"{store['n_models'] if 'n_models' in store else len(payloads)} "
                       f"models · cached {store['ts']}")


@app.callback(Output("total_detail", "children"),
              Input("total_group", "value"), Input("total_mode", "value"),
              State("datapath", "data"), State("target", "value"),
              State("window", "value"), State("strictness", "value"),
              State("family_caps", "data"))
def _total_detail(label, mode, dp, target, window, strictness, caps):
    if not dp or not label:
        return None
    store = cached_store(dp["path"], target, window,
                         tuning_of(strictness, caps))
    if not store:
        return None
    rolled = ru.rollup(list(store["models"].values()), mode)
    g = rolled["groups"].get(label)
    if not g:
        return None
    s = g["stats"]

    strip = html.Div(style={"display": "flex", "gap": "12px",
                            "marginBottom": "14px", "flexWrap": "wrap"},
                     children=[
        _metric("MODELS", str(s["n_slices"])),
        _metric("TOTAL VOLUME", fmt_km(s["volume"], 1)),
        _metric("R²", f"{s['r2']:.3f}"),
        _metric("MAPE", _pct(s["mape"]),
                "average weekly % error"),
        _metric("WMAPE", _pct(s["wmape"]),
                "error volume ÷ actual volume", tone=NAVY),
        _metric("BIAS", _pct(s["bias_pct"], 2),
                "over/under across the window"),
    ])

    # actual vs fitted at the total level
    dates = pd.to_datetime(g["dates"])
    fig = go.Figure()
    fig.add_trace(go.Scatter(x=dates, y=g["actual"], name="Actual",
                             mode="lines", line=dict(color=INK, width=1.8)))
    fig.add_trace(go.Scatter(x=dates, y=g["fitted"], name="Fitted",
                             mode="lines",
                             line=dict(color=NAVY, width=1.8, dash="dot")))
    fig = _fig_base(fig, 260)

    warn = None
    if s["weeks_full_coverage"] < s["n_weeks"]:
        warn = html.Div(
            f"⚠ {s['n_weeks'] - s['weeks_full_coverage']} of {s['n_weeks']} "
            f"weeks are covered by fewer than all {s['n_slices']} models "
            f"(as few as {s['min_slices_in_a_week']}). A slice that starts "
            "part-way through the window contributes nothing before it "
            "starts, so dips at the left edge of the chart can be coverage, "
            "not performance.",
            style={"fontSize": "12px", "color": WARN, "marginTop": "8px"})

    # contribution stack at the total level
    ctab = ru.contribution_table(g, top=14)
    ctab_disp = ctab.copy()
    ctab_disp["contribution"] = ctab_disp["contribution"].map(
        lambda v: fmt_km(v, 1))
    ctab_disp["share_pct"] = ctab_disp["share_pct"].map(lambda v: _pct(v))
    ctab_disp.columns = ["Driver", "Contribution", "Share of modeled volume"]

    due = ru.rollup_due_to(rolled, label)
    due = due[due["driver"] != "Intercept"].head(12)
    dfig = go.Figure(go.Bar(
        x=due["due_to"], y=due["driver"], orientation="h",
        marker_color=[GREEN if v >= 0 else RED for v in due["due_to"]]))
    dfig.update_layout(yaxis=dict(autorange="reversed"))
    dfig = _room_for_labels(_fig_base(dfig, 30 + 26 * len(due)),
                            due["due_to"].tolist())

    members = pd.DataFrame(g["members"])
    members["volume"] = members["volume"].map(lambda v: fmt_km(v, 1))
    for c in ("r2",):
        members[c] = members[c].map(lambda v: f"{v:.3f}")
    for c in ("mape", "holdout_mape"):
        members[c] = members[c].map(lambda v: _pct(v))
    members.columns = ["Product", "Channel", "R²", "MAPE", "Holdout MAPE",
                       "Volume", "Variables"]

    return html.Div([
        strip,
        card(f"{label} — actual vs fitted",
             f"{s['n_slices']} models summed · decomposition exact to "
             f"{s['decomposition_error']:.0e}",
             html.Div([dcc.Graph(figure=fig, config={"displayModeBar": False}),
                       warn] if warn is not None
                      else dcc.Graph(figure=fig,
                                     config={"displayModeBar": False}))),
        html.Div(style={"display": "flex", "gap": "14px",
                        "marginTop": "14px", "flexWrap": "wrap"}, children=[
            html.Div(style={"flex": "1 1 420px"}, children=[
                card("Total contributions", "level over the whole window",
                     dash_table.DataTable(
                         data=ctab_disp.to_dict("records"),
                         columns=[{"name": c, "id": c}
                                  for c in ctab_disp.columns],
                         page_size=15, **TBL_BASE), pad=False)]),
            html.Div(style={"flex": "1 1 420px"}, children=[
                card("Year-over-year due-to",
                     "change in the latest 52 weeks vs the 52 before",
                     dcc.Graph(figure=dfig,
                               config={"displayModeBar": False}))]),
        ]),
        html.Div(style={"marginTop": "14px"}, children=[
            card("Models in this total",
                 "the parts the total is made of — a good total can still "
                 "hide a bad slice",
                 dash_table.DataTable(
                     data=members.to_dict("records"),
                     columns=[{"name": c, "id": c} for c in members.columns],
                     sort_action="native", page_size=20, **TBL_BASE),
                 pad=False)]),
    ])


# ═══════════════════════════════════════════════════════════════════════════
# SATURATION — weekly transforms, correlations, ROI curve, live overrides
#
# Alex on the 2026-08-11 call: "I want something that can go: okay, well, what
# if I want to override this visually? Here's the graph, here's the decayed
# media, I'm going to change from a 0.2 to a 0.3. What does that look like?
# How much worse is it? … I really like this, and that's what I wanted was some
# kind of optimization. But I want to see the reasoning behind it and some kind
# of visual for that selection."
#
# So the sliders are the point of this screen, not a garnish. They redraw from
# the CACHED slice — no refit — which is what makes dragging them feel live.
# The coefficient is held at its fitted value while the transform changes, so
# what moves is the SHAPE of the input and its correlation with the target. The
# screen says so, because pretending the whole model updated would be a lie.
# ═══════════════════════════════════════════════════════════════════════════

def _sat_slice(dp, target, window, brand, channel, strictness, caps):
    return cached_slice(dp["path"], target, window, brand, channel,
                        tuning_of(strictness, caps)) if dp else None


@app.callback(Output("sat_var", "options"), Output("sat_var", "value"),
              Output("sat_note", "children"),
              Input("active_tab", "data"), Input("brand", "value"),
              Input("channel", "value"),
              State("datapath", "data"), State("target", "value"),
              State("window", "value"), State("strictness", "value"),
              State("family_caps", "data"), State("sat_var", "value"))
def _sat_vars(tab, brand, channel, dp, target, window, strictness, caps, cur):
    if tab != "sat":
        return no_update, no_update, no_update
    p = _sat_slice(dp, target, window, brand, channel, strictness, caps)
    if not p:
        return [], None, html.Span(
            "No cached model for this slice — run it on Model Runs first.",
            style={"color": WARN})
    media = (p.get("contrib_wk") or {}).get("media") or []
    if not media:
        return [], None, html.Span(
            f"{brand} × {channel} has no media variable in its model, so "
            "there is no saturation curve to draw. Media that was never "
            "selected has no fitted coefficient — the Variables screen can "
            "force one in if it should be there.", style={"color": WARN})
    val = cur if cur in media else media[0]
    return media, val, (
        f"{brand} × {channel} · {len(media)} media variable"
        f"{'s' if len(media) != 1 else ''} in the model. The weekly panel "
        "below shows exactly what the engine fed the regression.")


@app.callback(Output("sat_decay", "value"), Output("sat_mid", "value"),
              Output("sat_slope", "value"),
              Input("sat_var", "value"), Input("sat_reset", "n_clicks"),
              State("brand", "value"), State("channel", "value"),
              State("datapath", "data"), State("target", "value"),
              State("window", "value"), State("strictness", "value"),
              State("family_caps", "data"))
def _sat_defaults(var, _r, brand, channel, dp, target, window, strictness, caps):
    """Sliders start at — and Reset returns to — the values the model used."""
    p = _sat_slice(dp, target, window, brand, channel, strictness, caps)
    if not p or not var:
        return 0.5, 0.5, 1.0
    for mc in p.get("media_curves") or []:
        if mc["variable"] == var:
            return (float(mc["decay"]),
                    float(mc["sat_midpoint"]) if mc["sat_midpoint"] else 0.5,
                    float(mc["sat_slope"]) if mc["sat_slope"] else 1.0)
    return 0.5, 0.5, 1.0


@app.callback(Output("sat_weekly", "children"), Output("sat_corr", "children"),
              Output("sat_status", "children"),
              Input("sat_var", "value"), Input("sat_decay", "value"),
              Input("sat_mid", "value"), Input("sat_slope", "value"),
              State("brand", "value"), State("channel", "value"),
              State("datapath", "data"), State("target", "value"),
              State("window", "value"), State("strictness", "value"),
              State("family_caps", "data"))
def _sat_weekly(var, decay, mid, slope, brand, channel, dp, target, window,
                strictness, caps):
    if not dp or not var:
        return None, None, ""
    try:
        r = cp.run_slice(dp["path"], brand, channel,
                         config=cp.ModelConfig(
                             target=target, model_weeks=int(window),
                             window_end=dataset_anchor(dp["path"]),
                             variable_config=cp.resolve_config_path(
                                 dp["path"], brand, channel)))
    except Exception as exc:                                   # noqa: BLE001
        return (html.Div(f"Could not rebuild the slice: {exc}",
                         style={"color": RED, "fontSize": "13px"}), None, "")
    spec = r["specs_by_name"].get(var)
    if spec is None or var not in r["df"].columns:
        return None, None, ""

    # the MODEL's own settings, for the side-by-side comparison
    base = sc.media_transform_series(
        r["df"], var, decay=spec.adstock_decay,
        scale=getattr(spec, "scale", 1.0),
        sat_midpoint=spec.sat_midpoint, sat_slope=spec.sat_slope,
        sat_ref=(r.get("sat_refs") or {}).get(var),
        target=r["y"], dates=r["df"]["date"])
    # the OVERRIDDEN settings from the sliders
    live = sc.media_transform_series(
        r["df"], var, decay=float(decay),
        scale=getattr(spec, "scale", 1.0),
        sat_midpoint=float(mid), sat_slope=float(slope),
        sat_ref=None, target=r["y"], dates=r["df"]["date"])

    dates = pd.to_datetime(live["dates"])
    fig = go.Figure()
    fig.add_trace(go.Bar(x=dates, y=live["raw"], name="Raw execution",
                         marker_color="#D6E4F5",
                         hovertemplate="%{x|%b %d %Y}<br>raw %{y:,.0f}<extra></extra>"))
    fig.add_trace(go.Scatter(x=dates, y=live["decayed"], name="Decayed (adstock)",
                             mode="lines", line=dict(color=NAVY, width=2)))
    if live["saturated_rescaled"]:
        fig.add_trace(go.Scatter(
            x=dates, y=live["saturated_rescaled"],
            name="Decayed + saturated (rescaled)", mode="lines",
            line=dict(color="#0E7090", width=2, dash="dot")))
    if live["target_rescaled"]:
        # Alex's bonus ask: "also including the target variable plotted so we
        # can visualize the relationship"
        fig.add_trace(go.Scatter(
            x=dates, y=live["target_rescaled"], name=f"{target} (rescaled)",
            mode="lines", line=dict(color=MUTED, width=1.4)))
    fig = _fig_base(fig, 320)
    fig.update_layout(barmode="overlay",
                      yaxis_title="execution (raw / decayed units)")

    # the saturated series on its own true [0,1) axis — the scale warning Alex
    # flagged, answered by giving it its own chart rather than a shared axis
    fig2 = None
    if live["saturated"]:
        fig2 = go.Figure(go.Scatter(
            x=dates, y=live["saturated"], mode="lines",
            line=dict(color="#0E7090", width=2), name="saturated"))
        fig2.add_hline(y=0.5, line=dict(color=MUTED, width=1, dash="dash"),
                       annotation_text="half saturation",
                       annotation_font_size=10)
        fig2 = _fig_base(fig2, 170)
        fig2.update_layout(yaxis=dict(range=[0, 1],
                                      title="saturated (0–1)"))

    rows = []
    for key, label in (("raw", "Raw"), ("decayed", "Decayed"),
                       ("saturated", "Decayed + saturated")):
        b, l = base["corr"].get(key), live["corr"].get(key)
        rows.append({
            "Measure": label,
            "Correlation with target (model settings)":
                "—" if b is None else f"{b:+.3f}",
            "Correlation with target (your settings)":
                "—" if l is None else f"{l:+.3f}",
            "Change": "—" if (b is None or l is None)
                      else f"{abs(l) - abs(b):+.3f}",
        })
    best = live.get("best_transform")
    changed = (abs(float(decay) - float(spec.adstock_decay or 0)) > 1e-9
               or abs(float(mid) - float(spec.sat_midpoint or 0.5)) > 1e-9
               or abs(float(slope) - float(spec.sat_slope or 1.0)) > 1e-9)

    corr_card = card(
        "Correlation with the target",
        ("|r| for each transform. This is the evidence for the decay and "
         "midpoint that were chosen — the transform that tracks the target "
         "best is usually the one the optimizer picked."),
        html.Div([
            dash_table.DataTable(
                data=rows, columns=[{"name": c, "id": c} for c in rows[0]],
                **TBL_BASE),
            html.Div(
                [f"Strongest on your settings: {best}. " if best else "",
                 ("These are correlations only — a higher |r| on its own does "
                  "not guarantee a better model, because the variable competes "
                  "with everything else in the regression. Hit Run on Variables "
                  "to see what it does to holdout MAPE.")
                 if changed else
                 "Sliders are at the model's own settings."],
                style={"fontSize": "11.5px", "color": INK3,
                       "marginTop": "8px"}),
        ]))

    kids = [card(f"{var} — weekly execution",
                 ("raw vs decayed vs decayed+saturated. The saturated series "
                  "is bounded in [0,1) by construction, so it is rescaled to "
                  "the decayed peak here for shape comparison and shown on its "
                  "true axis below."),
                 dcc.Graph(figure=fig, config={"displayModeBar": False}))]
    if fig2 is not None:
        kids.append(html.Div(style={"marginTop": "12px"}, children=[
            card("Saturated series, true scale",
                 "how close each week runs to the ceiling",
                 dcc.Graph(figure=fig2, config={"displayModeBar": False}))]))
    status = (f"decay {decay} · midpoint {mid} · slope {slope}"
              + ("  (overridden)" if changed else "  (model settings)"))
    return html.Div(kids), corr_card, status


@app.callback(Output("sat_spend", "value"), Output("sat_price", "value"),
              Output("sat_inputs_note", "children"),
              Input("sat_var", "value"), State("brand", "value"),
              State("channel", "value"), State("datapath", "data"),
              State("target", "value"), State("window", "value"),
              State("strictness", "value"), State("family_caps", "data"))
def _sat_inputs(var, brand, channel, dp, target, window, strictness, caps):
    """Auto-derive spend and price; both stay editable."""
    if not dp or not var:
        return None, None, ""
    try:
        r = cp.run_slice(dp["path"], brand, channel,
                         config=cp.ModelConfig(
                             target=target, model_weeks=int(window),
                             window_end=dataset_anchor(dp["path"]),
                             variable_config=cp.resolve_config_path(
                                 dp["path"], brand, channel)))
        kw = sc.curve_inputs_from_result(r, var)
    except Exception:                                          # noqa: BLE001
        return None, None, ""
    return (round(float(kw["spend"]), 2), round(float(kw["price"]), 4),
            f"Spend = Σ {var} over the latest {kw['weeks']} weeks; price = "
            f"mean Price per Volume over the same weeks. Margin is not in the "
            f"data — it scales the ROI axis but cannot move the optimal point.")


@app.callback(Output("sat_roi", "children"),
              Input("sat_var", "value"), Input("sat_mid", "value"),
              Input("sat_slope", "value"), Input("sat_spend", "value"),
              Input("sat_price", "value"), Input("sat_margin", "value"),
              State("brand", "value"), State("channel", "value"),
              State("datapath", "data"), State("target", "value"),
              State("window", "value"), State("strictness", "value"),
              State("family_caps", "data"))
def _sat_roi(var, mid, slope, spend, price, margin, brand, channel, dp,
             target, window, strictness, caps):
    if not dp or not var:
        return None
    try:
        r = cp.run_slice(dp["path"], brand, channel,
                         config=cp.ModelConfig(
                             target=target, model_weeks=int(window),
                             window_end=dataset_anchor(dp["path"]),
                             variable_config=cp.resolve_config_path(
                                 dp["path"], brand, channel)))
    except Exception as exc:                                   # noqa: BLE001
        return html.Div(f"{exc}", style={"color": RED, "fontSize": "13px"})

    raw = pd.to_numeric(r["df"][var], errors="coerce").fillna(0).values
    sat_alex = sc.midpoint_to_alex_saturation(float(mid), None, raw[-52:])
    d = sc.build_sat_curve(r, var, margin=float(margin or 30) / 100.0,
                           price=(float(price) if price else None),
                           spend=(float(spend) if spend else None),
                           saturation=float(np.clip(sat_alex, 0.01, 0.99)),
                           slope=float(slope))
    if "error" in d:
        return html.Div(d["error"], style={"color": WARN, "fontSize": "13px"})

    fig = go.Figure()
    fig.add_trace(go.Scatter(
        x=d["ratio"], y=d["sales"], name="Sales", mode="lines",
        line=dict(color="#7FD3E8", width=1), fill="tozeroy",
        fillcolor="rgba(127,211,232,0.28)",
        hovertemplate="%{x:.0%} of spend<br>%{y:,.0f} sales<extra></extra>"))
    fig.add_trace(go.Scatter(
        x=d["ratio"], y=d["avg_return"], name="Average Return", yaxis="y2",
        mode="lines", line=dict(color=NAVY, width=2),
        hovertemplate="%{x:.0%}<br>avg $%{y:,.2f}<extra></extra>"))
    fig.add_trace(go.Scatter(
        x=d["ratio"], y=d["marginal_return"], name="Marginal Return",
        yaxis="y2", mode="lines", line=dict(color=RED, width=2),
        hovertemplate="%{x:.0%}<br>marginal $%{y:,.2f}<extra></extra>"))
    fig.add_vline(x=1.0, line=dict(color=MUTED, width=1, dash="dash"),
                  annotation_text="current", annotation_font_size=10)
    if not d["at_boundary"]:
        fig.add_vline(x=d["optimal_ratio"],
                      line=dict(color=GREEN, width=1.5, dash="dot"),
                      annotation_text=f"optimal {d['optimal_ratio']:.0%}",
                      annotation_font_size=10)
    fig = _fig_base(fig, 380)
    fig.update_layout(
        xaxis=dict(title="% of current spend", tickformat=".0%"),
        yaxis=dict(title="Sales driven"),
        yaxis2=dict(title="$ ROI", overlaying="y", side="right",
                    showgrid=False, tickprefix="$"))

    tone = (WARN if d["at_boundary"]
            else GREEN if d["direction"] == "increase"
            else RED if d["direction"] == "decrease" else INK)
    headline = (d["verdict"] if d["at_boundary"] else
                f"{d['verdict'].upper()} — optimal is "
                f"{d['optimal_ratio']:.0%} of current spend "
                f"({fmt_km(d['optimal_spend'], 1)} vs "
                f"{fmt_km(d['current_spend'], 1)} today)")

    rows = [
        {"Point": "Current (100%)",
         "Spend": fmt_km(d["current_spend"], 1),
         "Sales driven": fmt_km(d["current_sales"], 1),
         "Average return": f"${d['current_avg_return']:,.2f}",
         "Marginal return": f"${d['current_marginal_return']:,.2f}"},
        {"Point": f"Optimal ({d['optimal_ratio']:.0%})",
         "Spend": fmt_km(d["optimal_spend"], 1),
         "Sales driven": fmt_km(d["optimal_sales"], 1),
         "Average return": f"${d['optimal_avg_return']:,.2f}",
         "Marginal return": f"${d['optimal_marginal_return']:,.2f}"},
    ]
    caveats = []
    if d.get("assumed_curve"):
        caveats.append(
            "This variable was modeled LINEAR — no saturation was fitted — so "
            "the curve below uses the slider values as an assumption, not a "
            "result. Read it as a what-if.")
    if d["at_boundary"]:
        caveats.append(
            "With a slope below 1 the curve is concave from the origin: "
            "marginal return sits under average return at every spend level "
            "and the two never cross, so there is no interior optimum to "
            "report. Raise the slope above 1 for an S-curve if that matches "
            "what you believe about this medium.")

    return html.Div([
        html.Div(headline, style={"fontSize": "13.5px", "fontWeight": 600,
                                  "color": tone, "marginBottom": "8px"}),
        dcc.Graph(figure=fig, config={"displayModeBar": False}),
        dash_table.DataTable(
            data=rows, columns=[{"name": c, "id": c} for c in rows[0]],
            **TBL_BASE),
        html.Div(f"Half-saturation at {fmt_km(d['half_saturation'], 1)} "
                 f"({d['half_saturation_pct']:.0%} of average weekly "
                 f"execution) · slope {d['slope']:.2f} · "
                 f"{d['weeks']} weeks · margin {d['margin']:.0%}",
                 style={"fontSize": "11.5px", "color": INK3,
                        "marginTop": "8px"}),
    ] + [html.Div(c, style={"fontSize": "11.5px", "color": WARN,
                            "marginTop": "6px"}) for c in caveats])


@app.callback(Output("sat_apply_status", "children"),
              Output("config_rev", "data", allow_duplicate=True),
              Input("sat_apply", "n_clicks"),
              State("sat_var", "value"), State("sat_decay", "value"),
              State("sat_mid", "value"), State("sat_slope", "value"),
              State("brand", "value"), State("channel", "value"),
              State("datapath", "data"), State("config_rev", "data"),
              prevent_initial_call=True)
def _sat_apply(_n, var, decay, mid, slope, brand, channel, dp, rev):
    """Write the three slider values into this product's variable config.

    Writes the PRODUCT DEFAULT (channel = ALL) so the change applies wherever
    the variable appears, matching how the Variables screen treats a product
    scope. Deliberately does NOT run anything — Alex's "these are candidates,
    not results; hit Run to see what they do" rule from the previous round
    applies here too.
    """
    if not dp or not var:
        return "Nothing to apply.", no_update
    try:
        path = cp.default_config_path(dp["path"], brand)
        with _CFG_WRITE_LOCK:
            cfg = pd.read_csv(path)
            m = cfg["variable"].astype(str) == str(var)
            if not m.any():
                return f"{var} is not in this product's config.", no_update
            cfg.loc[m, "adstock_decay"] = float(decay)
            cfg.loc[m, "sat_midpoint"] = float(mid)
            cfg.loc[m, "sat_slope"] = float(slope)
            cfg.to_csv(path, index=False)
    except Exception as exc:                                   # noqa: BLE001
        return html.Span(f"Could not write config: {exc}",
                         style={"color": RED}), no_update
    return (html.Span(
        f"Saved to {brand}'s config — decay {decay}, midpoint {mid}, slope "
        f"{slope}. These are candidates, not results: hit Run on Variables "
        "(or Model Runs) to see what they do to the holdout.",
        style={"color": GREEN}), (rev or 0) + 1)


# ═══════════════════════════════════════════════════════════════════════════
# MULTI-LEVEL — unpooled / pooled / hierarchical  (Alex 2026-08-11)
# ═══════════════════════════════════════════════════════════════════════════

@app.callback(Output("multi_brand_note", "children"),
              Input("brand", "value"), Input("active_tab", "data"))
def _multi_note(brand, tab):
    if tab != "multi":
        return no_update
    return f"Product: {brand or '—'} (all channels)"


@app.callback(Output("multi_out", "children"),
              Input("multi_run", "n_clicks"),
              State("brand", "value"), State("datapath", "data"),
              State("target", "value"), State("window", "value"),
              State("multi_shrink", "value"), State("multi_sdlimit", "value"),
              State("multi_opts", "value"),
              prevent_initial_call=True)
def _multi_run(_n, brand, dp, target, window, shrink, sdlimit, opts):
    if not dp or not brand:
        return html.Div("Load a datafile and pick a product first.",
                        style={"fontSize": "13px", "color": WARN})
    opts = opts or []
    cfg = cp.ModelConfig(target=target, model_weeks=int(window),
                         window_end=dataset_anchor(dp["path"]),
                         variable_config=cp.resolve_config_path(
                             dp["path"], brand, None))
    t0 = time.time()
    try:
        cmp_tbl = ml.compare_levels(
            dp["path"], brand, config=cfg,
            sd_limit=float(sdlimit),
            shrink=(None if "fit" in opts else float(shrink)))
        used_shrink = cmp_tbl.attrs["shrink"]
        hier = ml.run_hierarchical(
            dp["path"], brand, config=cfg, sd_limit=float(sdlimit),
            shrink=float(used_shrink), enforce_sign=("sign" in opts))
        pooled = hier["pooled"]
    except Exception as exc:                                   # noqa: BLE001
        return html.Div(f"{type(exc).__name__}: {exc}",
                        style={"color": RED, "fontSize": "13px"})

    a = cmp_tbl.attrs
    best = min([("Unpooled", a["unpooled_weighted"]),
                ("Pooled", a["pooled_weighted"]),
                ("Hierarchical", a["hier_weighted"])], key=lambda kv: kv[1])

    strip = html.Div(style={"display": "flex", "gap": "12px",
                            "flexWrap": "wrap", "marginBottom": "14px"},
                     children=[
        _metric("UNPOOLED", _pct(a["unpooled_weighted"]),
                "one model per channel",
                tone=(GREEN if best[0] == "Unpooled" else INK)),
        _metric("POOLED", _pct(a["pooled_weighted"]),
                f"{a['pooled_rows']} rows, {a['n_pooled_vars']} predictors",
                tone=(GREEN if best[0] == "Pooled" else INK)),
        _metric("HIERARCHICAL", _pct(a["hier_weighted"]),
                f"λ = {used_shrink:g}",
                tone=(GREEN if best[0] == "Hierarchical" else INK)),
        _metric("WINNER", best[0], "lowest holdout WMAPE", tone=NAVY),
    ])

    # ── per-channel comparison ──
    ct = cmp_tbl.reset_index()
    disp = ct[["channel", "volume", "unpooled_wmape", "pooled_wmape",
               "hier_wmape", "best_level"]].copy()
    disp["volume"] = disp["volume"].map(lambda v: fmt_km(v, 1))
    for c in ("unpooled_wmape", "pooled_wmape", "hier_wmape"):
        disp[c] = disp[c].map(lambda v: _pct(v))
    disp.columns = ["Channel", "Volume", "Unpooled", "Pooled", "Hierarchical",
                    "Best"]

    # ── national predictors ──
    nat = pooled["national"]
    sum_only = pooled["national_check"].get("sum_equal_only") or []
    nat_card = card(
        "National predictors — split before pooling",
        f"{len(nat)} of {len(pooled['specs'])} candidates carry identical "
        f"values in every channel",
        html.Div([
            html.Div(
                "Each is multiplied by its channel's share of product volume, "
                "so the per-channel pieces sum back to the national figure and "
                "one shared coefficient produces the right product-level "
                "contribution instead of one per channel. Conservation checked "
                f"to {pooled['national_check']['max_rel_error']:.0e}.",
                style={"fontSize": "12px", "color": INK3}),
            html.Div(", ".join(nat[:14]) + ("…" if len(nat) > 14 else ""),
                     style={"fontSize": "11.5px", "color": INK2,
                            "fontFamily": "IBM Plex Mono", "marginTop": "8px"}),
        ] + ([html.Div(
            f"⚠ {', '.join(sum_only)} — equal SUMS in every channel but "
            "different weekly values, so NOT national. The quick "
            "sum-by-channel screen flags these; the engine compares the weekly "
            "values instead and leaves them channel-specific.",
            style={"fontSize": "11.5px", "color": WARN, "marginTop": "8px"})]
            if sum_only else [])))

    # ── coefficient index heat table ──
    idx = hier["index_capped"].round(3)
    idx_rows = [{"Channel": ch, **{c: f"{idx.loc[ch, c]:.2f}"
                                  for c in idx.columns}}
                for ch in idx.index]
    stab = hier["stability"]
    unstable = int((~stab["stable"]).sum())

    idx_card = card(
        "Coefficient index by channel",
        "final coefficient = index × pooled coefficient (1.00 = inherit the "
        "product-level effect unchanged)",
        html.Div([
            dash_table.DataTable(
                data=idx_rows,
                columns=[{"name": ("Channel" if c == "Channel"
                                   else c[:26]), "id": c}
                         for c in ["Channel"] + list(idx.columns)],
                style_table={"overflowX": "auto"}, page_size=12, **TBL_BASE),
            html.Div(
                [f"Index averages to 1 by construction (checked to "
                 f"{hier['checks']['index_mean_max_dev']:.0e}); uncapped "
                 f"coefficients average back to the pooled ones "
                 f"({hier['checks']['uncapped_recovers_pooled']:.0e}). ",
                 html.Br(),
                 f"⚠ {unstable} of {len(stab)} predictors have |mean| below "
                 "their spread across channels, so their index is dividing by "
                 "a number close to zero and is not reliable. Those are the "
                 "rows the cap and λ exist to tame."
                 if unstable else ""],
                style={"fontSize": "11.5px", "color": INK3,
                       "marginTop": "8px"}),
        ]), pad=False)

    # ── λ curve ──
    lam_card = None
    if a.get("shrink_curve"):
        cur = pd.DataFrame(a["shrink_curve"])
        lf = go.Figure(go.Scatter(
            x=cur["shrink"], y=cur["holdout_wmape"], mode="lines+markers",
            line=dict(color=NAVY, width=2)))
        lf.add_vline(x=used_shrink, line=dict(color=GREEN, width=1.5,
                                              dash="dot"),
                     annotation_text=f"chosen λ = {used_shrink:g}",
                     annotation_font_size=10)
        lf = _fig_base(lf, 220)
        lf.update_layout(xaxis_title="λ  (0 = pooled → 1 = full index)",
                         yaxis_title="holdout WMAPE %")
        flat = (cur["holdout_wmape"].max() - cur["holdout_wmape"].min()) \
            / max(cur["holdout_wmape"].min(), 1e-9) < 0.05
        lam_card = card(
            "How much channel variation the data supports",
            "each λ rebuilt from training rows only and scored on the "
            "reserved tails",
            html.Div([
                dcc.Graph(figure=lf, config={"displayModeBar": False}),
                html.Div(
                    "The curve is flat — the channels do not measurably differ "
                    "on this product, and the pooled model is the honest "
                    "answer." if flat else
                    ("λ = 0 wins: every bit of channel-level variation makes "
                     "the holdout worse, so these channels are better served "
                     "by one shared coefficient."
                     if used_shrink == 0 else
                     f"λ = {used_shrink:g} beats both extremes — genuine "
                     "partial pooling, which is the case Arko described."),
                    style={"fontSize": "11.5px", "color": INK3}),
            ]))

    # ── pooled per-channel fit ──
    pf = ml.pooled_channel_fit(pooled).reset_index()
    pf_disp = pf.copy()
    pf_disp["actual"] = pf_disp["actual"].map(lambda v: fmt_km(v, 1))
    pf_disp["fitted"] = pf_disp["fitted"].map(lambda v: fmt_km(v, 1))
    pf_disp["r2"] = pf_disp["r2"].map(lambda v: f"{v:.3f}")
    for c in ("mape", "bias_pct"):
        pf_disp[c] = pf_disp[c].map(lambda v: _pct(v))
    pf_disp.columns = ["Channel", "Weeks", "R²", "MAPE", "Bias", "Actual",
                       "Fitted"]

    skipped = pooled.get("skipped") or []
    skip_note = (html.Div(
        "Not pooled: " + "; ".join(f"{s['channel']} ({s['reason']})"
                                   for s in skipped),
        style={"fontSize": "11.5px", "color": WARN, "marginTop": "8px"})
        if skipped else None)

    return html.Div([
        strip,
        html.Div(f"Ran in {time.time() - t0:.1f}s · window "
                 f"{pooled['window_start']:%Y-%m-%d} to "
                 f"{pooled['window_end']:%Y-%m-%d} · "
                 f"{len(pooled['channels'])} channels · all three levels "
                 f"scored on the SAME held-out last 13 weeks per channel, by "
                 f"WMAPE (error volume ÷ actual volume — plain MAPE would let "
                 f"the smallest channel decide the winner).",
                 style={"fontSize": "12px", "color": INK3,
                        "marginBottom": "12px"}),
        card("Per-channel comparison",
             "same holdout, same metric, three levels",
             html.Div([
                 dash_table.DataTable(
                     data=disp.to_dict("records"),
                     columns=[{"name": c, "id": c} for c in disp.columns],
                     sort_action="native", page_size=15,
                     style_data_conditional=[
                         {"if": {"filter_query": "{Best} = 'unpooled'",
                                 "column_id": "Unpooled"},
                          "backgroundColor": "#E0F2EC", "fontWeight": "600"},
                         {"if": {"filter_query": "{Best} = 'pooled'",
                                 "column_id": "Pooled"},
                          "backgroundColor": "#E0F2EC", "fontWeight": "600"},
                         {"if": {"filter_query": "{Best} = 'hierarchical'",
                                 "column_id": "Hierarchical"},
                          "backgroundColor": "#E0F2EC", "fontWeight": "600"}],
                     **TBL_BASE),
                 skip_note] if skip_note is not None else
                 dash_table.DataTable(
                     data=disp.to_dict("records"),
                     columns=[{"name": c, "id": c} for c in disp.columns],
                     sort_action="native", page_size=15, **TBL_BASE)),
             pad=False),
        html.Div(style={"marginTop": "14px"}, children=[lam_card])
        if lam_card is not None else None,
        html.Div(style={"marginTop": "14px"}, children=[nat_card]),
        html.Div(style={"marginTop": "14px"}, children=[idx_card]),
        html.Div(style={"marginTop": "14px"}, children=[
            card("Pooled model — fit by channel",
                 "how well ONE set of coefficients serves each channel",
                 dash_table.DataTable(
                     data=pf_disp.to_dict("records"),
                     columns=[{"name": c, "id": c} for c in pf_disp.columns],
                     sort_action="native", page_size=15, **TBL_BASE),
                 pad=False)]),
    ])


def _validate_wiring():
    """Fail loudly at startup if any callback references a component id
    missing from the layout — the browser otherwise disables the callback
    silently when debug=False (bitten once by a falsy empty component)."""
    import collections
    import re as _re
    ids = []

    def walk(c):
        if getattr(c, "id", None) is not None:
            ids.append(c.id)
        ch = getattr(c, "children", None)
        if ch is None:
            return
        for x in (ch if isinstance(ch, (list, tuple)) else [ch]):
            if hasattr(x, "to_plotly_json"):
                walk(x)

    walk(app.layout)
    dupes = [i for i, n in collections.Counter(ids).items() if n > 1]
    missing = set()
    for spec in app._callback_list:
        for dep in spec["inputs"] + spec.get("state", []):
            if dep["id"] not in ids:
                missing.add(dep["id"])
        for oid in _re.findall(r"([\w-]+)\.[\w]+", spec["output"]):
            if oid not in ids:
                missing.add(oid)
    if dupes or missing:
        raise RuntimeError(
            f"Callback wiring broken - duplicate ids: {dupes or 'none'}; "
            f"ids referenced but missing from layout: {sorted(missing) or 'none'}")
    print(f"wiring OK: {len(ids)} components, "
          f"{len(app._callback_list)} callbacks", flush=True)


if __name__ == "__main__":
    _validate_wiring()
    app.run(debug=False, host="127.0.0.1", port=8050)
